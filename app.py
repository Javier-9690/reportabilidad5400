
import csv
import hashlib
import json
import os
import re
import threading
import tempfile
import unicodedata
from collections import defaultdict
from difflib import SequenceMatcher
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO

import pandas as pd
import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell
from dotenv import load_dotenv
from flask import Flask, Response, flash, jsonify, redirect, render_template, request, send_file, stream_with_context, url_for
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, func, or_

load_dotenv()
db = SQLAlchemy()

DATE_RE = re.compile(r"(\d{1,2})[-_/\.](\d{1,2})[-_/\.](\d{2,4})")
DEFAULT_GERENCIAS = [
    "BHP Projects", "Cátodos", "Concentradora", "Development&Strategic", "HSE",
    "Integrated Operations", "Mine Operations", "NPI&CHO", "PT&E", "VPP", "VPP Growth",
    "VP Sustaining", "Infraestructura & Servicios",
]


def now_utc():
    return datetime.utcnow()


class UploadedFile(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    file_type = db.Column(db.String(20), nullable=False)  # curva | censo
    size_bytes = db.Column(db.Integer, nullable=False, default=0)
    sha256 = db.Column(db.String(64), nullable=False, index=True)
    content = db.Column(db.LargeBinary, nullable=False)
    uploaded_at = db.Column(db.DateTime, nullable=False, default=now_utc)


class CurvaVersion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    file_id = db.Column(db.Integer, db.ForeignKey("uploaded_file.id"), nullable=False)
    name = db.Column(db.String(255), nullable=False)
    sheet_name = db.Column(db.String(120), nullable=False, default="Fcst_5400")
    is_active = db.Column(db.Boolean, nullable=False, default=True)
    total_items = db.Column(db.Integer, nullable=False, default=0)
    total_daily_values = db.Column(db.Integer, nullable=False, default=0)
    uploaded_at = db.Column(db.DateTime, nullable=False, default=now_utc)
    file = db.relationship("UploadedFile", backref="curvas")


class CurvaItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    curva_version_id = db.Column(db.Integer, db.ForeignKey("curva_version.id"), nullable=False, index=True)
    solicitud_id = db.Column(db.String(80), nullable=False, index=True)
    gerencia = db.Column(db.String(180), nullable=False, index=True)
    area = db.Column(db.String(180), default="")
    empresa = db.Column(db.String(180), default="")
    turno = db.Column(db.String(80), default="")
    tipo_contrato = db.Column(db.String(120), default="")
    formato = db.Column(db.String(120), default="")
    camp = db.Column(db.String(120), default="")
    version = db.relationship("CurvaVersion", backref="items")
    __table_args__ = (db.UniqueConstraint("curva_version_id", "solicitud_id", name="uq_curva_solicitud"),)


class CurvaDailyValue(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    curva_item_id = db.Column(db.Integer, db.ForeignKey("curva_item.id"), nullable=False, index=True)
    fecha = db.Column(db.Date, nullable=False, index=True)
    dotacion_planificada = db.Column(db.Float, nullable=False, default=0)
    item = db.relationship("CurvaItem", backref="daily_values")


class Censo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    file_id = db.Column(db.Integer, db.ForeignKey("uploaded_file.id"), nullable=False)
    fecha_censo = db.Column(db.Date, nullable=False, index=True)
    sheet_name = db.Column(db.String(120), default="")
    total_records = db.Column(db.Integer, nullable=False, default=0)
    total_occupied = db.Column(db.Float, nullable=False, default=0)
    matched_count = db.Column(db.Integer, nullable=False, default=0)
    unmatched_count = db.Column(db.Integer, nullable=False, default=0)
    imported_at = db.Column(db.DateTime, nullable=False, default=now_utc)
    file = db.relationship("UploadedFile", backref="censos")

    @property
    def fecha_label(self):
        return self.fecha_censo.strftime("%d/%m/%Y")


class CensoRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    censo_id = db.Column(db.Integer, db.ForeignKey("censo.id"), nullable=False, index=True)
    curva_item_id = db.Column(db.Integer, db.ForeignKey("curva_item.id"), nullable=True, index=True)
    solicitud_id = db.Column(db.String(80), index=True)
    modulo = db.Column(db.String(120), default="")
    lugar = db.Column(db.String(120), default="")
    habitacion = db.Column(db.String(120), default="")
    empresa = db.Column(db.String(180), default="")
    cama = db.Column(db.String(80), default="")
    dia = db.Column(db.String(80), default="")
    camas_ocupadas = db.Column(db.Float, nullable=False, default=1)
    turno = db.Column(db.String(80), default="")
    gerencia_censo = db.Column(db.String(180), default="")
    area = db.Column(db.String(180), default="")
    rut = db.Column(db.String(80), default="")
    estado = db.Column(db.String(120), default="")
    censo = db.relationship("Censo", backref="records")
    curva_item = db.relationship("CurvaItem", backref="censo_records")


class ExportJob(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    job_type = db.Column(db.String(80), nullable=False, default="dotacion_gerencia")
    status = db.Column(db.String(30), nullable=False, default="pending", index=True)  # pending | running | completed | failed
    message = db.Column(db.Text, default="")
    params_json = db.Column(db.Text, default="{}")
    filename = db.Column(db.String(255), default="")
    content_type = db.Column(db.String(160), default="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    content = db.Column(db.LargeBinary, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=now_utc)
    started_at = db.Column(db.DateTime, nullable=True)
    completed_at = db.Column(db.DateTime, nullable=True)


def get_database_url(app):
    """
    Resolve la URL de base de datos.

    En Render debe venir desde la variable DATABASE_URL de PostgreSQL.
    Para desarrollo local, crea automáticamente la carpeta instance/ y usa SQLite.
    """
    db_url = os.getenv("DATABASE_URL", "").strip()

    if db_url:
        # Render/PostgreSQL antiguos pueden entregar postgres://, pero SQLAlchemy usa postgresql://
        if db_url.startswith("postgres://"):
            db_url = db_url.replace("postgres://", "postgresql://", 1)
        return db_url

    # Fallback solo para desarrollo local.
    # Esto evita el error: sqlite3.OperationalError: unable to open database file
    # cuando la carpeta instance/ no existe.
    os.makedirs(app.instance_path, exist_ok=True)
    return "sqlite:///" + os.path.join(app.instance_path, "dotacion_reportes.db")


def create_app():
    app = Flask(__name__, instance_relative_config=True)
    app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-change-me")
    app.config["SQLALCHEMY_DATABASE_URI"] = get_database_url(app)
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {"pool_pre_ping": True}
    app.config["MAX_CONTENT_LENGTH"] = int(os.getenv("MAX_CONTENT_LENGTH_MB", "80")) * 1024 * 1024
    db.init_app(app)
    with app.app_context():
        db.create_all()
    register_routes(app)
    return app


def strip_accents(text):
    text = unicodedata.normalize("NFKD", str(text))
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def norm(value):
    if value is None or pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", strip_accents(value).lower().replace("_", " ")).strip()


def clean(value):
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def norm_id(value):
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text.upper()


def as_number(value):
    if value is None or pd.isna(value):
        return 0.0
    if isinstance(value, str):
        text = value.strip().replace(",", ".")
        if norm(text) in {"si", "sí", "x", "ok", "ocupado", "ocupada"}:
            return 1.0
        try:
            return float(text)
        except ValueError:
            return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def parse_date(value):
    if value is None or value == "" or pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and 20000 <= float(value) <= 60000:
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    parsed = pd.to_datetime(str(value), errors="coerce", dayfirst=True)
    return None if pd.isna(parsed) else parsed.date()


def date_from_text(text):
    m = DATE_RE.search(str(text or ""))
    if not m:
        return None
    d, mo, y = m.groups()
    y = int(y) + (2000 if int(y) < 100 else 0)
    try:
        return date(y, int(mo), int(d))
    except ValueError:
        return None


def find_header(raw, required):
    required = [norm(x) for x in required]
    for i in range(min(len(raw), 50)):
        values = {norm(v) for v in raw.iloc[i].tolist() if clean(v)}
        if sum(1 for x in required if x in values) >= min(2, len(required)):
            return i
    return 0


def colmap(columns):
    aliases = {
        "id": ["id de la solicitud", "id solicitud", "id", "solicitud"],
        "gerencia": ["gerencia general", "gerencia"],
        "area": ["area", "área"], "empresa": ["empresa"], "turno": ["turno"],
        "tipo_contrato": ["tipo contrato", "tipo de contrato"], "formato": ["formato"], "camp": ["camp", "campamento"],
        "modulo": ["modulo", "módulo"], "lugar": ["lugar"], "habitacion": ["habitacion", "habitación"],
        "cama": ["cama"], "inicio": ["inicio"], "termino": ["termino", "término"],
        "dia": ["dia", "día", "fecha", "fecha censo"],
        "ocupadas": ["camas ocupadas", "camas ocupdas", "camas ocupada", "ocupadas", "ocupada", "ocupacion", "ocupación"],
        "co_mel": ["co mel", "comel", "co_mel", "co", "centro costo", "centro de costo"],
        "sexo": ["sexo", "genero", "género"],
        "jornada": ["jornada", "rol", "sistema turno", "sistema de turno"],
        "pab": ["pab", "pabellon", "pabellón"],
        "rut": ["rut", "run"], "estado": ["estado", "status"],
    }
    available = {norm(c): c for c in columns}
    out = {}
    for key, opts in aliases.items():
        for opt in opts:
            if norm(opt) in available:
                out[key] = available[norm(opt)]
                break
    return out


def save_upload(file, file_type):
    content = file.read()
    f = UploadedFile(filename=file.filename, file_type=file_type, size_bytes=len(content), sha256=hashlib.sha256(content).hexdigest(), content=content)
    db.session.add(f); db.session.flush()
    return f, content



def xlsx_sheet_paths(content):
    """
    Lee los nombres de hojas de un .xlsx directamente desde el ZIP interno.
    No usa openpyxl para evitar cargar estilos pesados en Render.
    """
    import zipfile
    import xml.etree.ElementTree as ET

    ns_main = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    ns_rel = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}

    with zipfile.ZipFile(BytesIO(content)) as zf:
        workbook = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

        rel_map = {}
        for rel in rels.findall("r:Relationship", ns_rel):
            target = rel.attrib.get("Target", "")
            if not target.startswith("/"):
                target = "xl/" + target
            else:
                target = target.lstrip("/")
            rel_map[rel.attrib.get("Id")] = target

        sheets = []
        for sheet in workbook.findall("a:sheets/a:sheet", ns_main):
            name = sheet.attrib.get("name")
            rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            path = rel_map.get(rel_id)
            if name and path:
                sheets.append((name, path))
        return sheets


def get_sheet_names(content, filename):
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return [name for name, _ in xlsx_sheet_paths(content)]
    if name.endswith(".xlsb"):
        return pd.ExcelFile(BytesIO(content), engine="pyxlsb").sheet_names
    return pd.ExcelFile(BytesIO(content)).sheet_names


def column_index_from_cell_ref(cell_ref):
    letters = "".join(ch for ch in str(cell_ref or "") if ch.isalpha()).upper()
    if not letters:
        return 0
    idx = 0
    for ch in letters:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def read_xlsx_shared_strings(zf):
    import xml.etree.ElementTree as ET
    try:
        data = zf.read("xl/sharedStrings.xml")
    except KeyError:
        return []

    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    root = ET.fromstring(data)
    values = []
    for si in root.findall("a:si", ns):
        texts = [node.text or "" for node in si.findall(".//a:t", ns)]
        values.append("".join(texts))
    return values


def parse_xlsx_cell(cell, shared_strings):
    import xml.etree.ElementTree as ET
    cell_type = cell.attrib.get("t")
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

    if cell_type == "inlineStr":
        texts = [node.text or "" for node in cell.findall(f".//{ns}t")]
        return "".join(texts)

    value_node = cell.find(f"{ns}v")
    if value_node is None or value_node.text is None:
        return ""

    raw = value_node.text
    if cell_type == "s":
        try:
            return shared_strings[int(raw)]
        except Exception:
            return ""
    if cell_type == "b":
        return raw == "1"

    try:
        num = float(raw)
        return int(num) if num.is_integer() else num
    except Exception:
        return raw


def read_xlsx_sheet_df(content, sheet_name, header=None, nrows=None):
    """
    Convierte una hoja .xlsx a DataFrame leyendo XML crudo.
    Esto evita openpyxl/pandas para la curva, que en Render puede morir por estilos.
    """
    import zipfile
    import xml.etree.ElementTree as ET

    sheet_map = dict(xlsx_sheet_paths(content))
    if sheet_name not in sheet_map:
        raise ValueError(f"La hoja {sheet_name} no existe en el archivo.")

    rows = []
    max_cols = 0
    with zipfile.ZipFile(BytesIO(content)) as zf:
        shared_strings = read_xlsx_shared_strings(zf)
        sheet_path = sheet_map[sheet_name]
        with zf.open(sheet_path) as fh:
            context = ET.iterparse(fh, events=("end",))
            ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
            for _, elem in context:
                if elem.tag != f"{ns}row":
                    continue
                row_values = []
                for cell in elem.findall(f"{ns}c"):
                    col_idx = column_index_from_cell_ref(cell.attrib.get("r"))
                    while len(row_values) <= col_idx:
                        row_values.append("")
                    row_values[col_idx] = parse_xlsx_cell(cell, shared_strings)
                max_cols = max(max_cols, len(row_values))
                rows.append(row_values)
                elem.clear()
                if nrows is not None and len(rows) >= nrows:
                    break

    if not rows:
        return pd.DataFrame()

    for row in rows:
        if len(row) < max_cols:
            row.extend([""] * (max_cols - len(row)))

    if header is None:
        return pd.DataFrame(rows)

    header = int(header)
    columns = rows[header]
    # Asegura nombres únicos para columnas vacías o repetidas.
    cleaned_columns = []
    seen = {}
    for i, col in enumerate(columns):
        # Mantiene los encabezados numéricos de fecha como número Excel.
        # Así parse_date(46083) los convierte correctamente a fecha real.
        if col is None or col == "" or pd.isna(col):
            name = f"col_{i + 1}"
        else:
            name = col

        key = str(name)
        if key in seen:
            seen[key] += 1
            name = f"{key}_{seen[key]}"
        else:
            seen[key] = 0
        cleaned_columns.append(name)

    data_rows = rows[header + 1:]
    return pd.DataFrame(data_rows, columns=cleaned_columns)


def read_excel_df(content, filename, sheet_name, header=None, nrows=None):
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return read_xlsx_sheet_df(content, sheet_name=sheet_name, header=header, nrows=nrows)

    engine = "pyxlsb" if name.endswith(".xlsb") else None
    kwargs = {
        "io": BytesIO(content),
        "sheet_name": sheet_name,
        "header": header,
    }
    if engine:
        kwargs["engine"] = engine
    if nrows is not None:
        kwargs["nrows"] = nrows
    return pd.read_excel(**kwargs)


def import_curva(file, sheet_name="Fcst_5400", version_name=None):
    uploaded, content = save_upload(file, "curva")
    sheet_names = get_sheet_names(content, uploaded.filename)
    sheet = sheet_name if sheet_name in sheet_names else next((s for s in sheet_names if "5400" in s), sheet_names[0])
    raw = read_excel_df(content, uploaded.filename, sheet_name=sheet, header=None, nrows=80)
    header = find_header(raw, ["ID de la solicitud", "Gerencia General"])
    df = read_excel_df(content, uploaded.filename, sheet_name=sheet, header=header).dropna(how="all")
    cm = colmap(df.columns)
    if "id" not in cm or "gerencia" not in cm:
        raise ValueError("No se encontraron columnas ID de la solicitud y Gerencia General en la curva.")
    date_cols = [(c, parse_date(c)) for c in df.columns]
    date_cols = [(c, d) for c, d in date_cols if d]
    CurvaVersion.query.update({CurvaVersion.is_active: False})
    version = CurvaVersion(file_id=uploaded.id, name=version_name or uploaded.filename, sheet_name=sheet, is_active=True)
    db.session.add(version); db.session.flush()
    seen, n_items, n_values = set(), 0, 0
    for _, row in df.iterrows():
        sid = norm_id(row.get(cm["id"]))
        if not sid or sid in seen:
            continue
        seen.add(sid)
        item = CurvaItem(
            curva_version_id=version.id, solicitud_id=sid, gerencia=clean(row.get(cm["gerencia"])) or "SIN GERENCIA",
            area=clean(row.get(cm.get("area"))) if cm.get("area") else "",
            empresa=clean(row.get(cm.get("empresa"))) if cm.get("empresa") else "",
            turno=clean(row.get(cm.get("turno"))) if cm.get("turno") else "",
            tipo_contrato=clean(row.get(cm.get("tipo_contrato"))) if cm.get("tipo_contrato") else "",
            formato=clean(row.get(cm.get("formato"))) if cm.get("formato") else "",
            camp=clean(row.get(cm.get("camp"))) if cm.get("camp") else "",
        )
        db.session.add(item); db.session.flush(); n_items += 1
        vals = [CurvaDailyValue(curva_item_id=item.id, fecha=d, dotacion_planificada=as_number(row.get(c))) for c, d in date_cols]
        if vals:
            db.session.bulk_save_objects(vals); n_values += len(vals)
    version.total_items, version.total_daily_values = n_items, n_values
    db.session.commit()
    return version


def read_censo(content, filename):
    sheet_names = get_sheet_names(content, filename)
    sheet = next((s for s in sheet_names if date_from_text(s)), sheet_names[0])
    detected_date = date_from_text(sheet) or date_from_text(filename)
    raw = read_excel_df(content, filename, sheet_name=sheet, header=None, nrows=80)
    header = find_header(raw, ["Id", "Camas Ocupadas"])
    df = read_excel_df(content, filename, sheet_name=sheet, header=header).dropna(how="all")
    return df, sheet, detected_date


def lookup_active_curve_items():
    """Retorna la curva activa y un mapa ID -> CurvaItem.id para cruces rápidos."""
    active = CurvaVersion.query.filter_by(is_active=True).order_by(CurvaVersion.uploaded_at.desc()).first()
    item_map = {}
    if active:
        item_map = {
            i.solicitud_id: i.id
            for i in CurvaItem.query.filter_by(curva_version_id=active.id).all()
        }
    return active, item_map


def occupied_value(row, cm):
    """Normaliza Camas Ocupdas/Ocupadas a 1 o 0 para reservas del censo."""
    if cm.get("ocupadas"):
        return 1.0 if as_number(row.get(cm.get("ocupadas"))) > 0 else 0.0
    return 0.0


def rebuild_censo_from_dataframe(censo, df, preserved_matches=None):
    """
    Reconstruye los registros de un censo desde su DataFrame original.

    Regla de negocio:
    - Reservas = todas las filas del censo que tienen ID.
    - Ocupación = reservas con Camas Ocupdas/Ocupadas > 0.
    - Sin match/Cruzados se calculan sobre todas las reservas, no solo ocupados.
    - Si ya existía una corrección manual para un ID, se conserva y se aplica a
      todas las reservas con ese mismo ID, incluyendo Camas Ocupdas = 0.
    """
    cm = colmap(df.columns)
    if "id" not in cm:
        raise ValueError("No se encontró la columna Id en el censo.")

    _, item_map = lookup_active_curve_items()
    preserved_matches = preserved_matches or {}

    records = []
    matched = 0
    unmatched = 0
    occupied_total = 0

    for _, row in df.iterrows():
        sid = norm_id(row.get(cm["id"]))
        if not sid:
            continue

        occupied = occupied_value(row, cm)
        occupied_total += occupied

        item_id = preserved_matches.get(sid) or item_map.get(sid)
        if item_id:
            matched += 1
        else:
            unmatched += 1

        records.append(CensoRecord(
            censo_id=censo.id,
            curva_item_id=item_id,
            solicitud_id=sid,
            modulo=clean(row.get(cm.get("modulo"))) if cm.get("modulo") else "",
            lugar=clean(row.get(cm.get("lugar"))) if cm.get("lugar") else "",
            habitacion=clean(row.get(cm.get("habitacion"))) if cm.get("habitacion") else "",
            empresa=clean(row.get(cm.get("empresa"))) if cm.get("empresa") else "",
            cama=clean(row.get(cm.get("cama"))) if cm.get("cama") else "",
            dia=clean(row.get(cm.get("dia"))) if cm.get("dia") else censo.fecha_censo.strftime("%d/%m/%Y"),
            camas_ocupadas=occupied,
            turno=clean(row.get(cm.get("turno"))) if cm.get("turno") else "",
            gerencia_censo=clean(row.get(cm.get("gerencia"))) if cm.get("gerencia") else "",
            area=clean(row.get(cm.get("area"))) if cm.get("area") else "",
            rut=clean(row.get(cm.get("rut"))) if cm.get("rut") else "",
            estado=clean(row.get(cm.get("estado"))) if cm.get("estado") else "",
        ))

    CensoRecord.query.filter_by(censo_id=censo.id).delete(synchronize_session=False)
    if records:
        db.session.bulk_save_objects(records)

    censo.total_records = len(records)  # Reservas
    censo.total_occupied = occupied_total  # Ocupación
    censo.matched_count = matched
    censo.unmatched_count = unmatched
    return censo


def import_censo(file):
    uploaded, content = save_upload(file, "censo")
    df, sheet, fecha = read_censo(content, uploaded.filename)
    cm = colmap(df.columns)
    if "id" not in cm:
        raise ValueError("No se encontró la columna Id en el censo.")
    if not fecha and "dia" in cm:
        dates = [parse_date(x) for x in df[cm["dia"]].dropna().tolist()]
        dates = [x for x in dates if x]
        fecha = max(set(dates), key=dates.count) if dates else None
    if not fecha:
        raise ValueError("No se pudo detectar la fecha del censo.")

    censo = Censo(file_id=uploaded.id, fecha_censo=fecha, sheet_name=sheet)
    db.session.add(censo)
    db.session.flush()
    rebuild_censo_from_dataframe(censo, df)
    db.session.commit()
    return censo


def date_span(start, end):
    while start <= end:
        yield start
        start += timedelta(days=1)


def report_data(start=None, end=None, curve_id=None):
    if not start or not end:
        minmax = db.session.query(func.min(Censo.fecha_censo), func.max(Censo.fecha_censo)).one()
        start = start or minmax[0]; end = end or minmax[1]
    if not start or not end:
        return {"dates": [], "date_labels": [], "rows": [], "totals_by_date": [], "grand_total": 0, "conclusions": ["No hay censos importados."]}
    dates = list(date_span(start, end))
    curve = CurvaVersion.query.get(curve_id) if curve_id else CurvaVersion.query.filter_by(is_active=True).order_by(CurvaVersion.uploaded_at.desc()).first()
    actual, planned, gerencias = defaultdict(lambda: defaultdict(float)), defaultdict(lambda: defaultdict(float)), set()
    rows_actual = db.session.query(Censo.fecha_censo, func.coalesce(CurvaItem.gerencia, "SIN MATCH EN CURVA"), func.sum(CensoRecord.camas_ocupadas)).join(CensoRecord, CensoRecord.censo_id==Censo.id).outerjoin(CurvaItem, CurvaItem.id==CensoRecord.curva_item_id).filter(Censo.fecha_censo>=start, Censo.fecha_censo<=end).group_by(Censo.fecha_censo, func.coalesce(CurvaItem.gerencia, "SIN MATCH EN CURVA")).all()
    for d, g, val in rows_actual:
        actual[g][d] += float(val or 0); gerencias.add(g)
    if curve:
        rows_plan = db.session.query(CurvaItem.gerencia, CurvaDailyValue.fecha, func.sum(CurvaDailyValue.dotacion_planificada)).join(CurvaDailyValue, CurvaDailyValue.curva_item_id==CurvaItem.id).filter(CurvaItem.curva_version_id==curve.id, CurvaDailyValue.fecha>=start, CurvaDailyValue.fecha<=end).group_by(CurvaItem.gerencia, CurvaDailyValue.fecha).all()
        for g, d, val in rows_plan:
            planned[g][d] += float(val or 0); gerencias.add(g)
    order = [g for g in DEFAULT_GERENCIAS if g in gerencias] + sorted([g for g in gerencias if g not in DEFAULT_GERENCIAS and g != "SIN MATCH EN CURVA"])
    if "SIN MATCH EN CURVA" in gerencias: order.append("SIN MATCH EN CURVA")
    totals = [0.0 for _ in dates]; planned_totals = [0.0 for _ in dates]; out_rows=[]
    for g in order:
        vals=[actual[g].get(d,0.0) for d in dates]; plans=[planned[g].get(d,0.0) for d in dates]
        for i,v in enumerate(vals): totals[i]+=v
        for i,v in enumerate(plans): planned_totals[i]+=v
        out_rows.append({"gerencia":g,"values":vals,"planned_values":plans,"total":sum(vals),"planned_total":sum(plans),"difference":sum(vals)-sum(plans)})
    grand, plan_grand = sum(totals), sum(planned_totals)
    conclusions = [f"Dotación real acumulada: {grand:.0f}. Planificación acumulada: {plan_grand:.0f}."]
    if plan_grand: conclusions.append(f"Cumplimiento global contra curva: {(grand/plan_grand*100):.1f}%.")
    if dates: conclusions.append(f"Día con mayor dotación real: {dates[max(range(len(totals)), key=lambda i: totals[i])].strftime('%d/%m/%Y')} ({max(totals):.0f}).")
    if out_rows:
        top=max(out_rows, key=lambda r:r['total']); conclusions.append(f"Gerencia con mayor dotación real: {top['gerencia']} ({top['total']:.0f}).")
    no_match=next((r for r in out_rows if r['gerencia']=='SIN MATCH EN CURVA'), None)
    if no_match and no_match['total']>0: conclusions.append(f"Existen {no_match['total']:.0f} registros sin match de ID en la curva activa.")
    return {"start_date":start.isoformat(),"end_date":end.isoformat(),"dates":[d.isoformat() for d in dates],"date_labels":[d.strftime('%d/%m/%y') for d in dates],"rows":out_rows,"totals_by_date":totals,"planned_totals_by_date":planned_totals,"grand_total":grand,"planned_grand_total":plan_grand,"conclusions":conclusions,"curve":{"id":curve.id,"name":curve.name} if curve else None}


def report_xlsx(data):
    """Genera el Excel del reporte y agrega detalle completo de todos los censos importados."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Dotacion Gerencia'

    dates = data.get('date_labels', [])
    rows = data.get('rows', [])
    total_cols = len(dates) + 5  # Gerencia + fechas + Total Real + Total Plan + Diferencia + Cumplimiento

    colors = {
        'red': 'B42318',
        'red_dark': '7A1712',
        'pink': 'F8C9CD',
        'pink_soft': 'FFF0F0',
        'yellow': 'FFF200',
        'white': 'FFFFFF',
        'black': '111111',
        'gray': 'F2F4F7',
        'blue': 'E8F1FF',
        'green': 'D1FADF',
        'green_text': '027A48',
        'danger': 'FEE4E2',
        'danger_text': 'B42318',
        'warning': 'FFF3CD',
        'orange_text': 'B54708',
        'border': '7A1712',
    }

    thin_border = Border(
        left=Side(style='thin', color=colors['border']),
        right=Side(style='thin', color=colors['border']),
        top=Side(style='thin', color=colors['border']),
        bottom=Side(style='thin', color=colors['border']),
    )
    thick_top = Border(
        left=Side(style='thin', color=colors['border']),
        right=Side(style='thin', color=colors['border']),
        top=Side(style='medium', color=colors['border']),
        bottom=Side(style='thin', color=colors['border']),
    )

    def fill(color):
        return PatternFill('solid', fgColor=color)

    def safe_num(value):
        return int(round(float(value or 0)))

    def safe_pct(real, plan):
        real = float(real or 0)
        plan = float(plan or 0)
        if not plan:
            return 'Sin plan' if real else '0%'
        return f'{(real / plan) * 100:.1f}%'

    def apply_header(cell, rotate=False):
        cell.fill = fill(colors['red'])
        cell.font = Font(color=colors['white'], bold=True, size=8)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90 if rotate else 0, wrap_text=True)

    ws.sheet_view.showGridLines = False

    # Título principal
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title = ws.cell(1, 1, 'RESUMEN DE DOTACIÓN POR GERENCIA')
    title.fill = fill(colors['red_dark'])
    title.font = Font(color=colors['white'], bold=True, size=16)
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Metadata
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    curve_name = (data.get('curve') or {}).get('name', 'Sin curva seleccionada')
    meta = ws.cell(2, 1, f"Período: {data.get('start_date', '')} al {data.get('end_date', '')} | Curva: {curve_name} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    meta.fill = fill(colors['pink_soft'])
    meta.font = Font(color=colors['black'], bold=True, size=9)
    meta.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    # KPIs
    real_total = float(data.get('grand_total') or 0)
    plan_total = float(data.get('planned_grand_total') or 0)
    diff_total = real_total - plan_total
    kpis = [
        ('REAL ACUMULADO', safe_num(real_total), colors['pink']),
        ('PLAN ACUMULADO', safe_num(plan_total), colors['blue']),
        ('DIFERENCIA', safe_num(diff_total), colors['green'] if diff_total >= 0 else colors['danger']),
        ('CUMPLIMIENTO', safe_pct(real_total, plan_total), colors['gray']),
    ]
    start_kpi_col = 1
    for label, value, color in kpis:
        ws.merge_cells(start_row=4, start_column=start_kpi_col, end_row=4, end_column=start_kpi_col + 1)
        cell = ws.cell(4, start_kpi_col, label)
        cell.fill = fill(colors['red'])
        cell.font = Font(color=colors['white'], bold=True, size=8)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.cell(4, start_kpi_col + 1).border = thin_border

        ws.merge_cells(start_row=5, start_column=start_kpi_col, end_row=5, end_column=start_kpi_col + 1)
        cell = ws.cell(5, start_kpi_col, value)
        cell.fill = fill(color)
        cell.font = Font(color=colors['black'], bold=True, size=12)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.cell(5, start_kpi_col + 1).border = thin_border
        start_kpi_col += 2

    # Tabla principal
    table_header_row = 7
    headers = ['GERENCIAS'] + dates + ['TOTAL REAL', 'TOTAL PLAN', 'DIF.', 'CUMP.']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(table_header_row, col_idx, header)
        apply_header(cell, rotate=1 < col_idx <= len(dates) + 1)
    ws.row_dimensions[table_header_row].height = 82

    for row_idx, item in enumerate(rows, table_header_row + 1):
        is_unmatched = item.get('gerencia') == 'SIN MATCH EN CURVA'
        ws.cell(row_idx, 1, item.get('gerencia', ''))
        for col_idx, value in enumerate(item.get('values', []), 2):
            ws.cell(row_idx, col_idx, safe_num(value))

        total_real_col = len(dates) + 2
        total_plan_col = len(dates) + 3
        diff_col = len(dates) + 4
        pct_col = len(dates) + 5
        real = float(item.get('total') or 0)
        plan = float(item.get('planned_total') or 0)
        diff = real - plan
        ws.cell(row_idx, total_real_col, safe_num(real))
        ws.cell(row_idx, total_plan_col, safe_num(plan))
        ws.cell(row_idx, diff_col, safe_num(diff))
        ws.cell(row_idx, pct_col, safe_pct(real, plan))

        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'center', vertical='center')
            cell.font = Font(size=8, bold=col_idx in (1, total_real_col, total_plan_col, diff_col, pct_col))
            if col_idx == 1:
                cell.fill = fill(colors['warning'] if is_unmatched else colors['white'])
                if is_unmatched:
                    cell.font = Font(size=8, bold=True, color=colors['orange_text'])
            elif 2 <= col_idx <= len(dates) + 1:
                cell.fill = fill(colors['warning'] if is_unmatched else (colors['white'] if safe_num(cell.value) == 0 else colors['pink']))
            elif col_idx == total_real_col:
                cell.fill = fill(colors['yellow'] if real else colors['pink'])
            elif col_idx == total_plan_col:
                cell.fill = fill(colors['blue'])
            elif col_idx == diff_col:
                cell.fill = fill(colors['green'] if diff >= 0 else colors['danger'])
                cell.font = Font(size=8, bold=True, color=colors['green_text'] if diff >= 0 else colors['danger_text'])
            elif col_idx == pct_col:
                cell.fill = fill(colors['gray'])

    total_row = table_header_row + len(rows) + 1
    ws.cell(total_row, 1, 'TOTAL')
    for col_idx, value in enumerate(data.get('totals_by_date', []), 2):
        ws.cell(total_row, col_idx, safe_num(value))
    ws.cell(total_row, len(dates) + 2, safe_num(real_total))
    ws.cell(total_row, len(dates) + 3, safe_num(plan_total))
    ws.cell(total_row, len(dates) + 4, safe_num(diff_total))
    ws.cell(total_row, len(dates) + 5, safe_pct(real_total, plan_total))

    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(total_row, col_idx)
        cell.border = thick_top
        cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'center', vertical='center')
        cell.font = Font(size=8, bold=True, color=colors['white'])
        cell.fill = fill(colors['red'])
        if col_idx == len(dates) + 2:
            cell.fill = fill(colors['yellow'])
            cell.font = Font(size=8, bold=True, color=colors['black'])
        elif col_idx == len(dates) + 3:
            cell.fill = fill(colors['blue'])
            cell.font = Font(size=8, bold=True, color=colors['black'])
        elif col_idx == len(dates) + 4:
            cell.fill = fill(colors['green'] if diff_total >= 0 else colors['danger'])
            cell.font = Font(size=8, bold=True, color=colors['green_text'] if diff_total >= 0 else colors['danger_text'])
        elif col_idx == len(dates) + 5:
            cell.fill = fill(colors['gray'])
            cell.font = Font(size=8, bold=True, color=colors['black'])

    conclusion_row = total_row + 3
    ws.cell(conclusion_row, 1, 'CONCLUSIONES')
    ws.cell(conclusion_row, 1).font = Font(bold=True, color=colors['red'], size=11)
    for idx, text in enumerate(data.get('conclusions', []), conclusion_row + 1):
        ws.cell(idx, 1, '• ' + str(text))
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=min(total_cols, 10))
        ws.cell(idx, 1).alignment = Alignment(wrap_text=True, vertical='top')

    ws.column_dimensions['A'].width = 32
    for col_idx in range(2, len(dates) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 5
    for col_idx in range(len(dates) + 2, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13
    ws.freeze_panes = 'B8'
    ws.auto_filter.ref = f'A{table_header_row}:{get_column_letter(total_cols)}{total_row}'

    # Pestaña con detalle completo de todos los censos importados previamente
    detail = wb.create_sheet('Detalle Censos')
    detail.sheet_view.showGridLines = False
    detail_headers = [
        'Censo ID', 'Fecha Censo', 'Hoja Censo', 'Importado Censo',
        'Total Registros Censo', 'Total Ocupado Censo', 'Cruzados Censo', 'Sin Match Censo',
        'Archivo ID', 'Archivo Nombre', 'Archivo Tipo', 'Archivo Tamaño Bytes', 'Archivo SHA256', 'Archivo Subido',
        'Registro ID', 'Solicitud ID', 'Módulo', 'Lugar', 'Habitación', 'Empresa Censo', 'Cama', 'Día',
        'Camas Ocupadas', 'Turno Censo', 'Gerencia Censo', 'Área Censo', 'RUT', 'Estado',
        'Curva Item ID', 'Curva Versión ID', 'Curva Versión', 'Gerencia Curva', 'Área Curva', 'Empresa Curva',
        'Turno Curva', 'Tipo Contrato Curva', 'Formato Curva', 'Camp Curva', 'Estado Match'
    ]
    for col_idx, header in enumerate(detail_headers, 1):
        cell = detail.cell(1, col_idx, header)
        cell.fill = fill(colors['red'])
        cell.font = Font(color=colors['white'], bold=True, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    detail.row_dimensions[1].height = 32

    # IMPORTANTE:
    # No consultamos el modelo UploadedFile completo porque incluye la columna binaria `content`.
    # Si se selecciona el modelo completo, PostgreSQL replica el archivo original por cada registro
    # del censo en el JOIN y Render puede responder 502 por memoria/tiempo.
    # Por eso se seleccionan solo las columnas necesarias para el detalle.
    query = db.session.query(
        Censo.id.label('censo_id'),
        Censo.fecha_censo.label('fecha_censo'),
        Censo.sheet_name.label('censo_sheet_name'),
        Censo.imported_at.label('censo_imported_at'),
        Censo.total_records.label('censo_total_records'),
        Censo.total_occupied.label('censo_total_occupied'),
        Censo.matched_count.label('censo_matched_count'),
        Censo.unmatched_count.label('censo_unmatched_count'),
        UploadedFile.id.label('file_id'),
        UploadedFile.filename.label('file_filename'),
        UploadedFile.file_type.label('file_type'),
        UploadedFile.size_bytes.label('file_size_bytes'),
        UploadedFile.sha256.label('file_sha256'),
        UploadedFile.uploaded_at.label('file_uploaded_at'),
        CensoRecord.id.label('record_id'),
        CensoRecord.solicitud_id.label('record_solicitud_id'),
        CensoRecord.modulo.label('record_modulo'),
        CensoRecord.lugar.label('record_lugar'),
        CensoRecord.habitacion.label('record_habitacion'),
        CensoRecord.empresa.label('record_empresa'),
        CensoRecord.cama.label('record_cama'),
        CensoRecord.dia.label('record_dia'),
        CensoRecord.camas_ocupadas.label('record_camas_ocupadas'),
        CensoRecord.turno.label('record_turno'),
        CensoRecord.gerencia_censo.label('record_gerencia_censo'),
        CensoRecord.area.label('record_area'),
        CensoRecord.rut.label('record_rut'),
        CensoRecord.estado.label('record_estado'),
        CurvaItem.id.label('curva_item_id'),
        CurvaItem.curva_version_id.label('curva_version_id'),
        CurvaVersion.name.label('curva_version_name'),
        CurvaItem.gerencia.label('curva_gerencia'),
        CurvaItem.area.label('curva_area'),
        CurvaItem.empresa.label('curva_empresa'),
        CurvaItem.turno.label('curva_turno'),
        CurvaItem.tipo_contrato.label('curva_tipo_contrato'),
        CurvaItem.formato.label('curva_formato'),
        CurvaItem.camp.label('curva_camp'),
    ).select_from(CensoRecord).join(
        Censo, Censo.id == CensoRecord.censo_id
    ).join(
        UploadedFile, UploadedFile.id == Censo.file_id
    ).outerjoin(
        CurvaItem, CurvaItem.id == CensoRecord.curva_item_id
    ).outerjoin(
        CurvaVersion, CurvaVersion.id == CurvaItem.curva_version_id
    )

    row_idx = 2
    for row in query.yield_per(1000):
        is_matched = bool(row.curva_item_id)
        values = [
            row.censo_id,
            row.fecha_censo.strftime('%Y-%m-%d') if row.fecha_censo else '',
            row.censo_sheet_name or '',
            row.censo_imported_at.strftime('%Y-%m-%d %H:%M:%S') if row.censo_imported_at else '',
            row.censo_total_records,
            row.censo_total_occupied,
            row.censo_matched_count,
            row.censo_unmatched_count,
            row.file_id or '',
            row.file_filename or '',
            row.file_type or '',
            row.file_size_bytes or '',
            row.file_sha256 or '',
            row.file_uploaded_at.strftime('%Y-%m-%d %H:%M:%S') if row.file_uploaded_at else '',
            row.record_id,
            row.record_solicitud_id or '',
            row.record_modulo or '',
            row.record_lugar or '',
            row.record_habitacion or '',
            row.record_empresa or '',
            row.record_cama or '',
            row.record_dia or '',
            row.record_camas_ocupadas,
            row.record_turno or '',
            row.record_gerencia_censo or '',
            row.record_area or '',
            row.record_rut or '',
            row.record_estado or '',
            row.curva_item_id or '',
            row.curva_version_id or '',
            row.curva_version_name or '',
            row.curva_gerencia or '',
            row.curva_area or '',
            row.curva_empresa or '',
            row.curva_turno or '',
            row.curva_tipo_contrato or '',
            row.curva_formato or '',
            row.curva_camp or '',
            'Cruzado' if is_matched else 'Sin match',
        ]
        for col_idx, value in enumerate(values, 1):
            cell = detail.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(vertical='center')
            if col_idx == len(values):
                cell.fill = fill(colors['green'] if is_matched else colors['danger'])
                cell.font = Font(bold=True, color=colors['green_text'] if is_matched else colors['danger_text'])
        row_idx += 1

    detail.freeze_panes = 'A2'
    if row_idx > 2:
        detail.auto_filter.ref = f'A1:{get_column_letter(len(detail_headers))}{row_idx - 1}'
    widths = {
        1: 10, 2: 13, 3: 14, 4: 20, 5: 18, 6: 18, 7: 15, 8: 15,
        9: 10, 10: 34, 11: 12, 12: 18, 13: 24, 14: 20, 15: 12, 16: 18,
        17: 14, 18: 18, 19: 14, 20: 24, 21: 12, 22: 14, 23: 16, 24: 16,
        25: 24, 26: 20, 27: 14, 28: 18, 29: 14, 30: 16, 31: 30, 32: 28,
        33: 24, 34: 24, 35: 16, 36: 20, 37: 16, 38: 16, 39: 14,
    }
    for col_idx in range(1, len(detail_headers) + 1):
        detail.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_idx, 16)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio




def parse_date_text(value):
    value = (value or '').strip()
    if not value:
        return None
    return datetime.strptime(value, '%Y-%m-%d').date()


def excel_cell(ws, value, font=None, fill=None, alignment=None, border=None, number_format=None):
    cell = WriteOnlyCell(ws, value=value)
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format
    return cell


def report_xlsx_fast(data, progress_callback=None):
    """
    Genera un Excel gerencial y liviano con XlsxWriter.

    - La hoja "Dotación Gerencia" queda presentable para envío a gerencia.
    - La hoja "Detalle Censos" reconstruye los censos desde los archivos originales
      almacenados en la base, usando las columnas del archivo de censo, pero mostrando
      el ID corregido/aplicado en curva para que concuerde con el reporte.
    - Se escribe en archivo temporal para evitar alto consumo de memoria en Render.
    """
    if progress_callback:
        progress_callback('Preparando formato gerencial del Excel...')

    def as_int(value):
        try:
            number = float(value or 0)
            return int(round(number))
        except Exception:
            return 0

    def as_pct(real, plan):
        real = float(real or 0)
        plan = float(plan or 0)
        if plan <= 0:
            return 'Sin plan' if real else '0%'
        return f'{real / plan * 100:.1f}%'

    def excel_safe(value):
        if value is None:
            return ''
        try:
            if pd.isna(value):
                return ''
        except Exception:
            pass
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime()
        if isinstance(value, (datetime, date)):
            return value
        return value

    def parse_sort_date(value, fallback=None):
        parsed = parse_date(value)
        if parsed:
            return parsed
        return fallback or date.min

    dates = data.get('date_labels', [])
    rows = data.get('rows', [])
    total_cols = 1 + len(dates) + 4
    last_col = max(total_cols - 1, 7)
    real_total = float(data.get('grand_total') or 0)
    plan_total = float(data.get('planned_grand_total') or 0)
    diff_total = real_total - plan_total
    curve_name = (data.get('curve') or {}).get('name', 'Sin curva seleccionada')
    generated_at = datetime.now().strftime('%d/%m/%Y %H:%M')

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    tmp_path = tmp.name
    tmp.close()

    workbook = xlsxwriter.Workbook(tmp_path, {
        'constant_memory': True,
        'strings_to_urls': False,
        'nan_inf_to_errors': True,
    })

    # Paleta alineada con estética CampCheck / Escondida.
    RED_DARK = '#8F1510'
    RED = '#B42318'
    RED_SOFT = '#FFF1F1'
    RED_CELL = '#F9D7D9'
    YELLOW = '#FFF200'
    ORANGE = '#F58220'
    BLUE = '#E8F1FF'
    GREEN = '#D1FADF'
    GREEN_TEXT = '#027A48'
    DANGER = '#FEE4E2'
    DANGER_TEXT = '#B42318'
    GRAY = '#F2F4F7'
    DARK = '#1D2939'
    BORDER = '#8F1510'
    WHITE = '#FFFFFF'

    fmt_title = workbook.add_format({
        'bold': True, 'font_color': WHITE, 'font_size': 16,
        'align': 'center', 'valign': 'vcenter', 'bg_color': RED_DARK,
        'border': 1, 'border_color': BORDER,
    })
    fmt_subtitle = workbook.add_format({
        'bold': True, 'font_color': DARK, 'font_size': 9,
        'align': 'left', 'valign': 'vcenter', 'bg_color': RED_SOFT,
        'border': 1, 'border_color': '#E4B6B2',
    })
    fmt_header = workbook.add_format({
        'bold': True, 'font_color': WHITE, 'font_size': 8,
        'align': 'center', 'valign': 'vcenter', 'bg_color': RED,
        'border': 1, 'border_color': BORDER,
    })
    fmt_header_left = workbook.add_format({
        'bold': True, 'font_color': WHITE, 'font_size': 8,
        'align': 'left', 'valign': 'vcenter', 'bg_color': RED,
        'border': 1, 'border_color': BORDER,
    })
    fmt_date_header = workbook.add_format({
        'bold': True, 'font_color': WHITE, 'font_size': 8,
        'align': 'center', 'valign': 'vcenter', 'bg_color': RED,
        'border': 1, 'border_color': BORDER, 'rotation': 90,
    })
    fmt_text = workbook.add_format({
        'font_size': 8, 'font_color': DARK, 'align': 'left', 'valign': 'vcenter',
        'border': 1, 'border_color': '#B55B56',
    })
    fmt_num = workbook.add_format({
        'font_size': 8, 'font_color': DARK, 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#B55B56', 'num_format': '#,##0', 'bg_color': RED_CELL,
    })
    fmt_zero = workbook.add_format({
        'font_size': 8, 'font_color': '#98A2B3', 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#E4E7EC', 'num_format': '#,##0', 'bg_color': WHITE,
    })
    fmt_total_real = workbook.add_format({
        'bold': True, 'font_size': 8, 'font_color': '#111111', 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': BORDER, 'num_format': '#,##0', 'bg_color': YELLOW,
    })
    fmt_total_plan = workbook.add_format({
        'bold': True, 'font_size': 8, 'font_color': '#111111', 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#9BBBEA', 'num_format': '#,##0', 'bg_color': BLUE,
    })
    fmt_diff_pos = workbook.add_format({
        'bold': True, 'font_size': 8, 'font_color': GREEN_TEXT, 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#75E0A7', 'num_format': '#,##0', 'bg_color': GREEN,
    })
    fmt_diff_neg = workbook.add_format({
        'bold': True, 'font_size': 8, 'font_color': DANGER_TEXT, 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#FDA29B', 'num_format': '#,##0', 'bg_color': DANGER,
    })
    fmt_pct = workbook.add_format({
        'bold': True, 'font_size': 8, 'font_color': DARK, 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#D0D5DD', 'bg_color': GRAY,
    })
    fmt_total_row = workbook.add_format({
        'bold': True, 'font_size': 8, 'font_color': WHITE, 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': BORDER, 'num_format': '#,##0', 'bg_color': RED,
    })
    fmt_total_row_left = workbook.add_format({
        'bold': True, 'font_size': 8, 'font_color': WHITE, 'align': 'left', 'valign': 'vcenter',
        'border': 1, 'border_color': BORDER, 'bg_color': RED,
    })
    fmt_kpi_label = workbook.add_format({
        'bold': True, 'font_color': WHITE, 'font_size': 9,
        'align': 'center', 'valign': 'vcenter', 'bg_color': RED_DARK,
        'border': 1, 'border_color': BORDER,
    })
    fmt_kpi_value_yellow = workbook.add_format({
        'bold': True, 'font_color': '#111111', 'font_size': 11,
        'align': 'center', 'valign': 'vcenter', 'bg_color': YELLOW,
        'border': 1, 'border_color': BORDER, 'num_format': '#,##0',
    })
    fmt_kpi_value_blue = workbook.add_format({
        'bold': True, 'font_color': '#111111', 'font_size': 11,
        'align': 'center', 'valign': 'vcenter', 'bg_color': BLUE,
        'border': 1, 'border_color': '#9BBBEA', 'num_format': '#,##0',
    })
    fmt_kpi_value_text = workbook.add_format({
        'bold': True, 'font_color': DARK, 'font_size': 11,
        'align': 'center', 'valign': 'vcenter', 'bg_color': GRAY,
        'border': 1, 'border_color': '#D0D5DD',
    })
    fmt_note_title = workbook.add_format({
        'bold': True, 'font_color': WHITE, 'font_size': 10,
        'align': 'left', 'valign': 'vcenter', 'bg_color': RED,
        'border': 1, 'border_color': BORDER,
    })
    fmt_note = workbook.add_format({
        'font_color': DARK, 'font_size': 9, 'align': 'left', 'valign': 'top',
        'bg_color': WHITE, 'border': 1, 'border_color': '#E4E7EC', 'text_wrap': True,
    })
    fmt_unmatched = workbook.add_format({
        'bold': True, 'font_size': 8, 'font_color': '#B54708', 'align': 'left', 'valign': 'vcenter',
        'border': 1, 'border_color': '#FEC84B', 'bg_color': '#FFF3CD',
    })

    # ── Hoja reporte gerencial ─────────────────────────────────────────────
    ws = workbook.add_worksheet('Dotación Gerencia')
    ws.hide_gridlines(2)
    ws.set_tab_color(RED)
    ws.set_landscape()
    ws.fit_to_pages(1, 0)
    ws.set_margins(left=0.25, right=0.25, top=0.45, bottom=0.45)
    ws.repeat_rows(6)
    ws.freeze_panes(7, 1)

    ws.set_column(0, 0, 30)
    if dates:
        ws.set_column(1, len(dates), 5)
    ws.set_column(len(dates) + 1, len(dates) + 4, 13)

    ws.merge_range(0, 0, 0, last_col, 'RESUMEN DE DOTACIÓN POR GERENCIA', fmt_title)
    ws.merge_range(1, 0, 1, last_col, f"Período: {data.get('start_date', '')} al {data.get('end_date', '')}   |   Curva: {curve_name}   |   Generado: {generated_at}", fmt_subtitle)
    ws.set_row(0, 28)
    ws.set_row(1, 21)

    # KPI blocks
    ws.merge_range(3, 0, 3, 1, 'REAL ACUMULADO', fmt_kpi_label)
    ws.merge_range(4, 0, 4, 1, as_int(real_total), fmt_kpi_value_yellow)
    ws.merge_range(3, 2, 3, 3, 'PLAN ACUMULADO', fmt_kpi_label)
    ws.merge_range(4, 2, 4, 3, as_int(plan_total), fmt_kpi_value_blue)
    ws.merge_range(3, 4, 3, 5, 'DIFERENCIA', fmt_kpi_label)
    ws.merge_range(4, 4, 4, 5, as_int(diff_total), fmt_diff_pos if diff_total >= 0 else fmt_diff_neg)
    ws.merge_range(3, 6, 3, 7, 'CUMPLIMIENTO', fmt_kpi_label)
    ws.merge_range(4, 6, 4, 7, as_pct(real_total, plan_total), fmt_kpi_value_text)
    ws.set_row(3, 19)
    ws.set_row(4, 23)

    header_row = 6
    ws.write(header_row, 0, 'GERENCIAS', fmt_header_left)
    for idx, label in enumerate(dates, 1):
        ws.write(header_row, idx, label, fmt_date_header)
    ws.write(header_row, len(dates) + 1, 'TOTAL REAL', fmt_header)
    ws.write(header_row, len(dates) + 2, 'TOTAL PLAN', fmt_header)
    ws.write(header_row, len(dates) + 3, 'DIF.', fmt_header)
    ws.write(header_row, len(dates) + 4, 'CUMP.', fmt_header)
    ws.set_row(header_row, 74)

    row_num = header_row + 1
    for item in rows:
        is_unmatched = item.get('gerencia') == 'SIN MATCH EN CURVA'
        real = float(item.get('total') or 0)
        plan = float(item.get('planned_total') or 0)
        diff = real - plan
        ws.write(row_num, 0, item.get('gerencia', ''), fmt_unmatched if is_unmatched else fmt_text)
        for idx, value in enumerate(item.get('values', []), 1):
            val = as_int(value)
            ws.write_number(row_num, idx, val, fmt_zero if val == 0 else fmt_num)
        ws.write_number(row_num, len(dates) + 1, as_int(real), fmt_total_real)
        ws.write_number(row_num, len(dates) + 2, as_int(plan), fmt_total_plan)
        ws.write_number(row_num, len(dates) + 3, as_int(diff), fmt_diff_pos if diff >= 0 else fmt_diff_neg)
        ws.write(row_num, len(dates) + 4, as_pct(real, plan), fmt_pct)
        row_num += 1

    total_row = row_num
    ws.write(total_row, 0, 'TOTAL', fmt_total_row_left)
    for idx, value in enumerate(data.get('totals_by_date', []), 1):
        ws.write_number(total_row, idx, as_int(value), fmt_total_row)
    ws.write_number(total_row, len(dates) + 1, as_int(real_total), fmt_total_real)
    ws.write_number(total_row, len(dates) + 2, as_int(plan_total), fmt_total_plan)
    ws.write_number(total_row, len(dates) + 3, as_int(diff_total), fmt_diff_pos if diff_total >= 0 else fmt_diff_neg)
    ws.write(total_row, len(dates) + 4, as_pct(real_total, plan_total), fmt_pct)

    if total_row > header_row:
        ws.autofilter(header_row, 0, total_row, last_col)

    notes_start = total_row + 3
    ws.merge_range(notes_start, 0, notes_start, min(last_col, 7), 'CONCLUSIONES AUTOMÁTICAS', fmt_note_title)
    for i, text in enumerate(data.get('conclusions', []), notes_start + 1):
        ws.merge_range(i, 0, i, min(last_col, 7), '• ' + str(text), fmt_note)
        ws.set_row(i, 22)

    # ── Hoja detalle de censos, reconstruida desde archivos originales ─────
    if progress_callback:
        progress_callback('Reconstruyendo detalle de censos desde archivos originales...')

    detail = workbook.add_worksheet('Detalle Censos')
    detail.hide_gridlines(2)
    detail.set_tab_color(ORANGE)
    detail.freeze_panes(1, 0)
    detail.set_landscape()

    preferred_cols = [
        'Modulo', 'Lugar', 'Habitacion', 'Empresa', 'Id', 'Cama', 'Inicio', 'Termino', 'Dia',
        'Camas Ocupdas', 'Turno', 'Gerencia', 'CO MEL', 'AREA', 'SEXO', 'JORNADA', 'Rut', 'PAB', 'ESTADO'
    ]
    fmt_detail_header = workbook.add_format({
        'bold': True, 'font_color': WHITE, 'font_size': 9, 'align': 'center', 'valign': 'vcenter',
        'bg_color': RED, 'border': 1, 'border_color': BORDER, 'text_wrap': True,
    })
    fmt_detail_text = workbook.add_format({
        'font_size': 8, 'font_color': DARK, 'align': 'left', 'valign': 'vcenter',
        'border': 1, 'border_color': '#D0D5DD',
    })
    fmt_detail_center = workbook.add_format({
        'font_size': 8, 'font_color': DARK, 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#D0D5DD',
    })
    fmt_detail_num = workbook.add_format({
        'font_size': 8, 'font_color': DARK, 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#D0D5DD', 'num_format': '#,##0.00',
    })
    fmt_detail_date = workbook.add_format({
        'font_size': 8, 'font_color': DARK, 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#D0D5DD', 'num_format': 'yyyy-mm-dd',
    })
    fmt_detail_zero = workbook.add_format({
        'font_size': 8, 'font_color': '#98A2B3', 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#EAECF0', 'bg_color': '#FAFAFA', 'num_format': '#,##0.00',
    })

    for col_idx, header in enumerate(preferred_cols):
        detail.write(0, col_idx, header, fmt_detail_header)
    detail.set_row(0, 24)
    detail.autofilter(0, 0, 0, len(preferred_cols) - 1)
    widths = {
        'Modulo': 12, 'Lugar': 14, 'Habitacion': 12, 'Empresa': 28, 'Id': 18, 'Cama': 9,
        'Inicio': 12, 'Termino': 12, 'Dia': 13, 'Camas Ocupdas': 15, 'Turno': 14,
        'Gerencia': 30, 'CO MEL': 18, 'AREA': 24, 'SEXO': 10, 'JORNADA': 14,
        'Rut': 15, 'PAB': 14, 'ESTADO': 16,
    }
    for col_idx, header in enumerate(preferred_cols):
        detail.set_column(col_idx, col_idx, widths.get(header, 16))

    censos = Censo.query.join(UploadedFile, UploadedFile.id == Censo.file_id).filter(
        UploadedFile.file_type == 'censo'
    ).order_by(Censo.fecha_censo.asc(), Censo.id.asc()).all()

    detail_row = 1
    for censo in censos:
        # Mapa por censo para reemplazar el ID original del archivo por el ID corregido/aplicado
        # desde la curva. Esto hace que la hoja "Detalle Censos" concuerde con el reporte
        # gerencial y con las correcciones realizadas en "Sin match".
        corrected_id_by_key = {}
        curve_meta_by_key = {}
        correction_rows = db.session.query(
            CensoRecord.solicitud_id.label('id_original'),
            CurvaItem.solicitud_id.label('id_curva'),
            CurvaItem.gerencia.label('gerencia_curva'),
            CurvaItem.area.label('area_curva'),
            CurvaItem.empresa.label('empresa_curva'),
        ).outerjoin(
            CurvaItem, CurvaItem.id == CensoRecord.curva_item_id
        ).filter(
            CensoRecord.censo_id == censo.id,
            CurvaItem.id.isnot(None)
        ).all()
        for corr in correction_rows:
            keys = {norm_id(corr.id_original), compact_id(corr.id_original)}
            for key in keys:
                if key:
                    corrected_id_by_key[key] = corr.id_curva or corr.id_original
                    curve_meta_by_key[key] = {
                        'gerencia': corr.gerencia_curva or '',
                        'area': corr.area_curva or '',
                        'empresa': corr.empresa_curva or '',
                    }

        def corrected_value_for(row_value, fallback=''):
            key_norm = norm_id(row_value)
            key_compact = compact_id(row_value)
            return corrected_id_by_key.get(key_norm) or corrected_id_by_key.get(key_compact) or fallback or row_value

        def corrected_meta_for(row_value):
            key_norm = norm_id(row_value)
            key_compact = compact_id(row_value)
            return curve_meta_by_key.get(key_norm) or curve_meta_by_key.get(key_compact) or {}

        try:
            df, sheet, fecha = read_censo(censo.file.content, censo.file.filename)
        except Exception:
            # Fallback con los datos persistidos si el archivo original no se puede leer.
            fallback_rows = CensoRecord.query.filter_by(censo_id=censo.id).order_by(CensoRecord.id.asc()).all()
            for record in fallback_rows:
                applied_item = record.curva_item
                corrected_id = applied_item.solicitud_id if applied_item else record.solicitud_id
                values = {
                    'Modulo': record.modulo,
                    'Lugar': record.lugar,
                    'Habitacion': record.habitacion,
                    'Empresa': record.empresa,
                    'Id': corrected_id,
                    'Cama': record.cama,
                    'Inicio': '',
                    'Termino': '',
                    'Dia': record.dia,
                    'Camas Ocupdas': record.camas_ocupadas,
                    'Turno': record.turno,
                    'Gerencia': applied_item.gerencia if applied_item else record.gerencia_censo,
                    'CO MEL': '',
                    'AREA': applied_item.area if applied_item else record.area,
                    'SEXO': '',
                    'JORNADA': '',
                    'Rut': record.rut,
                    'PAB': '',
                    'ESTADO': record.estado,
                }
                for col_idx, col_name in enumerate(preferred_cols):
                    val = excel_safe(values.get(col_name, ''))
                    fmt = fmt_detail_center if col_name in ('Modulo', 'Lugar', 'Habitacion', 'Id', 'Cama', 'Turno', 'CO MEL', 'SEXO', 'JORNADA', 'PAB', 'ESTADO') else fmt_detail_text
                    if col_name == 'Camas Ocupdas':
                        fmt = fmt_detail_zero if float(val or 0) == 0 else fmt_detail_num
                    detail.write(detail_row, col_idx, val, fmt)
                detail_row += 1
            continue

        # Eliminar columnas vacías tipo Unnamed y conservar las columnas originales del censo.
        named_cols = [c for c in df.columns if not str(c).lower().startswith('unnamed')]
        df = df[named_cols].copy()
        col_lookup = {norm(c): c for c in df.columns}

        def source_col(expected):
            normalized = norm(expected)
            if normalized in col_lookup:
                return col_lookup[normalized]
            # Soporte para variaciones de escritura comunes.
            aliases = {
                'camas ocupdas': ['camas ocupadas', 'camas ocupdas', 'ocupadas'],
                'co mel': ['co mel', 'comel'],
                'area': ['area', 'área'],
                'termino': ['termino', 'término'],
            }.get(normalized, [])
            for alias in aliases:
                if norm(alias) in col_lookup:
                    return col_lookup[norm(alias)]
            return None

        # Ordenar por la fecha de la columna Día y luego por ubicación/cama, manteniendo formato original.
        dia_col = source_col('Dia')
        if dia_col:
            df['_sort_fecha_'] = df[dia_col].apply(lambda value: parse_sort_date(value, censo.fecha_censo))
        else:
            df['_sort_fecha_'] = censo.fecha_censo
        sort_cols = ['_sort_fecha_']
        for candidate in ('Modulo', 'Lugar', 'Habitacion', 'Cama'):
            col = source_col(candidate)
            if col:
                sort_cols.append(col)
        df = df.sort_values(sort_cols, kind='mergesort')

        id_src_col = source_col('Id')
        for _, row in df.iterrows():
            original_id_value = excel_safe(row[id_src_col]) if id_src_col else ''
            applied_id_value = corrected_value_for(original_id_value, original_id_value)
            applied_meta = corrected_meta_for(original_id_value)

            for col_idx, col_name in enumerate(preferred_cols):
                src = source_col(col_name)
                value = excel_safe(row[src]) if src else ''

                # Reemplazar el ID del archivo por el ID corregido/aplicado en curva.
                # También se actualizan Gerencia, AREA y Empresa cuando existe match con curva,
                # para que el detalle quede consistente con el reporte.
                if col_name == 'Id':
                    value = applied_id_value
                elif col_name == 'Gerencia' and applied_meta.get('gerencia'):
                    value = applied_meta['gerencia']
                elif col_name == 'AREA' and applied_meta.get('area'):
                    value = applied_meta['area']
                elif col_name == 'Empresa' and applied_meta.get('empresa'):
                    value = applied_meta['empresa']

                if col_name == 'Dia' and isinstance(value, str):
                    parsed = parse_date(value)
                    if parsed:
                        value = parsed
                if col_name in ('Inicio', 'Termino') and isinstance(value, str):
                    parsed = parse_date(value)
                    if parsed:
                        value = parsed
                if col_name == 'Camas Ocupdas':
                    number = as_number(value)
                    detail.write_number(detail_row, col_idx, number, fmt_detail_zero if number == 0 else fmt_detail_num)
                elif isinstance(value, (datetime, date)):
                    detail.write_datetime(detail_row, col_idx, datetime(value.year, value.month, value.day), fmt_detail_date)
                else:
                    fmt = fmt_detail_center if col_name in ('Modulo', 'Lugar', 'Habitacion', 'Id', 'Cama', 'Turno', 'CO MEL', 'SEXO', 'JORNADA', 'PAB', 'ESTADO') else fmt_detail_text
                    detail.write(detail_row, col_idx, value, fmt)
            detail_row += 1

        if progress_callback:
            progress_callback(f'Detalle de censos: {detail_row - 1:,} filas escritas...'.replace(',', '.'))

    # Actualizar autofiltro al rango completo.
    if detail_row > 1:
        detail.autofilter(0, 0, detail_row - 1, len(preferred_cols) - 1)

    # ── Hoja de auditoría del cruce, usando las correcciones aplicadas ────
    # Esta hoja NO reemplaza el detalle original. Sirve para verificar que las
    # correcciones hechas en "Sin match" sí están consideradas en el reporte.
    if progress_callback:
        progress_callback('Generando auditoría de cruces corregidos...')

    audit = workbook.add_worksheet('Auditoria Cruce')
    audit.hide_gridlines(2)
    audit.set_tab_color(BLUE)
    audit.freeze_panes(1, 0)
    audit.set_landscape()

    audit_headers = [
        'Fecha Censo', 'Archivo Censo', 'Modulo', 'Lugar', 'Habitacion', 'Empresa Censo',
        'ID Censo Original', 'ID Curva Aplicado', 'Gerencia Curva', 'Gerencia Censo',
        'Área Curva', 'Área Censo', 'Cama', 'Dia', 'Camas Ocupadas', 'Turno Censo',
        'Rut', 'Estado Censo', 'Estado Match'
    ]
    fmt_audit_header = workbook.add_format({
        'bold': True, 'font_color': WHITE, 'font_size': 9, 'align': 'center', 'valign': 'vcenter',
        'bg_color': BLUE_DARK, 'border': 1, 'border_color': '#9BBBEA', 'text_wrap': True,
    })
    fmt_audit_text = workbook.add_format({
        'font_size': 8, 'font_color': DARK, 'align': 'left', 'valign': 'vcenter',
        'border': 1, 'border_color': '#D0D5DD',
    })
    fmt_audit_center = workbook.add_format({
        'font_size': 8, 'font_color': DARK, 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#D0D5DD',
    })
    fmt_audit_num = workbook.add_format({
        'font_size': 8, 'font_color': DARK, 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#D0D5DD', 'num_format': '#,##0.00',
    })
    fmt_audit_date = workbook.add_format({
        'font_size': 8, 'font_color': DARK, 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#D0D5DD', 'num_format': 'yyyy-mm-dd',
    })
    fmt_audit_ok = workbook.add_format({
        'bold': True, 'font_size': 8, 'font_color': GREEN_TEXT, 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#75E0A7', 'bg_color': GREEN,
    })
    fmt_audit_bad = workbook.add_format({
        'bold': True, 'font_size': 8, 'font_color': DANGER_TEXT, 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#FDA29B', 'bg_color': DANGER,
    })

    for col_idx, header in enumerate(audit_headers):
        audit.write(0, col_idx, header, fmt_audit_header)
    audit.set_row(0, 30)
    audit_widths = [13, 34, 12, 14, 12, 28, 20, 20, 30, 30, 24, 24, 9, 13, 15, 14, 15, 16, 15]
    for col_idx, width in enumerate(audit_widths):
        audit.set_column(col_idx, col_idx, width)

    audit_row = 1
    last_seen_id = 0
    audit_batch_size = 2000
    while True:
        rows_batch = db.session.query(
            CensoRecord.id.label('record_id'),
            Censo.fecha_censo.label('fecha_censo'),
            UploadedFile.filename.label('filename'),
            CensoRecord.modulo.label('modulo'),
            CensoRecord.lugar.label('lugar'),
            CensoRecord.habitacion.label('habitacion'),
            CensoRecord.empresa.label('empresa_censo'),
            CensoRecord.solicitud_id.label('id_censo'),
            CurvaItem.solicitud_id.label('id_curva'),
            CurvaItem.gerencia.label('gerencia_curva'),
            CensoRecord.gerencia_censo.label('gerencia_censo'),
            CurvaItem.area.label('area_curva'),
            CensoRecord.area.label('area_censo'),
            CensoRecord.cama.label('cama'),
            CensoRecord.dia.label('dia'),
            CensoRecord.camas_ocupadas.label('camas_ocupadas'),
            CensoRecord.turno.label('turno'),
            CensoRecord.rut.label('rut'),
            CensoRecord.estado.label('estado'),
        ).join(
            Censo, Censo.id == CensoRecord.censo_id
        ).join(
            UploadedFile, UploadedFile.id == Censo.file_id
        ).outerjoin(
            CurvaItem, CurvaItem.id == CensoRecord.curva_item_id
        ).filter(
            CensoRecord.id > last_seen_id
        ).order_by(
            Censo.fecha_censo.asc(), CensoRecord.id.asc()
        ).limit(audit_batch_size).all()

        if not rows_batch:
            break

        for row in rows_batch:
            last_seen_id = row.record_id
            matched = bool(row.id_curva)
            values = [
                row.fecha_censo,
                row.filename,
                row.modulo,
                row.lugar,
                row.habitacion,
                row.empresa_censo,
                row.id_censo,
                row.id_curva or '',
                row.gerencia_curva or '',
                row.gerencia_censo,
                row.area_curva or '',
                row.area_censo,
                row.cama,
                row.dia,
                row.camas_ocupadas or 0,
                row.turno,
                row.rut,
                row.estado,
                'Cruzado' if matched else 'Sin match',
            ]
            for col_idx, value in enumerate(values):
                if col_idx == 0 and isinstance(value, date):
                    audit.write_datetime(audit_row, col_idx, datetime(value.year, value.month, value.day), fmt_audit_date)
                elif col_idx == 14:
                    audit.write_number(audit_row, col_idx, float(value or 0), fmt_audit_num)
                elif col_idx == 18:
                    audit.write(audit_row, col_idx, value, fmt_audit_ok if matched else fmt_audit_bad)
                elif col_idx in (2, 3, 4, 6, 7, 12, 13, 15, 16, 17):
                    audit.write(audit_row, col_idx, excel_safe(value), fmt_audit_center)
                else:
                    audit.write(audit_row, col_idx, excel_safe(value), fmt_audit_text)
            audit_row += 1

        if progress_callback:
            progress_callback(f'Auditoría de cruces: {audit_row - 1:,} filas escritas...'.replace(',', '.'))

    if audit_row > 1:
        audit.autofilter(0, 0, audit_row - 1, len(audit_headers) - 1)

    if progress_callback:
        progress_callback('Comprimiendo archivo Excel...')

    workbook.close()
    with open(tmp_path, 'rb') as fh:
        content = fh.read()
    try:
        os.remove(tmp_path)
    except OSError:
        pass
    return BytesIO(content)







def recalc_censo_totals(censo_id):
    """Recalcula métricas del censo sobre reservas: filas con ID."""
    censo = Censo.query.get(censo_id)
    if not censo:
        return None
    total = CensoRecord.query.filter_by(censo_id=censo_id).count()
    occupied = db.session.query(func.coalesce(func.sum(CensoRecord.camas_ocupadas), 0)).filter_by(censo_id=censo_id).scalar() or 0
    matched = CensoRecord.query.filter(
        CensoRecord.censo_id == censo_id,
        CensoRecord.curva_item_id.isnot(None)
    ).count()
    censo.total_records = total
    censo.total_occupied = float(occupied or 0)
    censo.matched_count = matched
    censo.unmatched_count = max(total - matched, 0)
    return censo


def recalc_many_censos(censo_ids):
    for censo_id in sorted(set(censo_ids)):
        recalc_censo_totals(censo_id)


def apply_correction_to_solicitud_id(solicitud_id, item):
    """
    Corrige todos los registros que tengan el mismo ID original de censo.
    Incluye reservas con Camas Ocupdas = 0.
    """
    sid = norm_id(solicitud_id)
    if not sid or not item:
        return 0, []
    rows = CensoRecord.query.filter(CensoRecord.solicitud_id == sid).all()
    affected_censos = set()
    for row in rows:
        row.curva_item_id = item.id
        affected_censos.add(row.censo_id)
    recalc_many_censos(affected_censos)
    return len(rows), list(affected_censos)


def apply_correction_to_same_id(record, item):
    """
    Corrige en lote todos los registros con el mismo ID del registro elegido.
    La corrección alcanza reservas ocupadas y no ocupadas.
    """
    if not record:
        return 0, []
    return apply_correction_to_solicitud_id(record.solicitud_id, item)


def preserved_matches_for_censo(censo_id):
    """Conserva correcciones manuales existentes antes de reprocesar un censo."""
    rows = CensoRecord.query.filter(
        CensoRecord.censo_id == censo_id,
        CensoRecord.curva_item_id.isnot(None),
        CensoRecord.solicitud_id.isnot(None),
        CensoRecord.solicitud_id != ''
    ).all()
    preserved = {}
    for row in rows:
        sid = norm_id(row.solicitud_id)
        if sid and row.curva_item_id:
            preserved[sid] = row.curva_item_id
    return preserved


def reprocess_censo_from_original_file(censo):
    """
    Reprocesa un censo desde el archivo original guardado en BD para incluir
    reservas con Camas Ocupdas = 0 y recalcular métricas.
    """
    preserved = preserved_matches_for_censo(censo.id)
    df, sheet, fecha = read_censo(censo.file.content, censo.file.filename)
    if fecha:
        censo.fecha_censo = fecha
    censo.sheet_name = sheet
    rebuild_censo_from_dataframe(censo, df, preserved_matches=preserved)
    return censo

def resolve_report_dates(start=None, end=None):
    """Devuelve rango de fechas válido para reportes/exportaciones."""
    if not start or not end:
        minmax = db.session.query(func.min(Censo.fecha_censo), func.max(Censo.fecha_censo)).one()
        start = start or minmax[0]
        end = end or minmax[1]
    return start, end


def format_export_filename(export_mode):
    stamp = datetime.now().strftime('%Y%m%d_%H%M')
    if export_mode == 'censos':
        return f'censos_acumulados_{stamp}.xlsx'
    return f'dotacion_gerencia_{stamp}.xlsx'


def csv_row(values):
    """Convierte una lista de valores en una línea CSV compatible con Excel."""
    output = StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL, lineterminator='\n')
    writer.writerow(['' if value is None else value for value in values])
    return output.getvalue()


def source_col_from_map(cm, key):
    """Devuelve el nombre de columna de un DataFrame según alias normalizados."""
    return cm.get(key)


def build_corrected_meta_map(censo_id):
    """
    Mapa por ID original de censo para aplicar correcciones hechas en Sin match.
    Si varios registros tienen el mismo ID, todos reciben el mismo ID/metadata de curva.
    """
    rows = db.session.query(
        CensoRecord.solicitud_id.label('id_censo'),
        CurvaItem.solicitud_id.label('id_curva'),
        CurvaItem.gerencia.label('gerencia'),
        CurvaItem.area.label('area'),
        CurvaItem.empresa.label('empresa'),
    ).outerjoin(
        CurvaItem, CurvaItem.id == CensoRecord.curva_item_id
    ).filter(
        CensoRecord.censo_id == censo_id,
        CensoRecord.solicitud_id.isnot(None),
        CensoRecord.solicitud_id != '',
    ).all()

    mapping = {}
    for row in rows:
        original = norm_id(row.id_censo)
        if not original:
            continue
        if row.id_curva:
            mapping[original] = {
                'id': row.id_curva or original,
                'gerencia': row.gerencia or '',
                'area': row.area or '',
                'empresa': row.empresa or '',
            }
        elif original not in mapping:
            mapping[original] = {'id': original, 'gerencia': '', 'area': '', 'empresa': ''}
    return mapping


def iter_censos_acumulados_csv(start_date=None, end_date=None, batch_size=10):
    """
    Exportación liviana de censos acumulados como CSV.

    Reconstruye la data desde los archivos originales de censo guardados en
    PostgreSQL. Así conserva columnas como Inicio, Termino, CO MEL, SEXO y
    JORNADA, e incluye también filas con Camas Ocupdas = 0.

    Los IDs corregidos en Sin match se aplican por ID original; por eso todos los
    registros iguales quedan corregidos en el CSV.
    """
    headers = [
        'Modulo', 'Lugar', 'Habitacion', 'Empresa', 'Id', 'Cama', 'Inicio', 'Termino', 'Dia',
        'Camas Ocupdas', 'Turno', 'Gerencia', 'CO MEL', 'AREA', 'SEXO', 'JORNADA', 'Rut', 'PAB', 'ESTADO'
    ]

    # BOM UTF-8 + instrucción sep=; para que Excel separe columnas automáticamente.
    yield '\ufeffsep=;\n'
    yield csv_row(headers)

    base_query = Censo.query.join(UploadedFile, UploadedFile.id == Censo.file_id).filter(
        UploadedFile.file_type == 'censo'
    )
    if start_date:
        base_query = base_query.filter(Censo.fecha_censo >= start_date)
    if end_date:
        base_query = base_query.filter(Censo.fecha_censo <= end_date)

    last_date = None
    last_id = 0
    while True:
        query = base_query
        if last_date is not None:
            query = query.filter(or_(
                Censo.fecha_censo > last_date,
                and_(Censo.fecha_censo == last_date, Censo.id > last_id)
            ))
        censos = query.order_by(Censo.fecha_censo.asc(), Censo.id.asc()).limit(batch_size).all()
        if not censos:
            break

        for censo in censos:
            corrected_map = build_corrected_meta_map(censo.id)
            try:
                df, sheet, fecha_detectada = read_censo(censo.file.content, censo.file.filename)
            except Exception:
                # Fallback si algún archivo original no se puede abrir: usa lo persistido.
                fallback = CensoRecord.query.filter_by(censo_id=censo.id).order_by(
                    CensoRecord.modulo.asc(), CensoRecord.lugar.asc(),
                    CensoRecord.habitacion.asc(), CensoRecord.cama.asc(), CensoRecord.id.asc()
                ).all()
                for record in fallback:
                    item = record.curva_item
                    yield csv_row([
                        record.modulo or '',
                        record.lugar or '',
                        record.habitacion or '',
                        item.empresa if item else (record.empresa or ''),
                        item.solicitud_id if item else (record.solicitud_id or ''),
                        record.cama or '',
                        '',
                        '',
                        record.dia or (censo.fecha_censo.strftime('%Y-%m-%d') if censo.fecha_censo else ''),
                        float(record.camas_ocupadas or 0),
                        record.turno or '',
                        item.gerencia if item else (record.gerencia_censo or ''),
                        '',
                        item.area if item else (record.area or ''),
                        '',
                        '',
                        record.rut or '',
                        '',
                        record.estado or '',
                    ])
                continue

            # Quita columnas vacías generadas por pandas, pero conserva nombres reales.
            valid_cols = [col for col in df.columns if clean(col) and not str(col).lower().startswith('unnamed')]
            df = df[valid_cols].copy()
            cm = colmap(df.columns)

            # Orden estable por fecha/ubicación/cama, manteniendo todos los registros del archivo.
            dia_col = source_col_from_map(cm, 'dia')
            if dia_col:
                df['_sort_fecha_'] = df[dia_col].apply(lambda value: parse_date(value) or censo.fecha_censo)
            else:
                df['_sort_fecha_'] = censo.fecha_censo
            sort_cols = ['_sort_fecha_']
            for key in ('modulo', 'lugar', 'habitacion', 'cama'):
                col = source_col_from_map(cm, key)
                if col:
                    sort_cols.append(col)
            df = df.sort_values(sort_cols, kind='stable')

            def get_value(row, key):
                col = source_col_from_map(cm, key)
                return clean(row.get(col)) if col else ''

            for _, row in df.iterrows():
                id_col = source_col_from_map(cm, 'id')
                original_id = norm_id(row.get(id_col)) if id_col else ''
                meta = corrected_map.get(original_id, {}) if original_id else {}
                dia_value = row.get(dia_col) if dia_col else censo.fecha_censo
                parsed_dia = parse_date(dia_value)
                dia_text = parsed_dia.strftime('%Y-%m-%d') if parsed_dia else clean(dia_value) or censo.fecha_censo.strftime('%Y-%m-%d')
                ocupadas_col = source_col_from_map(cm, 'ocupadas')

                yield csv_row([
                    get_value(row, 'modulo'),
                    get_value(row, 'lugar'),
                    get_value(row, 'habitacion'),
                    meta.get('empresa') or get_value(row, 'empresa'),
                    meta.get('id') or original_id,
                    get_value(row, 'cama'),
                    get_value(row, 'inicio'),
                    get_value(row, 'termino'),
                    dia_text,
                    as_number(row.get(ocupadas_col)) if ocupadas_col else 0,
                    get_value(row, 'turno'),
                    meta.get('gerencia') or get_value(row, 'gerencia'),
                    get_value(row, 'co_mel'),
                    meta.get('area') or get_value(row, 'area'),
                    get_value(row, 'sexo'),
                    get_value(row, 'jornada'),
                    get_value(row, 'rut'),
                    get_value(row, 'pab'),
                    get_value(row, 'estado'),
                ])

            last_date = censo.fecha_censo
            last_id = censo.id



def compact_id(value):
    """Normaliza un ID para comparación flexible: quita espacios, guiones y símbolos."""
    return re.sub(r"[^A-Z0-9]", "", norm_id(value))


def digits_id(value):
    """Retorna solo los dígitos del ID para detectar diferencias por ceros, puntos o guiones."""
    return re.sub(r"\D", "", str(value or ""))


def recalc_censo_stats(censo_id):
    """Alias compatible: recalcula métricas del censo sobre reservas, no solo ocupados."""
    return recalc_censo_totals(censo_id)


def set_latest_curve_active_if_needed():
    active = CurvaVersion.query.filter_by(is_active=True).first()
    if active:
        return active
    latest = CurvaVersion.query.order_by(CurvaVersion.uploaded_at.desc()).first()
    if latest:
        latest.is_active = True
    return latest


def delete_uploaded_file_data(file_id):
    """
    Elimina un archivo subido y toda la información derivada.

    - Si es censo: elimina sus reservas/registros y el censo.
    - Si es curva: elimina versiones, IDs y planificación; los registros de censo que
      apuntaban a esa curva quedan como sin match para poder corregirlos con otra curva.
    """
    uploaded = UploadedFile.query.get_or_404(file_id)
    impacted_censo_ids = set()

    if uploaded.file_type == "censo":
        censos = Censo.query.filter_by(file_id=uploaded.id).all()
        censo_ids = [c.id for c in censos]
        if censo_ids:
            CensoRecord.query.filter(CensoRecord.censo_id.in_(censo_ids)).delete(synchronize_session=False)
            Censo.query.filter(Censo.id.in_(censo_ids)).delete(synchronize_session=False)

    elif uploaded.file_type == "curva":
        curves = CurvaVersion.query.filter_by(file_id=uploaded.id).all()
        for curve in curves:
            item_ids = [row[0] for row in db.session.query(CurvaItem.id).filter_by(curva_version_id=curve.id).all()]
            if item_ids:
                impacted_censo_ids.update(
                    row[0] for row in db.session.query(CensoRecord.censo_id)
                    .filter(CensoRecord.curva_item_id.in_(item_ids))
                    .distinct()
                    .all()
                )
                CensoRecord.query.filter(CensoRecord.curva_item_id.in_(item_ids)).update(
                    {CensoRecord.curva_item_id: None}, synchronize_session=False
                )
                CurvaDailyValue.query.filter(CurvaDailyValue.curva_item_id.in_(item_ids)).delete(synchronize_session=False)
                CurvaItem.query.filter(CurvaItem.id.in_(item_ids)).delete(synchronize_session=False)
            db.session.delete(curve)
    else:
        raise ValueError("Tipo de archivo no reconocido.")

    db.session.delete(uploaded)
    db.session.flush()

    for censo_id in impacted_censo_ids:
        recalc_censo_totals(censo_id)

    set_latest_curve_active_if_needed()
    db.session.commit()
    return uploaded


def get_curve_for_matching(curve_id=None):
    if curve_id:
        return CurvaVersion.query.get(curve_id)
    return CurvaVersion.query.filter_by(is_active=True).order_by(CurvaVersion.uploaded_at.desc()).first()


def build_curve_candidates(curve):
    if not curve:
        return []
    return CurvaItem.query.filter_by(curva_version_id=curve.id).all()


def build_curve_match_index(curve):
    """
    Índice rápido para proponer correcciones Sin match sin comparar cada reserva
    contra toda la curva. Usa ID compacto, dígitos y sufijos.
    """
    items = build_curve_candidates(curve)
    index = {
        "items": items,
        "by_compact": defaultdict(list),
        "by_digits": defaultdict(list),
        "by_suffix": defaultdict(list),
    }

    for item in items:
        compact = compact_id(item.solicitud_id)
        digits = digits_id(item.solicitud_id).lstrip("0")

        if compact:
            index["by_compact"][compact].append(item)
        if digits:
            index["by_digits"][digits].append(item)
            for n in range(4, min(len(digits), 12) + 1):
                index["by_suffix"][(n, digits[-n:])].append(item)

    return index


def candidate_score(record, item):
    target = compact_id(record.solicitud_id)
    candidate = compact_id(item.solicitud_id)
    target_digits = digits_id(record.solicitud_id).lstrip("0")
    candidate_digits = digits_id(item.solicitud_id).lstrip("0")

    if not target or not candidate:
        return 0, ""

    score = 0
    reason = "Similitud de texto"

    if target == candidate:
        score, reason = 100, "Coincidencia exacta ignorando formato"
    elif target_digits and candidate_digits and target_digits == candidate_digits:
        score, reason = 98, "Coinciden los dígitos ignorando ceros o símbolos"
    elif target in candidate or candidate in target:
        score, reason = 90, "Un ID contiene al otro"
    elif target_digits and candidate_digits:
        max_suffix = 0
        for n in range(min(len(target_digits), len(candidate_digits)), 3, -1):
            if target_digits[-n:] == candidate_digits[-n:]:
                max_suffix = n
                break
        if max_suffix:
            score = min(88, 70 + max_suffix * 2)
            reason = f"Coinciden los últimos {max_suffix} dígitos"

    ratio = SequenceMatcher(None, target, candidate).ratio()
    if ratio >= 0.70 and int(ratio * 85) > score:
        score = int(ratio * 85)
        reason = f"Similitud de ID {ratio * 100:.0f}%"

    if record.empresa and item.empresa and norm(record.empresa) == norm(item.empresa):
        score += 4
        reason += " + misma empresa"
    if record.gerencia_censo and item.gerencia and norm(record.gerencia_censo) == norm(item.gerencia):
        score += 4
        reason += " + misma gerencia"

    return min(score, 100), reason


def candidate_pool_for_record(record, match_index, max_candidates=250):
    if not match_index:
        return []

    target = compact_id(record.solicitud_id)
    target_digits = digits_id(record.solicitud_id).lstrip("0")
    pool = []
    seen = set()

    def add(items):
        for item in items or []:
            if item.id not in seen:
                seen.add(item.id)
                pool.append(item)
                if len(pool) >= max_candidates:
                    return

    if target:
        add(match_index["by_compact"].get(target))
    if target_digits:
        add(match_index["by_digits"].get(target_digits))
        for n in range(min(len(target_digits), 12), 3, -1):
            add(match_index["by_suffix"].get((n, target_digits[-n:])))
            if len(pool) >= max_candidates:
                break

    if not pool and target:
        for item in match_index["items"][:max_candidates]:
            candidate = compact_id(item.solicitud_id)
            if candidate and (candidate[:4] == target[:4] or candidate[-4:] == target[-4:]):
                add([item])

    return pool[:max_candidates]


def suggestions_for_record(record, match_index, limit=3):
    suggestions = []
    for item in candidate_pool_for_record(record, match_index):
        score, reason = candidate_score(record, item)
        if score >= 60:
            suggestions.append({
                "item": item,
                "score": score,
                "reason": reason,
            })
    suggestions.sort(key=lambda x: x["score"], reverse=True)
    return suggestions[:limit]


def find_curve_item_by_manual_id(curve, manual_id):
    """Busca un ID manual en la curva, aceptando diferencias simples de formato."""
    if not curve or not manual_id:
        return None

    manual_id = norm_id(manual_id)
    item = CurvaItem.query.filter_by(curva_version_id=curve.id, solicitud_id=manual_id).first()
    if item:
        return item

    target = compact_id(manual_id)
    target_digits = digits_id(manual_id).lstrip("0")
    for item in CurvaItem.query.filter_by(curva_version_id=curve.id).all():
        if target and compact_id(item.solicitud_id) == target:
            return item
        if target_digits and digits_id(item.solicitud_id).lstrip("0") == target_digits:
            return item
    return None


def no_match_payload(censo_id=None, curve_id=None, search_text="", limit=200):
    curve = get_curve_for_matching(curve_id)
    match_index = build_curve_match_index(curve)

    base_query = CensoRecord.query.join(Censo, Censo.id == CensoRecord.censo_id).filter(
        CensoRecord.curva_item_id.is_(None),
        CensoRecord.solicitud_id.isnot(None),
        CensoRecord.solicitud_id != '',
    )
    if censo_id:
        base_query = base_query.filter(CensoRecord.censo_id == censo_id)

    search_text = (search_text or "").strip()
    if search_text:
        like = f"%{search_text}%"
        base_query = base_query.filter(
            or_(
                CensoRecord.solicitud_id.ilike(like),
                CensoRecord.empresa.ilike(like),
                CensoRecord.gerencia_censo.ilike(like),
                CensoRecord.rut.ilike(like),
                CensoRecord.habitacion.ilike(like),
            )
        )

    total_records = base_query.count()
    total_ids = base_query.with_entities(CensoRecord.solicitud_id).distinct().count()

    grouped = base_query.with_entities(
        CensoRecord.solicitud_id.label("solicitud_id"),
        func.count(CensoRecord.id).label("same_id_count"),
        func.min(Censo.fecha_censo).label("first_date"),
        func.max(Censo.fecha_censo).label("last_date"),
        func.min(CensoRecord.id).label("sample_id"),
    ).group_by(
        CensoRecord.solicitud_id
    ).order_by(
        func.max(Censo.fecha_censo).desc(),
        CensoRecord.solicitud_id.asc()
    ).limit(limit).all()

    sample_ids = [row.sample_id for row in grouped]
    sample_records = {
        record.id: record
        for record in CensoRecord.query.filter(CensoRecord.id.in_(sample_ids)).all()
    } if sample_ids else {}

    rows = []
    for group in grouped:
        record = sample_records.get(group.sample_id)
        if not record:
            continue
        rows.append({
            "record": record,
            "censo": record.censo,
            "same_id_count": int(group.same_id_count or 0),
            "first_date": group.first_date,
            "last_date": group.last_date,
            "suggestions": suggestions_for_record(record, match_index),
        })

    return {
        "curve": curve,
        "total": total_ids,
        "total_records": total_records,
        "limit": limit,
        "rows": rows,
    }


def refresh_curve_totals(curve_id):
    """Recalcula totales de una curva después de agregar/editar IDs por pantalla."""
    curve = CurvaVersion.query.get(curve_id)
    if not curve:
        return None
    curve.total_items = CurvaItem.query.filter_by(curva_version_id=curve.id).count()
    curve.total_daily_values = db.session.query(func.count(CurvaDailyValue.id)).join(
        CurvaItem, CurvaItem.id == CurvaDailyValue.curva_item_id
    ).filter(CurvaItem.curva_version_id == curve.id).scalar() or 0
    return curve


def create_or_update_curve_item_from_form(form):
    """
    Crea un nuevo ID en la curva desde pantalla.
    Si el ID ya existe en la curva seleccionada, actualiza sus datos maestros.
    Opcionalmente crea valores diarios de planificación para un rango de fechas.
    """
    curve_id = form.get('curve_id', type=int)
    curve = CurvaVersion.query.get(curve_id) if curve_id else get_curve_for_matching(None)
    if not curve:
        raise ValueError('No hay una curva disponible. Primero importa una curva planificada.')

    solicitud_id = norm_id(form.get('solicitud_id'))
    if not solicitud_id:
        raise ValueError('El ID de solicitud es obligatorio.')

    gerencia = clean(form.get('gerencia')) or 'SIN GERENCIA'
    area = clean(form.get('area'))
    empresa = clean(form.get('empresa'))
    turno = clean(form.get('turno'))
    tipo_contrato = clean(form.get('tipo_contrato'))
    formato = clean(form.get('formato'))
    camp = clean(form.get('camp'))

    item = CurvaItem.query.filter_by(curva_version_id=curve.id, solicitud_id=solicitud_id).first()
    created = item is None
    if created:
        item = CurvaItem(curva_version_id=curve.id, solicitud_id=solicitud_id)
        db.session.add(item)

    item.gerencia = gerencia
    item.area = area
    item.empresa = empresa
    item.turno = turno
    item.tipo_contrato = tipo_contrato
    item.formato = formato
    item.camp = camp
    db.session.flush()

    date_from = parse_date(form.get('date_from'))
    date_to = parse_date(form.get('date_to'))
    dotacion = as_number(form.get('dotacion_planificada'))
    created_values = 0

    if date_from or date_to or dotacion:
        if not date_from or not date_to:
            raise ValueError('Para cargar planificación diaria debes indicar fecha desde y hasta.')
        if date_to < date_from:
            raise ValueError('La fecha hasta no puede ser menor que la fecha desde.')
        if dotacion < 0:
            raise ValueError('La dotación planificada no puede ser negativa.')

        CurvaDailyValue.query.filter(
            CurvaDailyValue.curva_item_id == item.id,
            CurvaDailyValue.fecha >= date_from,
            CurvaDailyValue.fecha <= date_to,
        ).delete(synchronize_session=False)

        values = [
            CurvaDailyValue(curva_item_id=item.id, fecha=d, dotacion_planificada=dotacion)
            for d in date_span(date_from, date_to)
        ]
        if values:
            db.session.bulk_save_objects(values)
            created_values = len(values)

    refresh_curve_totals(curve.id)
    return curve, item, created, created_values


def sample_record_context(record_id):
    """Obtiene un registro sin match para prefijar el formulario de nuevo ID."""
    if not record_id:
        return None
    return CensoRecord.query.get(record_id)


def get_curve_items_payload(curve_id=None, search_text='', gerencia='', page=1, per_page=50):
    """Prepara la vista paginada de datos maestros de la curva."""
    curve = get_curve_for_matching(curve_id)
    curves = CurvaVersion.query.order_by(CurvaVersion.is_active.desc(), CurvaVersion.uploaded_at.desc()).all()

    page = max(int(page or 1), 1)
    per_page = min(max(int(per_page or 50), 10), 200)
    search_text = (search_text or '').strip()
    gerencia = (gerencia or '').strip()

    if not curve:
        return {
            'curve': None, 'curves': curves, 'items': [], 'stats': {}, 'gerencias': [],
            'total': 0, 'page': page, 'per_page': per_page, 'pages': 0,
            'q': search_text, 'selected_gerencia': gerencia,
        }

    query = CurvaItem.query.filter(CurvaItem.curva_version_id == curve.id)
    if search_text:
        like = f'%{search_text}%'
        query = query.filter(or_(
            CurvaItem.solicitud_id.ilike(like),
            CurvaItem.gerencia.ilike(like),
            CurvaItem.area.ilike(like),
            CurvaItem.empresa.ilike(like),
            CurvaItem.turno.ilike(like),
            CurvaItem.tipo_contrato.ilike(like),
            CurvaItem.formato.ilike(like),
            CurvaItem.camp.ilike(like),
        ))
    if gerencia:
        query = query.filter(CurvaItem.gerencia == gerencia)

    total = query.count()
    pages = (total + per_page - 1) // per_page if total else 0
    if pages and page > pages:
        page = pages

    items = query.order_by(CurvaItem.gerencia.asc(), CurvaItem.solicitud_id.asc()).offset((page - 1) * per_page).limit(per_page).all()
    item_ids = [item.id for item in items]

    stats = {}
    if item_ids:
        rows = db.session.query(
            CurvaDailyValue.curva_item_id,
            func.sum(CurvaDailyValue.dotacion_planificada).label('plan_total'),
            func.min(CurvaDailyValue.fecha).label('first_date'),
            func.max(CurvaDailyValue.fecha).label('last_date'),
        ).filter(
            CurvaDailyValue.curva_item_id.in_(item_ids)
        ).group_by(
            CurvaDailyValue.curva_item_id
        ).all()
        stats = {row.curva_item_id: row for row in rows}

    gerencias = [row[0] for row in db.session.query(CurvaItem.gerencia).filter(
        CurvaItem.curva_version_id == curve.id
    ).distinct().order_by(CurvaItem.gerencia.asc()).all()]

    return {
        'curve': curve, 'curves': curves, 'items': items, 'stats': stats, 'gerencias': gerencias,
        'total': total, 'page': page, 'per_page': per_page, 'pages': pages,
        'q': search_text, 'selected_gerencia': gerencia,
    }


def get_curve_item_for_edit(item_id):
    if not item_id:
        return None
    return CurvaItem.query.get_or_404(item_id)



# ── Reporte: Ocupabilidad ────────────────────────────────────────────────

def parse_beds_payload(value):
    """
    Convierte texto/JSON de camas habilitadas por fecha a dict {date: camas}.
    Acepta formatos:
      - JSON: {"2026-05-20": 1160}
      - líneas: 2026-05-20=1160, 20/05/2026:1160, 20-05-26;1160
    """
    if not value:
        return {}
    if isinstance(value, dict):
        raw_items = value.items()
    else:
        text = str(value).strip()
        if not text:
            return {}
        try:
            parsed = json.loads(text)
            if isinstance(parsed, dict):
                raw_items = parsed.items()
            else:
                raw_items = []
        except Exception:
            pairs = []
            for line in re.split(r"[\n,]+", text):
                line = line.strip()
                if not line:
                    continue
                if '=' in line:
                    left, right = line.split('=', 1)
                elif ':' in line:
                    left, right = line.split(':', 1)
                elif ';' in line:
                    left, right = line.split(';', 1)
                else:
                    parts = line.split()
                    if len(parts) < 2:
                        continue
                    left, right = parts[0], parts[1]
                pairs.append((left.strip(), right.strip()))
            raw_items = pairs

    out = {}
    for key, val in raw_items:
        d = parse_date(key)
        if not d:
            try:
                d = datetime.strptime(str(key)[:10], '%Y-%m-%d').date()
            except Exception:
                d = None
        if not d:
            continue
        out[d] = as_number(val)
    return out


def resolve_beds_by_date(dates, beds_default=0, beds_by_date=None):
    beds_by_date = beds_by_date or {}
    default = as_number(beds_default)
    return [float(beds_by_date.get(d, default) or 0) for d in dates]


def build_series_rows(order, dates, matrix):
    rows = []
    for g in order:
        vals = [float(matrix[g].get(d, 0) or 0) for d in dates]
        rows.append({'gerencia': g, 'values': vals, 'total': sum(vals)})
    return rows


def occupancy_report_data(start=None, end=None, curve_id=None, beds_default=0, beds_by_date=None):
    """
    Reporte de ocupabilidad.

    CURVA: planificación por gerencia desde la curva.
    RESERVAS: todas las filas del censo con ID, incluyendo Camas Ocupdas = 0.
    CENSO: ocupación real, suma de Camas Ocupdas.
    RESERVAS NO PRESENTES: Reservas - Censo.
    PORCENTAJES:
      - Ocupación = Censo / Camas habilitadas
      - Eficiencia = Censo / Reservas
      - Disponibilidad = (Camas habilitadas - Censo) / Camas habilitadas
    """
    if not start or not end:
        minmax = db.session.query(func.min(Censo.fecha_censo), func.max(Censo.fecha_censo)).one()
        start = start or minmax[0]
        end = end or minmax[1]
    if not start or not end:
        return {
            'start_date': '', 'end_date': '', 'dates': [], 'date_labels': [], 'sections': {},
            'totals': {}, 'percentages': {}, 'beds_by_date': [], 'conclusions': ['No hay censos importados.'],
            'curve': None,
        }

    dates = list(date_span(start, end))
    curve = CurvaVersion.query.get(curve_id) if curve_id else CurvaVersion.query.filter_by(is_active=True).order_by(CurvaVersion.uploaded_at.desc()).first()

    curva = defaultdict(lambda: defaultdict(float))
    reservas = defaultdict(lambda: defaultdict(float))
    censo = defaultdict(lambda: defaultdict(float))
    gerencias = set()

    if curve:
        rows_plan = db.session.query(
            CurvaItem.gerencia,
            CurvaDailyValue.fecha,
            func.sum(CurvaDailyValue.dotacion_planificada),
        ).join(
            CurvaDailyValue, CurvaDailyValue.curva_item_id == CurvaItem.id
        ).filter(
            CurvaItem.curva_version_id == curve.id,
            CurvaDailyValue.fecha >= start,
            CurvaDailyValue.fecha <= end,
        ).group_by(
            CurvaItem.gerencia, CurvaDailyValue.fecha
        ).all()
        for g, d, val in rows_plan:
            g = g or 'SIN GERENCIA'
            curva[g][d] += float(val or 0)
            gerencias.add(g)

    # Reservas: todas las filas con ID. Si ya fueron corregidas en Sin match,
    # curva_item_id apunta al ID correcto y, por tanto, a la gerencia correcta.
    rows_res = db.session.query(
        Censo.fecha_censo,
        func.coalesce(CurvaItem.gerencia, 'SIN MATCH EN CURVA'),
        func.count(CensoRecord.id),
    ).join(
        CensoRecord, CensoRecord.censo_id == Censo.id
    ).outerjoin(
        CurvaItem, CurvaItem.id == CensoRecord.curva_item_id
    ).filter(
        Censo.fecha_censo >= start,
        Censo.fecha_censo <= end,
        CensoRecord.solicitud_id.isnot(None),
        CensoRecord.solicitud_id != '',
    ).group_by(
        Censo.fecha_censo, func.coalesce(CurvaItem.gerencia, 'SIN MATCH EN CURVA')
    ).all()
    for d, g, val in rows_res:
        reservas[g][d] += float(val or 0)
        gerencias.add(g)

    # Censo/Ocupación: suma de Camas Ocupadas.
    rows_occ = db.session.query(
        Censo.fecha_censo,
        func.coalesce(CurvaItem.gerencia, 'SIN MATCH EN CURVA'),
        func.sum(CensoRecord.camas_ocupadas),
    ).join(
        CensoRecord, CensoRecord.censo_id == Censo.id
    ).outerjoin(
        CurvaItem, CurvaItem.id == CensoRecord.curva_item_id
    ).filter(
        Censo.fecha_censo >= start,
        Censo.fecha_censo <= end,
        CensoRecord.solicitud_id.isnot(None),
        CensoRecord.solicitud_id != '',
    ).group_by(
        Censo.fecha_censo, func.coalesce(CurvaItem.gerencia, 'SIN MATCH EN CURVA')
    ).all()
    for d, g, val in rows_occ:
        censo[g][d] += float(val or 0)
        gerencias.add(g)

    order = [g for g in DEFAULT_GERENCIAS if g in gerencias]
    order += sorted([g for g in gerencias if g not in DEFAULT_GERENCIAS and g != 'SIN MATCH EN CURVA'])
    if 'SIN MATCH EN CURVA' in gerencias:
        order.append('SIN MATCH EN CURVA')

    no_presentes = defaultdict(lambda: defaultdict(float))
    for g in order:
        for d in dates:
            no_presentes[g][d] = max(float(reservas[g].get(d, 0) or 0) - float(censo[g].get(d, 0) or 0), 0)

    curva_rows = build_series_rows(order, dates, curva)
    reservas_rows = build_series_rows(order, dates, reservas)
    censo_rows = build_series_rows(order, dates, censo)
    no_presentes_rows = build_series_rows(order, dates, no_presentes)

    total_curva = [sum(curva[g].get(d, 0) for g in order) for d in dates]
    total_reservas = [sum(reservas[g].get(d, 0) for g in order) for d in dates]
    total_censo = [sum(censo[g].get(d, 0) for g in order) for d in dates]
    total_no_presentes = [sum(no_presentes[g].get(d, 0) for g in order) for d in dates]
    beds = resolve_beds_by_date(dates, beds_default, beds_by_date)

    ocupacion_pct = [(total_censo[i] / beds[i] * 100) if beds[i] else 0 for i in range(len(dates))]
    eficiencia_pct = [(total_censo[i] / total_reservas[i] * 100) if total_reservas[i] else 0 for i in range(len(dates))]
    disponibilidad_pct = [((beds[i] - total_censo[i]) / beds[i] * 100) if beds[i] else 0 for i in range(len(dates))]

    conclusions = []
    if dates:
        max_idx = max(range(len(dates)), key=lambda i: total_censo[i]) if total_censo else 0
        conclusions.append(f"Mayor ocupación: {dates[max_idx].strftime('%d/%m/%Y')} con {total_censo[max_idx]:.0f} camas ocupadas.")
    if sum(beds):
        conclusions.append(f"Ocupación acumulada: {(sum(total_censo) / sum(beds) * 100):.1f}% sobre camas habilitadas informadas.")
    if sum(total_reservas):
        conclusions.append(f"Eficiencia acumulada: {(sum(total_censo) / sum(total_reservas) * 100):.1f}% sobre reservas.")
    if sum(total_no_presentes):
        conclusions.append(f"Reservas no presentes acumuladas: {sum(total_no_presentes):.0f}.")
    no_match_total = 0
    for row in reservas_rows:
        if row['gerencia'] == 'SIN MATCH EN CURVA':
            no_match_total = row['total']
            break
    if no_match_total:
        conclusions.append(f"Reservas sin match en curva: {no_match_total:.0f}. Corrige estos IDs para mejorar la clasificación por gerencia.")

    return {
        'start_date': start.isoformat(),
        'end_date': end.isoformat(),
        'dates': [d.isoformat() for d in dates],
        'date_labels': [d.strftime('%d/%m/%y') for d in dates],
        'sections': {
            'curva': curva_rows,
            'reservas': reservas_rows,
            'censo': censo_rows,
            'no_presentes': no_presentes_rows,
        },
        'totals': {
            'curva': total_curva,
            'reservas': total_reservas,
            'censo': total_censo,
            'no_presentes': total_no_presentes,
            'curva_total': sum(total_curva),
            'reservas_total': sum(total_reservas),
            'censo_total': sum(total_censo),
            'no_presentes_total': sum(total_no_presentes),
            'beds_total': sum(beds),
        },
        'beds_by_date': beds,
        'percentages': {
            'ocupacion': ocupacion_pct,
            'eficiencia': eficiencia_pct,
            'disponibilidad': disponibilidad_pct,
        },
        'conclusions': conclusions,
        'curve': {'id': curve.id, 'name': curve.name} if curve else None,
    }


def occupancy_xlsx_fast(data, progress_callback=None):
    """Genera Excel profesional del reporte de ocupabilidad con tablas y gráficos."""
    if progress_callback:
        progress_callback('Preparando reporte de ocupabilidad...')

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    tmp_path = tmp.name
    tmp.close()

    workbook = xlsxwriter.Workbook(tmp_path, {
        'constant_memory': False,
        'strings_to_urls': False,
        'nan_inf_to_errors': True,
    })

    RED_DARK = '#8F1510'
    RED = '#B42318'
    RED_SOFT = '#FFF1F1'
    YELLOW = '#FFF200'
    ORANGE = '#F58220'
    BLUE = '#4472C4'
    GRAY = '#F2F4F7'
    DARK = '#1D2939'
    WHITE = '#FFFFFF'
    BORDER = '#8F1510'

    fmt_title = workbook.add_format({'bold': True, 'font_color': WHITE, 'font_size': 16, 'align': 'center', 'valign': 'vcenter', 'bg_color': RED_DARK, 'border': 1, 'border_color': BORDER})
    fmt_meta = workbook.add_format({'bold': True, 'font_color': DARK, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'bg_color': RED_SOFT, 'border': 1, 'border_color': BORDER})
    fmt_section = workbook.add_format({'bold': True, 'font_color': WHITE, 'bg_color': RED, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': BORDER})
    fmt_head = workbook.add_format({'bold': True, 'font_color': WHITE, 'bg_color': RED, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': BORDER})
    fmt_date = workbook.add_format({'bold': True, 'font_color': WHITE, 'bg_color': RED, 'align': 'center', 'valign': 'vcenter', 'rotation': 90, 'border': 1, 'border_color': BORDER})
    fmt_text = workbook.add_format({'font_size': 8, 'border': 1, 'border_color': BORDER})
    fmt_num = workbook.add_format({'font_size': 8, 'align': 'center', 'border': 1, 'border_color': BORDER, 'num_format': '#,##0'})
    fmt_total = workbook.add_format({'bold': True, 'font_color': WHITE, 'bg_color': RED, 'align': 'center', 'border': 1, 'border_color': BORDER, 'num_format': '#,##0'})
    fmt_total_y = workbook.add_format({'bold': True, 'font_color': '#000000', 'bg_color': YELLOW, 'align': 'center', 'border': 1, 'border_color': BORDER, 'num_format': '#,##0'})
    fmt_pct = workbook.add_format({'font_size': 8, 'align': 'center', 'border': 1, 'border_color': BORDER, 'num_format': '0%'})
    fmt_pct_total = workbook.add_format({'bold': True, 'font_color': WHITE, 'bg_color': RED, 'align': 'center', 'border': 1, 'border_color': BORDER, 'num_format': '0%'})
    fmt_kpi_label = workbook.add_format({'bold': True, 'font_color': WHITE, 'bg_color': RED_DARK, 'align': 'center', 'border': 1, 'border_color': BORDER})
    fmt_kpi_val = workbook.add_format({'bold': True, 'font_color': '#000000', 'bg_color': YELLOW, 'align': 'center', 'border': 1, 'border_color': BORDER, 'num_format': '#,##0'})

    ws = workbook.add_worksheet('Ocupabilidad')
    ws.hide_gridlines(2)
    ws.freeze_panes(8, 1)
    dates = data.get('date_labels', [])
    date_count = len(dates)
    last_col = max(date_count + 1, 8)
    ws.set_column(0, 0, 26)
    if date_count:
        ws.set_column(1, date_count, 7)
    ws.set_column(date_count + 1, date_count + 1, 14)

    ws.merge_range(0, 0, 0, last_col, 'REPORTE DE OCUPABILIDAD', fmt_title)
    curve_name = (data.get('curve') or {}).get('name', 'Sin curva seleccionada')
    ws.merge_range(1, 0, 1, last_col, f"Período: {data.get('start_date','')} al {data.get('end_date','')} | Curva: {curve_name} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", fmt_meta)

    kpis = [
        ('CURVA ACUM.', data.get('totals', {}).get('curva_total', 0)),
        ('RESERVAS ACUM.', data.get('totals', {}).get('reservas_total', 0)),
        ('CENSO ACUM.', data.get('totals', {}).get('censo_total', 0)),
        ('NO PRESENTES', data.get('totals', {}).get('no_presentes_total', 0)),
        ('CAMAS HAB.', data.get('totals', {}).get('beds_total', 0)),
    ]
    col = 0
    for label, value in kpis:
        ws.write(3, col, label, fmt_kpi_label)
        ws.write_number(4, col, float(value or 0), fmt_kpi_val)
        col += 2

    current_row = 7

    def write_section(title, rows, total_values, total_label, total_format=fmt_total):
        nonlocal current_row
        ws.merge_range(current_row, 0, current_row, date_count + 1, title, fmt_section)
        current_row += 1
        ws.write(current_row, 0, 'GERENCIAS', fmt_head)
        for i, label in enumerate(dates):
            ws.write(current_row, i + 1, label, fmt_date)
        ws.write(current_row, date_count + 1, 'TOTAL', fmt_head)
        current_row += 1
        for row in rows:
            ws.write(current_row, 0, row.get('gerencia', ''), fmt_text)
            for i, val in enumerate(row.get('values', [])):
                ws.write_number(current_row, i + 1, float(val or 0), fmt_num)
            ws.write_number(current_row, date_count + 1, float(row.get('total') or 0), fmt_total_y)
            current_row += 1
        ws.write(current_row, 0, total_label, total_format)
        for i, val in enumerate(total_values):
            ws.write_number(current_row, i + 1, float(val or 0), total_format)
        ws.write_number(current_row, date_count + 1, float(sum(total_values) or 0), total_format)
        current_row += 2

    sections = data.get('sections', {})
    totals = data.get('totals', {})
    write_section('CURVA', sections.get('curva', []), totals.get('curva', []), 'TOTAL GENERAL CURVA')
    write_section('RESERVAS', sections.get('reservas', []), totals.get('reservas', []), 'TOTAL RESERVAS')
    write_section('CENSO', sections.get('censo', []), totals.get('censo', []), 'TOTAL CENSO')
    write_section('RESERVAS NO PRESENTES', sections.get('no_presentes', []), totals.get('no_presentes', []), 'TOTAL RESERVAS NO PRESENTES')

    # Porcentajes relevantes
    ws.merge_range(current_row, 0, current_row, date_count, 'PORCENTAJES RELEVANTES', fmt_section)
    current_row += 1
    ws.write(current_row, 0, '', fmt_head)
    for i, label in enumerate(dates):
        ws.write(current_row, i + 1, label, fmt_date)
    current_row += 1
    pct_start_row = current_row
    pct_rows = [
        ('CAMAS HABILITADAS', data.get('beds_by_date', []), 'num'),
        ('OCUPACIÓN', data.get('percentages', {}).get('ocupacion', []), 'pct'),
        ('EFICIENCIA', data.get('percentages', {}).get('eficiencia', []), 'pct'),
        ('DISPONIBILIDAD', data.get('percentages', {}).get('disponibilidad', []), 'pct'),
    ]
    for label, values, kind in pct_rows:
        ws.write(current_row, 0, label, fmt_text)
        for i, val in enumerate(values):
            if kind == 'pct':
                ws.write_number(current_row, i + 1, float(val or 0) / 100, fmt_pct)
            else:
                ws.write_number(current_row, i + 1, float(val or 0), fmt_num)
        current_row += 1

    # Datos para gráficos
    chart_ws = workbook.add_worksheet('_Datos Graficos Ocupabilidad')
    chart_ws.hide()
    headers = ['Fecha', 'TOTAL GENERAL CURVA', 'TOTAL RESERVAS', 'TOTAL CENSO', 'TOTAL RESERVAS NO PRESENTES', 'CAMAS HABILITADAS', 'OCUPACIÓN', 'EFICIENCIA', 'DISPONIBILIDAD']
    for c, h in enumerate(headers):
        chart_ws.write(0, c, h)
    for r, label in enumerate(dates, start=1):
        chart_ws.write(r, 0, label)
        chart_ws.write_number(r, 1, float(totals.get('curva', [])[r-1] if r-1 < len(totals.get('curva', [])) else 0))
        chart_ws.write_number(r, 2, float(totals.get('reservas', [])[r-1] if r-1 < len(totals.get('reservas', [])) else 0))
        chart_ws.write_number(r, 3, float(totals.get('censo', [])[r-1] if r-1 < len(totals.get('censo', [])) else 0))
        chart_ws.write_number(r, 4, float(totals.get('no_presentes', [])[r-1] if r-1 < len(totals.get('no_presentes', [])) else 0))
        chart_ws.write_number(r, 5, float(data.get('beds_by_date', [])[r-1] if r-1 < len(data.get('beds_by_date', [])) else 0))
        chart_ws.write_number(r, 6, float(data.get('percentages', {}).get('ocupacion', [])[r-1] if r-1 < len(data.get('percentages', {}).get('ocupacion', [])) else 0) / 100)
        chart_ws.write_number(r, 7, float(data.get('percentages', {}).get('eficiencia', [])[r-1] if r-1 < len(data.get('percentages', {}).get('eficiencia', [])) else 0) / 100)
        chart_ws.write_number(r, 8, float(data.get('percentages', {}).get('disponibilidad', [])[r-1] if r-1 < len(data.get('percentages', {}).get('disponibilidad', [])) else 0) / 100)

    if date_count:
        chart1 = workbook.add_chart({'type': 'line'})
        series_info = [
            (1, 'TOTAL GENERAL CURVA', BLUE),
            (2, 'TOTAL RESERVAS', ORANGE),
            (3, 'TOTAL CENSO', '#A6A6A6'),
            (4, 'TOTAL RESERVAS NO PRESENTES', '#FFC000'),
            (5, 'CAMAS HABILITADAS', '#5B9BD5'),
        ]
        for col_idx, name, color in series_info:
            chart1.add_series({
                'name': ['_Datos Graficos Ocupabilidad', 0, col_idx],
                'categories': ['_Datos Graficos Ocupabilidad', 1, 0, date_count, 0],
                'values': ['_Datos Graficos Ocupabilidad', 1, col_idx, date_count, col_idx],
                'line': {'color': color, 'width': 2.25},
                'marker': {'type': 'circle', 'size': 4, 'border': {'color': color}, 'fill': {'color': color}},
                'data_labels': {'value': True, 'font': {'size': 8, 'bold': True, 'color': WHITE}, 'position': 'above'},
            })
        chart1.set_title({'name': 'RESUMEN GENERAL DE OCUPACIÓN', 'name_font': {'bold': True, 'color': WHITE, 'size': 11}})
        chart1.set_legend({'position': 'bottom', 'font': {'color': WHITE, 'size': 8}})
        chart1.set_chartarea({'fill': {'color': '#3A3A3A'}, 'border': {'none': True}})
        chart1.set_plotarea({'fill': {'color': '#4A4A4A'}, 'border': {'none': True}})
        chart1.set_x_axis({'label_position': 'low', 'num_font': {'color': WHITE, 'size': 8}, 'line': {'color': '#777777'}})
        chart1.set_y_axis({'major_gridlines': {'visible': True, 'line': {'color': '#666666'}}, 'num_font': {'color': WHITE, 'size': 8}, 'line': {'color': '#777777'}})
        chart1.set_size({'width': 900, 'height': 420})
        ws.insert_chart(current_row + 1, 0, chart1)

        chart2 = workbook.add_chart({'type': 'line'})
        pct_series = [(6, 'OCUPACIÓN', '#4472C4'), (7, 'EFICIENCIA', '#ED7D31'), (8, 'DISPONIBILIDAD', '#A5A5A5')]
        for col_idx, name, color in pct_series:
            chart2.add_series({
                'name': ['_Datos Graficos Ocupabilidad', 0, col_idx],
                'categories': ['_Datos Graficos Ocupabilidad', 1, 0, date_count, 0],
                'values': ['_Datos Graficos Ocupabilidad', 1, col_idx, date_count, col_idx],
                'line': {'color': color, 'width': 2.5},
                'marker': {'type': 'circle', 'size': 4, 'border': {'color': color}, 'fill': {'color': color}},
                'data_labels': {'value': True, 'num_format': '0%', 'font': {'size': 8, 'bold': True, 'color': WHITE}, 'position': 'above'},
            })
        chart2.set_title({'name': 'PORCENTAJES RELEVANTES', 'name_font': {'bold': True, 'color': WHITE, 'size': 11}})
        chart2.set_legend({'position': 'bottom', 'font': {'color': WHITE, 'size': 8}})
        chart2.set_chartarea({'fill': {'color': '#3A3A3A'}, 'border': {'none': True}})
        chart2.set_plotarea({'fill': {'color': '#4A4A4A'}, 'border': {'none': True}})
        chart2.set_x_axis({'label_position': 'low', 'num_font': {'color': WHITE, 'size': 8}, 'line': {'color': '#777777'}})
        chart2.set_y_axis({'num_format': '0%', 'major_gridlines': {'visible': True, 'line': {'color': '#666666'}}, 'num_font': {'color': WHITE, 'size': 8}, 'line': {'color': '#777777'}, 'min': 0})
        chart2.set_size({'width': 900, 'height': 380})
        ws.insert_chart(current_row + 23, 0, chart2)

    ws.set_landscape()
    ws.fit_to_pages(1, 0)
    ws.set_margins(left=0.3, right=0.3, top=0.4, bottom=0.4)

    workbook.close()
    with open(tmp_path, 'rb') as f:
        content = f.read()
    try:
        os.unlink(tmp_path)
    except OSError:
        pass
    return content


def format_ocupabilidad_filename():
    return f'ocupabilidad_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'


def build_report_export_job(app, job_id):
    """Genera exportaciones Excel en segundo plano."""
    with app.app_context():
        job = ExportJob.query.get(job_id)
        if not job:
            return
        try:
            job.status = 'running'
            job.started_at = now_utc()
            job.message = 'Preparando datos...'
            db.session.commit()

            params = json.loads(job.params_json or '{}')
            start = parse_date(params.get('start_date'))
            end = parse_date(params.get('end_date'))
            curve_id = int(params.get('curve_id')) if str(params.get('curve_id') or '').strip() else None

            def progress(message):
                job.message = message
                db.session.commit()

            if job.job_type.startswith('ocupabilidad'):
                beds_default = as_number(params.get('beds_default'))
                beds_map = parse_beds_payload(params.get('beds_by_date'))
                data = occupancy_report_data(start, end, curve_id, beds_default, beds_map)
                content = occupancy_xlsx_fast(data, progress_callback=progress)
                job.filename = job.filename or format_ocupabilidad_filename()
            elif job.job_type.startswith('area_report'):
                kind = params.get('kind') or job.job_type.replace('area_report_', '') or 'egp'
                if not start or not end:
                    raise ValueError('Exportación cancelada: debes indicar Desde y Hasta. El Excel de EGP/F&A solo usa el rango solicitado, no el histórico completo.')
                data = area_report_data(kind, start, end, curve_id)
                content = area_report_xlsx(data, progress_callback=progress)
                job.filename = job.filename or format_area_report_filename(kind)
            else:
                data = report_data(start, end, curve_id)
                content = report_xlsx_fast(data, progress_callback=progress)
                job.filename = job.filename or format_export_filename(params.get('export_mode') or 'gerencia')

            job.content = content
            job.content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            job.status = 'completed'
            job.message = 'Archivo Excel listo para descargar.'
            job.completed_at = now_utc()
            db.session.commit()
        except Exception as exc:
            db.session.rollback()
            job = ExportJob.query.get(job_id)
            if job:
                job.status = 'failed'
                job.message = f'Error al generar Excel: {exc}'
                job.completed_at = now_utc()
                db.session.commit()


# ─────────────────────────────────────────────────────────────────────────────
# Reportes por área/proyecto: EGP y F&A
# ─────────────────────────────────────────────────────────────────────────────

def area_report_config(kind):
    kind = (kind or '').lower().strip()
    if kind in {'fa', 'fya', 'f&a'}:
        return {
            'kind': 'fa',
            'title': 'F&A',
            'long_title': 'Reporte alojamiento F&A',
            'subtitle': 'IDs asociados a F&A, con planificación por curva, reservas y censo.',
            'patterns': ['%F&A%', '%F & A%', '%Fast%Agile%', '%Fast&Agile%', '%Fast and Agile%', '%Fast y Agile%'],
        }
    return {
        'kind': 'egp',
        'title': 'EGP',
        'long_title': 'Reporte alojamiento EGP',
        'subtitle': 'IDs asociados a Escondida Growth Program, con planificación por curva, reservas y censo.',
        'patterns': ['%EGP%', '%Escondida Growth Program%', '%Growth Program%'],
    }


def area_report_filter(kind):
    cfg = area_report_config(kind)
    conditions = [CurvaItem.area.ilike(pattern) for pattern in cfg['patterns']]
    if cfg['kind'] == 'egp':
        conditions.extend([CurvaItem.empresa.ilike('%EGP%'), CurvaItem.camp.ilike('%EGP%')])
    else:
        conditions.extend([CurvaItem.empresa.ilike('%F&A%'), CurvaItem.camp.ilike('%F&A%')])
    return or_(*conditions)


def empty_area_report_payload(cfg, start=None, end=None, message=None):
    """Payload vacío para EGP/F&A cuando falta rango, curva o datos."""
    start_label = start.isoformat() if start else ''
    end_label = end.isoformat() if end else ''
    return {
        'kind': cfg['kind'],
        'title': cfg['title'],
        'long_title': cfg['long_title'],
        'subtitle': cfg['subtitle'],
        'start_date': start_label,
        'end_date': end_label,
        'dates': [],
        'date_labels': [],
        'rows': [],
        'totals': {
            'plan': [],
            'reservas': [],
            'censo': [],
            'no_show_reservas': [],
            'grand_plan': 0,
            'grand_reservas': 0,
            'grand_censo': 0,
            'grand_no_show_reservas': 0,
            'cumplimiento': 0,
            'eficiencia': 0,
        },
        'conclusions': [message or 'Selecciona un rango Desde/Hasta para generar el reporte.'],
        'curve': None,
    }


def area_report_data(kind, start=None, end=None, curve_id=None):
    cfg = area_report_config(kind)

    # Regla operacional: los informes EGP/F&A no deben usar histórico implícito.
    # Siempre deben generarse con el rango indicado por pantalla/exportación.
    if not start or not end:
        return empty_area_report_payload(
            cfg,
            start,
            end,
            'Selecciona Desde y Hasta. Este informe no se genera con histórico automático.'
        )

    if end < start:
        start, end = end, start

    dates = list(date_span(start, end))
    curve = CurvaVersion.query.get(curve_id) if curve_id else CurvaVersion.query.filter_by(is_active=True).order_by(CurvaVersion.uploaded_at.desc()).first()
    if not curve:
        payload = empty_area_report_payload(
            cfg,
            start,
            end,
            'No existe una curva activa. Importa una curva o selecciona una versión.'
        )
        payload.update({
            'dates': [d.isoformat() for d in dates],
            'date_labels': [d.strftime('%d/%m/%y') for d in dates],
        })
        return payload

    items = CurvaItem.query.filter(
        CurvaItem.curva_version_id == curve.id,
        area_report_filter(cfg['kind']),
    ).order_by(CurvaItem.solicitud_id.asc()).all()
    item_ids = [i.id for i in items]
    date_index = {d: idx for idx, d in enumerate(dates)}

    planned = {i.id: [0.0 for _ in dates] for i in items}
    reservas = {i.id: [0.0 for _ in dates] for i in items}
    censo = {i.id: [0.0 for _ in dates] for i in items}

    if item_ids:
        rows_plan = db.session.query(
            CurvaDailyValue.curva_item_id,
            CurvaDailyValue.fecha,
            func.sum(CurvaDailyValue.dotacion_planificada),
        ).filter(
            CurvaDailyValue.curva_item_id.in_(item_ids),
            CurvaDailyValue.fecha >= start,
            CurvaDailyValue.fecha <= end,
        ).group_by(CurvaDailyValue.curva_item_id, CurvaDailyValue.fecha).all()
        for item_id, d, value in rows_plan:
            if d in date_index:
                planned[item_id][date_index[d]] = float(value or 0)

        rows_censo = db.session.query(
            CensoRecord.curva_item_id,
            Censo.fecha_censo,
            func.count(CensoRecord.id),
            func.sum(CensoRecord.camas_ocupadas),
        ).join(Censo, Censo.id == CensoRecord.censo_id).filter(
            CensoRecord.curva_item_id.in_(item_ids),
            Censo.fecha_censo >= start,
            Censo.fecha_censo <= end,
            CensoRecord.solicitud_id.isnot(None),
            CensoRecord.solicitud_id != '',
        ).group_by(CensoRecord.curva_item_id, Censo.fecha_censo).all()
        for item_id, d, reserve_count, occupied_count in rows_censo:
            if d in date_index:
                reservas[item_id][date_index[d]] = float(reserve_count or 0)
                censo[item_id][date_index[d]] = float(occupied_count or 0)

    rows = []
    totals_plan = [0.0 for _ in dates]
    totals_res = [0.0 for _ in dates]
    totals_censo = [0.0 for _ in dates]

    for item in items:
        pvals = planned.get(item.id, [0.0 for _ in dates])
        rvals = reservas.get(item.id, [0.0 for _ in dates])
        cvals = censo.get(item.id, [0.0 for _ in dates])
        for idx in range(len(dates)):
            totals_plan[idx] += pvals[idx]
            totals_res[idx] += rvals[idx]
            totals_censo[idx] += cvals[idx]
        total_plan = sum(pvals)
        total_res = sum(rvals)
        total_censo = sum(cvals)
        rows.append({
            'id': item.solicitud_id,
            'empresa': item.empresa or '',
            'area': item.area or '',
            'turno': item.turno or '',
            'tipo_contrato': item.tipo_contrato or '',
            'formato': item.formato or '',
            'camp': item.camp or '',
            'planned_values': pvals,
            'reservation_values': rvals,
            'census_values': cvals,
            'total_plan': total_plan,
            'total_reservas': total_res,
            'total_censo': total_censo,
            'no_show_reservas': total_res - total_censo,
            'cumplimiento': (total_censo / total_plan * 100) if total_plan else None,
        })

    totals_no_show = [totals_res[i] - totals_censo[i] for i in range(len(dates))]
    grand_plan = sum(totals_plan)
    grand_res = sum(totals_res)
    grand_censo = sum(totals_censo)
    grand_no_show = grand_res - grand_censo
    cumplimiento = (grand_censo / grand_plan * 100) if grand_plan else 0
    eficiencia = (grand_censo / grand_res * 100) if grand_res else 0

    conclusions = [
        f"Rango exportado/visualizado: {start.strftime('%d/%m/%Y')} al {end.strftime('%d/%m/%Y')}.",
        f"{cfg['title']}: {len(items)} ID(s) identificados en la curva {curve.name}.",
        f"Planificación acumulada del rango: {grand_plan:.0f}. Reservas del rango: {grand_res:.0f}. Censo del rango: {grand_censo:.0f}.",
        f"No show de reservas del rango: {grand_no_show:.0f}. Eficiencia censo/reserva: {eficiencia:.1f}%.",
    ]
    if grand_plan:
        conclusions.append(f"Cumplimiento del censo contra curva: {cumplimiento:.1f}%.")
    if dates and totals_censo:
        idx_max = max(range(len(totals_censo)), key=lambda idx: totals_censo[idx])
        conclusions.append(f"Día con mayor censo: {dates[idx_max].strftime('%d/%m/%Y')} ({totals_censo[idx_max]:.0f}).")
    if rows:
        top = max(rows, key=lambda r: r['total_censo'])
        conclusions.append(f"ID con mayor censo acumulado: {top['id']} - {top['empresa']} ({top['total_censo']:.0f}).")

    return {
        'kind': cfg['kind'], 'title': cfg['title'], 'long_title': cfg['long_title'], 'subtitle': cfg['subtitle'],
        'start_date': start.isoformat(), 'end_date': end.isoformat(),
        'dates': [d.isoformat() for d in dates],
        'date_labels': [d.strftime('%d/%m/%y') for d in dates],
        'rows': rows,
        'totals': {
            'plan': totals_plan, 'reservas': totals_res, 'censo': totals_censo,
            'no_show_reservas': totals_no_show, 'grand_plan': grand_plan,
            'grand_reservas': grand_res, 'grand_censo': grand_censo,
            'grand_no_show_reservas': grand_no_show, 'cumplimiento': cumplimiento,
            'eficiencia': eficiencia,
        },
        'conclusions': conclusions,
        'curve': {'id': curve.id, 'name': curve.name},
    }

def format_area_report_filename(kind):
    cfg = area_report_config(kind)
    return f"reporte_{cfg['kind']}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"


def area_report_xlsx(data, progress_callback=None):
    if progress_callback:
        progress_callback(f"Preparando Excel {data.get('title', '')}...")
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    tmp_path = tmp.name
    tmp.close()

    dates = data.get('date_labels', [])
    rows = data.get('rows', [])
    totals = data.get('totals', {})
    title = data.get('long_title') or data.get('title') or 'Reporte'

    workbook = xlsxwriter.Workbook(tmp_path, {'constant_memory': True})
    ws = workbook.add_worksheet('Reporte')
    data_ws = workbook.add_worksheet('_Datos Graficos')
    data_ws.hide()

    red = '#B42318'; red_dark = '#7A1712'; yellow = '#FFF200'; gray = '#F2F4F7'; border = '#7A1712'
    fmt_title = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': red, 'font_size': 16, 'align': 'center', 'valign': 'vcenter'})
    fmt_sub = workbook.add_format({'bold': True, 'font_color': '#111111', 'bg_color': gray, 'font_size': 9})
    fmt_head = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': red, 'border': 1, 'border_color': border, 'align': 'center', 'valign': 'vcenter'})
    fmt_section = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': red_dark, 'border': 1, 'border_color': border, 'align': 'left'})
    fmt_date = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': red, 'border': 1, 'border_color': border, 'align': 'center', 'valign': 'vcenter', 'rotation': 90})
    fmt_text = workbook.add_format({'border': 1, 'border_color': border, 'font_size': 8, 'valign': 'vcenter'})
    fmt_num = workbook.add_format({'border': 1, 'border_color': border, 'font_size': 8, 'num_format': '0', 'align': 'center'})
    fmt_total = workbook.add_format({'bold': True, 'border': 1, 'border_color': border, 'bg_color': yellow, 'font_size': 8, 'num_format': '0', 'align': 'center'})
    fmt_total_label = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': red, 'border': 1, 'border_color': border, 'font_size': 8})
    fmt_kpi_label = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': red_dark, 'border': 1, 'border_color': border, 'align': 'center'})
    fmt_kpi_val = workbook.add_format({'bold': True, 'bg_color': '#FFFFFF', 'border': 1, 'border_color': border, 'num_format': '#,##0', 'align': 'center'})
    fmt_pct = workbook.add_format({'bold': True, 'bg_color': '#FFFFFF', 'border': 1, 'border_color': border, 'num_format': '0.0%', 'align': 'center'})

    last_col = 7 + len(dates)
    ws.merge_range(0, 0, 0, max(last_col, 12), title.upper(), fmt_title)
    ws.write(1, 0, f"Período: {data.get('start_date','')} al {data.get('end_date','')}", fmt_sub)
    ws.write(2, 0, f"Curva: {(data.get('curve') or {}).get('name','Sin curva')}", fmt_sub)
    ws.write(3, 0, data.get('subtitle', ''), fmt_sub)

    kpis = [
        ('Plan acumulado', totals.get('grand_plan', 0)), ('Reservas acumuladas', totals.get('grand_reservas', 0)),
        ('Censo acumulado', totals.get('grand_censo', 0)), ('No show reservas', totals.get('grand_no_show_reservas', 0)),
        ('Cumplimiento', (totals.get('cumplimiento', 0) or 0) / 100), ('Eficiencia', (totals.get('eficiencia', 0) or 0) / 100),
    ]
    ws.write_row(5, 0, [k[0] for k in kpis], fmt_kpi_label)
    for idx, (_, val) in enumerate(kpis):
        ws.write(6, idx, val, fmt_pct if idx >= 4 else fmt_kpi_val)

    data_ws.write_row(0, 0, ['Fecha', 'Plan', 'Reservas', 'Censo', 'No show reservas'], fmt_head)
    for idx, label in enumerate(dates, 1):
        data_ws.write(idx, 0, label)
        data_ws.write(idx, 1, (totals.get('plan') or [])[idx - 1] if idx - 1 < len(totals.get('plan') or []) else 0)
        data_ws.write(idx, 2, (totals.get('reservas') or [])[idx - 1] if idx - 1 < len(totals.get('reservas') or []) else 0)
        data_ws.write(idx, 3, (totals.get('censo') or [])[idx - 1] if idx - 1 < len(totals.get('censo') or []) else 0)
        data_ws.write(idx, 4, (totals.get('no_show_reservas') or [])[idx - 1] if idx - 1 < len(totals.get('no_show_reservas') or []) else 0)

    if dates:
        chart = workbook.add_chart({'type': 'line'})
        chart.set_title({'name': f"Resumen {data.get('title','')}: Curva, reservas y censo"})
        for col, name, color in [(1, 'Plan', '#4472C4'), (2, 'Reservas', '#ED7D31'), (3, 'Censo', '#A5A5A5'), (4, 'No show reservas', '#FFC000')]:
            chart.add_series({'name': name, 'categories': ['_Datos Graficos', 1, 0, len(dates), 0], 'values': ['_Datos Graficos', 1, col, len(dates), col], 'line': {'color': color, 'width': 2.25}, 'marker': {'type': 'circle', 'size': 4}})
        chart.set_legend({'position': 'bottom'})
        chart.set_y_axis({'major_gridlines': {'visible': True, 'line': {'color': '#D9D9D9'}}})
        chart.set_size({'width': 850, 'height': 310})
        ws.insert_chart(8, 0, chart)

    def write_section(start_row, name, key, total_key):
        ws.merge_range(start_row, 0, start_row, max(7 + len(dates), 12), name, fmt_section)
        headers = ['ID', 'Empresa', 'Área', 'Turno', 'Tipo Contrato', 'Formato', 'Camp']
        for col, h in enumerate(headers): ws.write(start_row + 1, col, h, fmt_head)
        for idx, label in enumerate(dates): ws.write(start_row + 1, 7 + idx, label, fmt_date)
        ws.write(start_row + 1, 7 + len(dates), 'TOTAL', fmt_head)
        row_idx = start_row + 2
        for item in rows:
            for col, val in enumerate([item['id'], item['empresa'], item['area'], item['turno'], item['tipo_contrato'], item['formato'], item['camp']]): ws.write(row_idx, col, val, fmt_text)
            vals = item.get(key) or []
            for idx, val in enumerate(vals): ws.write(row_idx, 7 + idx, val, fmt_num)
            ws.write(row_idx, 7 + len(dates), item.get(total_key, 0), fmt_total)
            row_idx += 1
        ws.write(row_idx, 0, f'TOTAL {name}', fmt_total_label)
        for col in range(1, 7): ws.write(row_idx, col, '', fmt_total_label)
        total_by_date = totals.get({'CURVA': 'plan', 'RESERVAS': 'reservas', 'CENSO': 'censo'}[name], [])
        for idx, val in enumerate(total_by_date): ws.write(row_idx, 7 + idx, val, fmt_total)
        ws.write(row_idx, 7 + len(dates), sum(total_by_date), fmt_total)
        return row_idx + 3

    row = 27 if dates else 9
    row = write_section(row, 'CURVA', 'planned_values', 'total_plan')
    row = write_section(row, 'RESERVAS', 'reservation_values', 'total_reservas')
    row = write_section(row, 'CENSO', 'census_values', 'total_censo')

    ws.merge_range(row, 0, row, 7, 'RESUMEN POR ID', fmt_section)
    ws.write_row(row + 1, 0, ['ID', 'Empresa', 'Área', 'Total Plan', 'Total Reservas', 'Total Censo', 'No Show Reservas', 'Cumplimiento'], fmt_head)
    for ridx, item in enumerate(rows, row + 2):
        ws.write(ridx, 0, item['id'], fmt_text); ws.write(ridx, 1, item['empresa'], fmt_text); ws.write(ridx, 2, item['area'], fmt_text)
        ws.write(ridx, 3, item['total_plan'], fmt_num); ws.write(ridx, 4, item['total_reservas'], fmt_num); ws.write(ridx, 5, item['total_censo'], fmt_num); ws.write(ridx, 6, item['no_show_reservas'], fmt_num)
        ws.write(ridx, 7, '' if item.get('cumplimiento') is None else item['cumplimiento'] / 100, fmt_pct if item.get('cumplimiento') is not None else fmt_num)

    row += len(rows) + 4
    ws.merge_range(row, 0, row, 7, 'CONCLUSIONES', fmt_section)
    for idx, conclusion in enumerate(data.get('conclusions') or [], row + 1): ws.write(idx, 0, f'• {conclusion}', fmt_text)

    ws.freeze_panes(7, 7)
    ws.set_column(0, 0, 14); ws.set_column(1, 1, 24); ws.set_column(2, 2, 42); ws.set_column(3, 6, 13); ws.set_column(7, 7 + len(dates), 5)
    ws.set_landscape(); ws.fit_to_pages(1, 0); ws.set_margins(left=0.25, right=0.25, top=0.45, bottom=0.45)

    workbook.close()
    with open(tmp_path, 'rb') as fh: content = fh.read()
    os.unlink(tmp_path)
    return content

def parse_arg(name):
    v = (request.args.get(name) or '').strip()
    if not v:
        return None
    return parse_date(v)


def parse_arg_any(*names):
    """Lee fechas aceptando nombres nuevos y antiguos de filtros."""
    for name in names:
        value = (request.args.get(name) or '').strip()
        if value:
            return parse_date(value)
    return None


def register_routes(app):
    @app.route('/')
    def dashboard():
        active=CurvaVersion.query.filter_by(is_active=True).order_by(CurvaVersion.uploaded_at.desc()).first()
        latest=Censo.query.order_by(Censo.fecha_censo.desc()).first()
        return render_template('dashboard.html', active=active, latest=latest, total_censos=Censo.query.count(), total_files=UploadedFile.query.count(), total_unmatched=db.session.query(func.sum(Censo.unmatched_count)).scalar() or 0)

    @app.route('/imports')
    def imports_page():
        return render_template('imports.html', files=UploadedFile.query.order_by(UploadedFile.uploaded_at.desc()).all(), curves=CurvaVersion.query.order_by(CurvaVersion.uploaded_at.desc()).all())

    @app.post('/api/import/curva')
    def upload_curva():
        f=request.files.get('file')
        if not f or not f.filename: flash('Selecciona un archivo de curva.', 'danger'); return redirect(url_for('imports_page'))
        try:
            v=import_curva(f, request.form.get('sheet_name') or 'Fcst_5400', request.form.get('version_name') or None)
            flash(f'Curva importada: {v.total_items} IDs y {v.total_daily_values} valores diarios.', 'success')
        except Exception as e:
            db.session.rollback(); flash(f'Error al importar curva: {e}', 'danger')
        return redirect(url_for('imports_page'))

    @app.post('/api/import/censo')
    def upload_censo():
        f=request.files.get('file')
        if not f or not f.filename: flash('Selecciona un archivo de censo.', 'danger'); return redirect(url_for('imports_page'))
        try:
            c=import_censo(f); flash(f'Censo {c.fecha_label} importado: {c.total_records} reservas, {int(c.total_occupied)} ocupadas, {c.matched_count} cruzadas, {c.unmatched_count} sin match.', 'success')
        except Exception as e:
            db.session.rollback(); flash(f'Error al importar censo: {e}', 'danger')
        return redirect(url_for('imports_page'))

    @app.post('/api/curves/<int:curve_id>/activate')
    def activate_curve(curve_id):
        CurvaVersion.query.update({CurvaVersion.is_active: False}); curve=CurvaVersion.query.get_or_404(curve_id); curve.is_active=True; db.session.commit(); flash(f'Curva activa: {curve.name}', 'success'); return redirect(url_for('imports_page'))


    @app.post('/api/files/<int:file_id>/delete')
    def delete_file(file_id):
        try:
            uploaded = delete_uploaded_file_data(file_id)
            flash(f'Archivo eliminado: {uploaded.filename}', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error al eliminar archivo: {e}', 'danger')
        return redirect(request.referrer or url_for('imports_page'))

    @app.route('/no-match')
    def no_match_page():
        censo_id = request.args.get('censo_id', type=int)
        curve_id = request.args.get('curve_id', type=int)
        search_text = request.args.get('q', '')
        limit = min(request.args.get('limit', 100, type=int) or 100, 1000)
        payload = no_match_payload(censo_id=censo_id, curve_id=curve_id, search_text=search_text, limit=limit)
        return render_template(
            'no_match.html',
            rows=payload['rows'],
            curve=payload['curve'],
            total=payload['total'],
            total_records=payload['total_records'],
            limit=payload['limit'],
            censos=Censo.query.order_by(Censo.fecha_censo.desc()).all(),
            curves=CurvaVersion.query.order_by(CurvaVersion.uploaded_at.desc()).all(),
            selected_censo_id=censo_id,
            selected_curve_id=curve_id,
            q=search_text,
        )

    @app.post('/api/no-match/<int:record_id>/apply')
    def apply_no_match(record_id):
        record = CensoRecord.query.get_or_404(record_id)
        item_id = request.form.get('curva_item_id', type=int)
        item = CurvaItem.query.get_or_404(item_id)
        updated_count, censo_ids = apply_correction_to_same_id(record, item)
        db.session.commit()
        flash(
            f'Corrección aplicada a {updated_count} registro(s) con ID censo {record.solicitud_id} → curva {item.solicitud_id} / {item.gerencia}.',
            'success'
        )
        return redirect(request.referrer or url_for('no_match_page'))

    @app.post('/api/no-match/<int:record_id>/manual')
    def manual_no_match(record_id):
        record = CensoRecord.query.get_or_404(record_id)
        curve_id = request.form.get('curve_id', type=int)
        manual_id = norm_id(request.form.get('manual_id'))
        curve = get_curve_for_matching(curve_id)
        if not curve:
            flash('No hay curva disponible para buscar el ID manual.', 'danger')
            return redirect(request.referrer or url_for('no_match_page'))
        item = find_curve_item_by_manual_id(curve, manual_id)
        if not item:
            flash(f'No se encontró el ID {manual_id} en la curva {curve.name}.', 'warning')
            return redirect(request.referrer or url_for('no_match_page'))
        updated_count, censo_ids = apply_correction_to_same_id(record, item)
        db.session.commit()
        flash(
            f'Corrección manual aplicada a {updated_count} registro(s): {record.solicitud_id} → {item.solicitud_id} / {item.gerencia}.',
            'success'
        )
        return redirect(request.referrer or url_for('no_match_page'))


    @app.route('/curva')
    def curva_page():
        payload = get_curve_items_payload(
            curve_id=request.args.get('curve_id', type=int),
            search_text=request.args.get('q', ''),
            gerencia=request.args.get('gerencia', ''),
            page=request.args.get('page', 1, type=int),
            per_page=request.args.get('per_page', 50, type=int),
        )
        return render_template('curva.html', **payload)

    @app.route('/curva/<int:item_id>/editar')
    def edit_curve_item_page(item_id):
        item = get_curve_item_for_edit(item_id)
        curves = CurvaVersion.query.order_by(CurvaVersion.is_active.desc(), CurvaVersion.uploaded_at.desc()).all()
        return render_template(
            'curva_item_form.html',
            curves=curves,
            record=None,
            item=item,
            selected_curve_id=item.curva_version_id,
            default_curve=item.version,
        )

    @app.route('/curva/nuevo')
    def new_curve_item_page():
        record_id = request.args.get('record_id', type=int)
        record = sample_record_context(record_id)
        selected_curve_id = request.args.get('curve_id', type=int)
        curves = CurvaVersion.query.order_by(CurvaVersion.is_active.desc(), CurvaVersion.uploaded_at.desc()).all()
        return render_template(
            'curva_item_form.html',
            curves=curves,
            record=record,
            item=None,
            selected_curve_id=selected_curve_id,
            default_curve=get_curve_for_matching(selected_curve_id),
        )

    @app.post('/api/curva/items')
    def create_curve_item():
        record_id = request.form.get('record_id', type=int)
        apply_to_unmatched = request.form.get('apply_to_unmatched') == '1'
        try:
            curve, item, created, created_values = create_or_update_curve_item_from_form(request.form)

            applied_count = 0
            if apply_to_unmatched:
                if record_id:
                    record = CensoRecord.query.get(record_id)
                    if record:
                        applied_count, _ = apply_correction_to_same_id(record, item)
                else:
                    applied_count, _ = apply_correction_to_solicitud_id(item.solicitud_id, item)

            db.session.commit()

            action = 'creado' if created else 'actualizado'
            msg = f'ID {item.solicitud_id} {action} en curva {curve.name} / {item.gerencia}.'
            if created_values:
                msg += f' Se cargaron {created_values} valores diarios de planificación.'
            if applied_count:
                msg += f' Además se corrigieron {applied_count} registro(s) sin match con ese ID.'
            flash(msg, 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error al guardar ID en curva: {e}', 'danger')

        next_url = request.form.get('next') or ''
        if next_url.startswith('/'):
            return redirect(next_url)
        return redirect(url_for('new_curve_item_page'))

    @app.post('/censos/reprocesar')
    def reprocess_censos_page():
        censos = Censo.query.order_by(Censo.fecha_censo.asc(), Censo.id.asc()).all()
        ok = 0
        errores = 0
        for censo in censos:
            try:
                reprocess_censo_from_original_file(censo)
                ok += 1
            except Exception as exc:
                errores += 1
                app.logger.exception('Error reprocesando censo %s: %s', censo.id, exc)
        db.session.commit()
        if errores:
            flash(f'Se reprocesaron {ok} censo(s), pero {errores} tuvieron error. Revisa logs de Render.', 'warning')
        else:
            flash(f'Se reprocesaron {ok} censo(s). Ahora Reservas incluye todas las filas con ID, aunque Camas Ocupdas sea 0.', 'success')
        return redirect(url_for('censos_page'))

    @app.route('/censos')
    def censos_page():
        return render_template('censos.html', censos=Censo.query.order_by(Censo.fecha_censo.desc()).all())



    @app.route('/reports/egp')
    def egp_report_page():
        return render_template('area_report.html', curves=CurvaVersion.query.order_by(CurvaVersion.uploaded_at.desc()).all(), kind='egp', title='EGP', subtitle='Escondida Growth Program')

    @app.route('/reports/fa')
    def fa_report_page():
        return render_template('area_report.html', curves=CurvaVersion.query.order_by(CurvaVersion.uploaded_at.desc()).all(), kind='fa', title='F&A', subtitle='Fast & Agile / F&A')

    @app.route('/api/reports/<kind>')
    def area_report_api(kind):
        if kind not in {'egp', 'fa'}:
            return jsonify({'error': 'Reporte no soportado'}), 404
        return jsonify(area_report_data(kind, parse_arg_any('start_date', 'date_from', 'desde', 'from'), parse_arg_any('end_date', 'date_to', 'hasta', 'to'), request.args.get('curve_id', type=int) or request.args.get('curveId', type=int) or request.args.get('curva_id', type=int)))

    @app.post('/api/reports/<kind>/export/start')
    def area_report_export_start(kind):
        if kind not in {'egp', 'fa'}:
            return jsonify({'error': 'Reporte no soportado'}), 404
        payload = request.get_json(silent=True) or request.form.to_dict() or request.args.to_dict()
        start_value = payload.get('start_date') or payload.get('startDate') or payload.get('desde') or payload.get('from')
        end_value = payload.get('end_date') or payload.get('endDate') or payload.get('hasta') or payload.get('to')
        start_parsed = parse_date(start_value)
        end_parsed = parse_date(end_value)
        if not start_parsed or not end_parsed:
            return jsonify({
                'ok': False,
                'error': 'Debes indicar Desde y Hasta antes de exportar. El Excel solo se genera con el rango seleccionado.'
            }), 400
        if end_parsed < start_parsed:
            start_parsed, end_parsed = end_parsed, start_parsed
        params = {
            'kind': kind,
            'start_date': start_parsed.isoformat(),
            'end_date': end_parsed.isoformat(),
            'curve_id': (payload.get('curve_id') or payload.get('curveId') or '').strip(),
        }
        job = ExportJob(
            job_type=f'area_report_{kind}',
            status='pending',
            message=f'Exportación {kind.upper()} en cola...',
            params_json=json.dumps(params),
            filename=format_area_report_filename(kind),
        )
        db.session.add(job); db.session.commit()
        thread = threading.Thread(target=build_report_export_job, args=(app, job.id), daemon=True); thread.start()
        return jsonify({'ok': True, 'job_id': job.id, 'status_url': url_for('export_job_status', job_id=job.id), 'download_url': url_for('export_job_download', job_id=job.id), 'page_url': url_for('export_job_page', job_id=job.id)}), 202

    @app.route('/api/reports/<kind>/export')
    def area_report_export_redirect(kind):
        if kind not in {'egp', 'fa'}:
            return jsonify({'error': 'Reporte no soportado'}), 404
        start_parsed = parse_arg_any('start_date', 'date_from', 'desde', 'from')
        end_parsed = parse_arg_any('end_date', 'date_to', 'hasta', 'to')
        if not start_parsed or not end_parsed:
            flash('Debes indicar Desde y Hasta antes de exportar. El Excel no se genera con rango histórico automático.', 'warning')
            return redirect(url_for('egp_report_page' if kind == 'egp' else 'fa_report_page'))
        if end_parsed < start_parsed:
            start_parsed, end_parsed = end_parsed, start_parsed
        params = {
            'kind': kind,
            'start_date': start_parsed.isoformat(),
            'end_date': end_parsed.isoformat(),
            'curve_id': (request.args.get('curve_id') or request.args.get('curveId') or request.args.get('curva_id') or '').strip(),
        }
        job = ExportJob(
            job_type=f'area_report_{kind}',
            status='pending',
            message=f'Exportación {kind.upper()} en cola...',
            params_json=json.dumps(params),
            filename=format_area_report_filename(kind),
        )
        db.session.add(job); db.session.commit()
        thread = threading.Thread(target=build_report_export_job, args=(app, job.id), daemon=True); thread.start()
        return redirect(url_for('export_job_page', job_id=job.id))

    @app.route('/reports/ocupabilidad')
    def ocupabilidad_page():
        return render_template('ocupabilidad.html', curves=CurvaVersion.query.order_by(CurvaVersion.uploaded_at.desc()).all())

    @app.route('/api/reports/ocupabilidad')
    def ocupabilidad_api():
        beds_default = request.args.get('beds_default', 0)
        beds_by_date = parse_beds_payload(request.args.get('beds_by_date'))
        return jsonify(occupancy_report_data(
            parse_arg('start_date'),
            parse_arg('end_date'),
            request.args.get('curve_id', type=int),
            beds_default=beds_default,
            beds_by_date=beds_by_date,
        ))

    @app.post('/api/reports/ocupabilidad/export/start')
    def ocupabilidad_export_start():
        payload = request.get_json(silent=True) or request.form.to_dict() or request.args.to_dict()
        params = {
            'start_date': (payload.get('start_date') or '').strip(),
            'end_date': (payload.get('end_date') or '').strip(),
            'curve_id': (payload.get('curve_id') or '').strip(),
            'beds_default': str(payload.get('beds_default') or '').strip(),
            'beds_by_date': payload.get('beds_by_date') or '',
        }
        job = ExportJob(
            job_type='ocupabilidad',
            status='pending',
            message='Exportación de ocupabilidad en cola...',
            params_json=json.dumps(params),
            filename=format_ocupabilidad_filename(),
        )
        db.session.add(job)
        db.session.commit()
        thread = threading.Thread(target=build_report_export_job, args=(app, job.id), daemon=True)
        thread.start()
        return jsonify({
            'ok': True,
            'job_id': job.id,
            'status_url': url_for('export_job_status', job_id=job.id),
            'download_url': url_for('export_job_download', job_id=job.id),
            'page_url': url_for('export_job_page', job_id=job.id),
        }), 202

    @app.route('/api/reports/ocupabilidad/export')
    def ocupabilidad_export():
        job = ExportJob(
            job_type='ocupabilidad',
            status='pending',
            message='Exportación de ocupabilidad en cola...',
            params_json=json.dumps(request.args.to_dict()),
            filename=format_ocupabilidad_filename(),
        )
        db.session.add(job)
        db.session.commit()
        thread = threading.Thread(target=build_report_export_job, args=(app, job.id), daemon=True)
        thread.start()
        return redirect(url_for('export_job_page', job_id=job.id), code=303)

    @app.route('/reports/dotacion-gerencia')
    def report_page():
        return render_template('report.html', curves=CurvaVersion.query.order_by(CurvaVersion.uploaded_at.desc()).all())

    @app.route('/api/reports/dotacion-gerencia')
    def report_api():
        return jsonify(report_data(parse_arg('start_date'), parse_arg('end_date'), request.args.get('curve_id', type=int)))

    def create_report_export_job_from_payload(payload):
        export_mode = (payload.get('export_mode') or payload.get('mode') or 'gerencia').strip().lower()
        if export_mode not in {'gerencia', 'censos', 'full'}:
            export_mode = 'gerencia'
        params = {
            'start_date': (payload.get('start_date') or '').strip(),
            'end_date': (payload.get('end_date') or '').strip(),
            'curve_id': (payload.get('curve_id') or '').strip(),
            'export_mode': export_mode,
        }
        job = ExportJob(
            job_type=f'dotacion_gerencia_{export_mode}',
            status='pending',
            message='Exportación en cola...',
            params_json=json.dumps(params),
            filename=format_export_filename(export_mode),
        )
        db.session.add(job)
        db.session.commit()

        thread = threading.Thread(target=build_report_export_job, args=(app, job.id), daemon=True)
        thread.start()
        return job

    @app.post('/api/reports/dotacion-gerencia/export/start')
    def report_export_start():
        payload = request.get_json(silent=True) or request.form.to_dict() or request.args.to_dict()
        job = create_report_export_job_from_payload(payload)

        return jsonify({
            'ok': True,
            'job_id': job.id,
            'status_url': url_for('export_job_status', job_id=job.id),
            'download_url': url_for('export_job_download', job_id=job.id),
            'page_url': url_for('export_job_page', job_id=job.id),
        }), 202

    @app.get('/api/reports/dotacion-gerencia/export/censos-csv')
    def report_export_censos_csv():
        """Descarga directa y liviana de censos acumulados como CSV, con IDs corregidos."""
        start = parse_arg('start_date')
        end = parse_arg('end_date')
        start, end = resolve_report_dates(start, end)
        start_text = start.strftime('%Y%m%d') if start else 'inicio'
        end_text = end.strftime('%Y%m%d') if end else 'fin'
        filename = f'censos_acumulados_{start_text}_{end_text}.csv'
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Cache-Control': 'no-store',
        }
        return Response(
            stream_with_context(iter_censos_acumulados_csv(start, end)),
            mimetype='text/csv; charset=utf-8',
            headers=headers,
        )

    @app.get('/exports/<int:job_id>')
    def export_job_page(job_id):
        job = ExportJob.query.get_or_404(job_id)
        return render_template('export_status.html', job=job)

    @app.get('/api/exports/<int:job_id>/status')
    def export_job_status(job_id):
        job = ExportJob.query.get_or_404(job_id)
        return jsonify({
            'id': job.id,
            'status': job.status,
            'message': job.message or '',
            'filename': job.filename or '',
            'created_at': job.created_at.isoformat() if job.created_at else None,
            'started_at': job.started_at.isoformat() if job.started_at else None,
            'completed_at': job.completed_at.isoformat() if job.completed_at else None,
            'download_url': url_for('export_job_download', job_id=job.id) if job.status == 'completed' else None,
        })

    @app.get('/api/exports/<int:job_id>/download')
    def export_job_download(job_id):
        job = ExportJob.query.get_or_404(job_id)
        if job.status != 'completed' or not job.content:
            return jsonify({'error': 'El archivo aún no está listo.'}), 409
        return send_file(
            BytesIO(job.content),
            as_attachment=True,
            download_name=job.filename or f'export_{job.id}.xlsx',
            mimetype=job.content_type or 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    @app.route('/api/reports/dotacion-gerencia/export')
    def report_export():
        # Compatibilidad con botones/enlaces antiguos o cacheados.
        export_mode = (request.args.get('export_mode') or request.args.get('mode') or 'gerencia').strip().lower()
        if export_mode == 'censos':
            return redirect(url_for('report_export_censos_csv', **request.args.to_dict()), code=303)
        job = create_report_export_job_from_payload(request.args.to_dict())
        return redirect(url_for('export_job_page', job_id=job.id), code=303)

app=create_app()
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.getenv('PORT',5000)), debug=os.getenv('FLASK_DEBUG')=='1')
