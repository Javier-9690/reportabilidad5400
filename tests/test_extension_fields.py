import csv
import html
import re
import unittest
from datetime import date, datetime
from io import BytesIO, StringIO
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

from tests.test_integration import app
from tests.test_edit_records import FormValues
from gestion5s import web


HEADERS = [
    "FECHA DE SOLICITUD", "ID", "EMPRESA", "CO", "GERENCIA", "CENTRO COSTOS",
    "CANT. CLIENTES", "TIPO DE SOLICITUD", "DESDE", "HASTA", "APROBADOR", "OBSERVACION",
]
NAMES = [
    "fecha_solicitud", "id_interno", "empresa", "co", "gerencia", "centro_costos",
    "cant_clientes", "tipo_solicitud", "desde", "hasta", "aprobador", "observacion",
]
FORM = dict(zip(NAMES, [
    "2026-09-08", "EXT-001", "Empresa & Servicios", "CO-12", "Gerencia Operaciones",
    "0005400", "0", "Excepción", "2026-09-10", "2026-09-12", "Aprobador de turno",
    "Autorizar <extensión> & coordinar alojamiento.\nConfirmar salida con recepción.",
]))


class ExtensionFieldsTest(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite://")
        web.Base.metadata.create_all(self.engine)
        self.sessions = sessionmaker(bind=self.engine)
        self.session_patch = patch.object(web, "SessionLocal", self.sessions)
        self.session_patch.start()
        self.client = app.test_client()

    def tearDown(self):
        self.session_patch.stop()
        self.engine.dispose()

    def record(self):
        with self.sessions() as db:
            record = db.query(web.ExtensionExcepcionEntry).one()
            return {column.name: getattr(record, column.name) for column in record.__table__.columns}

    def upload(self, workbook):
        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        response = self.client.post(
            "/gestion-5s/import/extensiones", data={"file": (payload, "extensiones.xlsx")},
            content_type="multipart/form-data", follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        return response.get_data(as_text=True)

    def assert_labels(self, page):
        labels = re.findall(r'<label[^>]*class="form-label[^\"]*"[^>]*>(.*?)</label>', page, re.S)
        self.assertEqual([html.unescape(label.strip()) for label in labels], HEADERS)
        values = FormValues(page).values
        self.assertEqual([name for name in values if name in NAMES], NAMES)
        self.assertNotIn("proyecto", values)

    def assert_listed_and_exported(self, expected):
        response = self.client.get("/gestion-5s/registros?vista=extensiones")
        self.assertEqual(response.status_code, 200)
        listing = response.get_data(as_text=True)
        self.assertEqual(re.findall(r"<th>(.*?)</th>", listing), ["Acciones", *HEADERS, "Registro"])
        cells = re.findall(r"<td(?:\s[^>]*)?>(.*?)</td>", listing, re.S)[1:13]
        self.assertEqual(
            [html.unescape(re.sub(r"<[^>]+>", "", cell)).strip() for cell in cells],
            [expected[name] if expected[name] != "" else "-" for name in NAMES],
        )
        self.assertIn("/gestion-5s/edit/extensiones/", listing)
        self.assertIn("/gestion-5s/delete/extensiones/", listing)
        self.assertIn("Eliminar todos", listing)

        response = self.client.get("/gestion-5s/download/extensiones.csv")
        self.assertEqual(response.status_code, 200)
        reader = csv.DictReader(StringIO(response.data.decode("utf-8-sig")))
        self.assertEqual(reader.fieldnames, HEADERS)
        self.assertEqual(list(reader), [dict(zip(HEADERS, (expected[name] for name in NAMES)))])

    def test_twelve_fields_save_display_export_and_edit_original_record(self):
        page = self.client.get("/gestion-5s/panel?tab=extensiones")
        self.assertEqual(page.status_code, 200)
        self.assert_labels(page.get_data(as_text=True))
        response = self.client.post("/gestion-5s/panel?tab=extensiones", data=FORM)
        self.assertEqual(response.status_code, 302)
        before = self.record()
        self.assertIsNone(before["proyecto"])
        for name in NAMES:
            self.assertEqual(str(before[name]), FORM[name])
        self.assert_listed_and_exported(FORM)

        route = f"/gestion-5s/edit/extensiones/{before['id']}"
        page = self.client.get(route).get_data(as_text=True)
        self.assert_labels(page)
        data = FormValues(page).values
        self.assertEqual({name: data[name] for name in NAMES}, FORM)
        updated = dict(FORM, centro_costos="CC-900", tipo_solicitud="Extensión autorizada", observacion="Confirmado.\nNueva observación.")
        data.update(updated)
        response = self.client.post(route, data=data, follow_redirects=True)
        self.assertEqual(response.status_code, 200)
        after = self.record()
        self.assertEqual(after["id"], before["id"])
        self.assertEqual(after["creado"], before["creado"])
        for name in NAMES:
            self.assertEqual(str(after[name]), updated[name])
        self.assert_listed_and_exported(updated)

        data = FormValues(self.client.get(route).get_data(as_text=True)).values
        data.update(centro_costos="", tipo_solicitud="")
        self.assertEqual(self.client.post(route, data=data).status_code, 302)
        self.assertIsNone(self.record()["centro_costos"])
        self.assertIsNone(self.record()["tipo_solicitud"])

    def test_downloaded_template_imports_all_fields_without_shifting_dates(self):
        response = self.client.get("/gestion-5s/template/extensiones.xlsx")
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.data))
        self.assertEqual([cell.value for cell in workbook.active[1]], HEADERS)
        values = dict(FORM, fecha_solicitud=datetime(2026, 9, 8), centro_costos=0,
                      cant_clientes=0, desde=date(2026, 9, 10), hasta=datetime(2026, 9, 12))
        workbook.active.append([values[name] for name in NAMES])
        self.assertIn("Importación de extensiones OK: 1 filas.", self.upload(workbook))
        expected = dict(FORM, centro_costos="0")
        record = self.record()
        for name in NAMES:
            self.assertEqual(str(record[name]), expected[name])
        self.assert_listed_and_exported(expected)

    def test_previous_template_is_rejected_without_importing_misaligned_data(self):
        workbook = Workbook()
        workbook.active.append([
            "FECHA_SOLICITUD", "ID", "EMPRESA", "CO", "GERENCIA", "PROYECTO",
            "CANT_CLIENTES", "DESDE", "HASTA", "APROBADOR", "OBSERVACION",
        ])
        workbook.active.append([
            "2026-09-08", "LEGACY-01", "Empresa", "CO", "Gerencia", "Proyecto antiguo",
            3, "2026-09-10", "2026-09-12", "Aprobador", "Observación antigua",
        ])
        page = self.upload(workbook)
        self.assertIn("Encabezados inválidos", page)
        self.assertIn("CENTRO COSTOS", page)
        self.assertIn("TIPO DE SOLICITUD", page)
        with self.sessions() as db:
            self.assertEqual(db.query(web.ExtensionExcepcionEntry).count(), 0)

    def test_existing_database_migration_preserves_history_and_allows_editing(self):
        with self.engine.begin() as conn:
            conn.execute(text("DROP TABLE extension_excepcion"))
            conn.execute(text("""
                CREATE TABLE extension_excepcion (
                    id INTEGER PRIMARY KEY, fecha_solicitud DATE NOT NULL,
                    id_interno VARCHAR(100), empresa VARCHAR(200), co VARCHAR(100),
                    gerencia VARCHAR(200), proyecto VARCHAR(200), cant_clientes INTEGER,
                    desde DATE, hasta DATE, aprobador VARCHAR(200), observacion TEXT,
                    creado DATETIME NOT NULL
                )
            """))
            conn.execute(text("""
                INSERT INTO extension_excepcion (id, fecha_solicitud, id_interno, proyecto, creado)
                VALUES (42, '2026-09-01', 'EXT-ANTERIOR', 'Proyecto histórico', '2026-09-01 10:00:00')
            """))
            before = dict(conn.execute(text("SELECT * FROM extension_excepcion")).mappings().one())
        web.ensure_extension_columns(self.engine)
        web.ensure_extension_columns(self.engine)
        with self.engine.connect() as conn:
            after = dict(conn.execute(text("SELECT * FROM extension_excepcion")).mappings().one())
        self.assertIsNone(after.pop("centro_costos"))
        self.assertIsNone(after.pop("tipo_solicitud"))
        self.assertEqual(after, before)

        route = "/gestion-5s/edit/extensiones/42"
        data = FormValues(self.client.get(route).get_data(as_text=True)).values
        self.assertEqual(data["id_interno"], "EXT-ANTERIOR")
        self.assertEqual(data["centro_costos"], "")
        self.assertEqual(data["tipo_solicitud"], "")
        data.update(centro_costos="CC-ANTERIOR", tipo_solicitud="Extensión", proyecto="No sobrescribir")
        self.assertEqual(self.client.post(route, data=data).status_code, 302)
        web.ensure_extension_columns(self.engine)
        record = self.record()
        self.assertEqual(record["id"], 42)
        self.assertEqual(record["creado"], datetime(2026, 9, 1, 10))
        self.assertEqual(record["proyecto"], "Proyecto histórico")
        self.assertEqual(record["centro_costos"], "CC-ANTERIOR")
        self.assertEqual(record["tipo_solicitud"], "Extensión")


if __name__ == "__main__":
    unittest.main()
