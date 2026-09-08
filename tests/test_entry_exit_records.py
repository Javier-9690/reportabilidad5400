import csv
import html
import json
import re
import unittest
from datetime import date, datetime, time
from io import BytesIO, StringIO
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import CALENDAR_MAC_1904, to_excel
from sqlalchemy import MetaData, create_engine, event, inspect, text
from sqlalchemy.orm import sessionmaker

from tests.test_integration import app
from tests.test_edit_records import FormValues
from gestion5s import web


HEADERS = [
    "FECHA INGRESO", "FECHA SALIDA", "HORA ENTRADA", "HORA SALIDA", "EMPRESA", "ID",
    "NOMBRE", "RUT", "TURNO", "PABELLON", "HABITACIÓN", "MOTIVO", "AUTORIZADO",
    "PENDULO", "N° TARJETA", "DEVOLUCIÓN",
]
NAMES = [
    "fecha_ingreso", "fecha_salida", "hora_entrada", "hora_salida", "empresa", "id_interno",
    "nombre", "rut", "turno", "pabellon", "habitacion", "motivo", "autorizado",
    "pendulo", "n_tarjeta", "devolucion",
]
FORM = dict(zip(NAMES, [
    "2026-09-08", "2026-09-09", "23:55:12", "00:15:30", "Empresa & Servicios", "000042",
    "Nombre de prueba", "00123456-K", "Noche", "Pabellón A", "001", "Alojamiento <temporal>.\nCoordinar devolución.",
    "Supervisión", "Sí", "000123", "Entregada en recepción",
]))


class EntryExitRecordsTest(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite://")
        web.Base.metadata.create_all(self.engine)
        self.sessions = sessionmaker(bind=self.engine)
        self.session_patch = patch.object(web, "SessionLocal", self.sessions)
        self.session_patch.start()
        self.client = app.test_client()

    def tearDown(self):
        self.session_patch.stop()
        self.engine.dispose()

    def records(self):
        with self.sessions() as db:
            return [{column.name: getattr(record, column.name) for column in record.__table__.columns}
                    for record in db.query(web.EntradaSalidaEntry).order_by(web.EntradaSalidaEntry.id).all()]

    def create(self, data):
        response = self.client.post("/gestion-5s/panel?tab=entradas_salidas", data=data)
        self.assertEqual(response.status_code, 302, response.get_data(as_text=True))
        return self.records()[-1]

    def upload(self, workbook):
        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        response = self.client.post(
            "/gestion-5s/import/entradas_salidas", data={"file": (payload, "entradas_salidas.xlsx")},
            content_type="multipart/form-data", follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        return response.get_data(as_text=True)

    def assert_labels(self, page):
        labels = re.findall(r'<label[^>]*class="form-label[^\"]*"[^>]*>(.*?)</label>', page, re.S)
        self.assertEqual([html.unescape(label.strip()) for label in labels], HEADERS)
        values = FormValues(page).values
        self.assertEqual([name for name in values if name in NAMES], NAMES)

    def csv_rows(self, **filters):
        response = self.client.get("/gestion-5s/download/entradas_salidas.csv", query_string=filters)
        self.assertEqual(response.status_code, 200)
        reader = csv.DictReader(StringIO(response.data.decode("utf-8-sig")))
        self.assertEqual(reader.fieldnames, HEADERS)
        return list(reader)

    def dashboard(self, **filters):
        response = self.client.get("/gestion-5s/dashboard", query_string=filters)
        self.assertEqual(response.status_code, 200)
        page = response.get_data(as_text=True)
        self.assertIn('id="entradasSalidasChart"', page)
        self.assertIn("Entradas y salidas por día", page)
        self.assertIn("Registros de entrada", page)
        self.assertIn("Registros de salida", page)
        labels = json.loads(re.search(r"const labels = (.*?);", page).group(1))
        series = json.loads(re.search(r"const series = (.*?);", page).group(1))
        return labels, series

    def test_all_sixteen_fields_round_trip_through_form_listing_csv_and_editor(self):
        page = self.client.get("/gestion-5s/panel?tab=entradas_salidas")
        self.assertEqual(page.status_code, 200)
        self.assert_labels(page.get_data(as_text=True))
        before = self.create(FORM)
        for name in NAMES:
            self.assertEqual(str(before[name]), FORM[name])
        page = self.client.get("/gestion-5s/registros?vista=entradas_salidas").get_data(as_text=True)
        self.assertEqual(re.findall(r"<th>(.*?)</th>", page), ["Acciones", *HEADERS, "Registro"])
        cells = re.findall(r"<td(?:\s[^>]*)?>(.*?)</td>", page, re.S)[1:17]
        self.assertEqual([html.unescape(re.sub(r"<[^>]+>", "", cell)).strip() for cell in cells],
                         [FORM[name] for name in NAMES])
        self.assertIn(f"/gestion-5s/edit/entradas_salidas/{before['id']}", page)
        self.assertIn(f"/gestion-5s/delete/entradas_salidas/{before['id']}", page)
        self.assertIn("Eliminar seleccionados", page)
        self.assertIn("Eliminar todos", page)
        self.assertEqual(self.csv_rows(), [dict(zip(HEADERS, FORM.values()))])

        route = f"/gestion-5s/edit/entradas_salidas/{before['id']}"
        page = self.client.get(route).get_data(as_text=True)
        self.assert_labels(page)
        data = FormValues(page).values
        self.assertEqual({name: data[name] for name in NAMES}, FORM)
        changed = dict(FORM, n_tarjeta="000999", devolucion="Pendiente", fecha_salida="2026-09-10", hora_salida="07:25:00")
        data.update(changed)
        self.assertEqual(self.client.post(route, data=data, follow_redirects=True).status_code, 200)
        after = self.records()[0]
        self.assertEqual(after["id"], before["id"])
        self.assertEqual(after["creado"], before["creado"])
        self.assertEqual(self.csv_rows(), [dict(zip(HEADERS, changed.values()))])

    def test_pending_departure_midnight_and_invalid_dates_or_times(self):
        record = self.create(dict(FORM, fecha_salida="", hora_entrada="00:00", hora_salida=""))
        self.assertEqual(record["hora_entrada"], time(0))
        self.assertIsNone(record["fecha_salida"])
        self.assertIsNone(record["hora_salida"])
        self.assertEqual(self.csv_rows()[0]["HORA ENTRADA"], "00:00:00")
        self.assertEqual(self.csv_rows()[0]["HORA SALIDA"], "")

        invalid = [
            {"hora_entrada": "25:30"},
            {"fecha_salida": "2026-09-08", "hora_salida": "22:00"},
            {"fecha_salida": "2026-09-07"},
        ]
        for changes in invalid:
            with self.subTest(changes=changes):
                response = self.client.post("/gestion-5s/panel?tab=entradas_salidas", data=dict(FORM, **changes))
                self.assertEqual(response.status_code, 422)
                self.assertIn("is-invalid", response.get_data(as_text=True))
                self.assertIn(FORM["nombre"], response.get_data(as_text=True))
                self.assertEqual(self.records(), [record])

        route = f"/gestion-5s/edit/entradas_salidas/{record['id']}"
        data = FormValues(self.client.get(route).get_data(as_text=True)).values
        data.update(fecha_salida="2026-09-07", hora_salida="23:59")
        self.assertEqual(self.client.post(route, data=data).status_code, 422)
        self.assertEqual(self.records(), [record])
        data.update(fecha_salida="2026-09-08", hora_salida="00:00")
        self.assertEqual(self.client.post(route, data=data).status_code, 302)
        self.assertEqual(self.records()[0]["hora_salida"], time(0))

    def test_template_import_accepts_excel_dates_times_and_preserves_identifiers(self):
        response = self.client.get("/gestion-5s/template/entradas_salidas.xlsx")
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.data))
        self.assertEqual([cell.value for cell in workbook.active[1]], HEADERS)
        self.assertEqual(workbook.active["F2"].number_format, "@")
        self.assertEqual(workbook.active["O2"].number_format, "@")
        data = dict(FORM, fecha_ingreso=datetime(2026, 9, 8), fecha_salida=date(2026, 9, 9),
                    hora_entrada=time(23, 55, 12), hora_salida=0)
        for column, name in enumerate(NAMES, 1):
            workbook.active.cell(2, column).value = data[name]
        self.assertIn("OK: 1 filas.", self.upload(workbook))
        first = self.records()[0]
        self.assertEqual(first["fecha_ingreso"], date(2026, 9, 8))
        self.assertEqual(first["fecha_salida"], date(2026, 9, 9))
        self.assertEqual(first["hora_salida"], time(0))
        for name in NAMES[4:]:
            self.assertEqual(first[name], FORM[name])

        workbook = Workbook()
        workbook.epoch = CALENDAR_MAC_1904
        workbook.active.append(HEADERS)
        data = dict(FORM, fecha_ingreso=to_excel(datetime(2026, 9, 8), epoch=workbook.epoch),
                    fecha_salida="09/09/2026", hora_entrada=0.5, hora_salida="07:45", id_interno="000043")
        workbook.active.append([data[name] for name in NAMES])
        self.assertIn("OK: 1 filas.", self.upload(workbook))
        second = self.records()[1]
        self.assertEqual(second["fecha_ingreso"], date(2026, 9, 8))
        self.assertEqual(second["hora_entrada"], time(12))
        self.assertEqual(second["hora_salida"], time(7, 45))
        self.assertEqual(second["id_interno"], "000043")
        self.assertEqual(second["n_tarjeta"], "000123")

    def test_invalid_import_reports_row_and_rolls_back_all_rows(self):
        for changes in ({"hora_entrada": "incorrecta"}, {"hora_salida": 1.5},
                        {"fecha_salida": "2026-09-07"}, {"fecha_salida": "fecha inválida"}):
            with self.subTest(changes=changes):
                workbook = Workbook()
                workbook.active.append(HEADERS)
                workbook.active.append([FORM[name] for name in NAMES])
                invalid = dict(FORM, **changes)
                workbook.active.append([invalid[name] for name in NAMES])
                page = self.upload(workbook)
                self.assertIn("Error importando entradas_salidas: Fila 3:", page)
                self.assertEqual(self.records(), [])

    def test_any_field_can_be_blank_and_rows_with_only_one_value_import(self):
        workbook = Workbook()
        workbook.active.append(HEADERS)
        workbook.active.append([None] * len(NAMES))
        for index, name in enumerate(NAMES):
            values = [None if column % 2 else "   " for column in range(len(NAMES))]
            values[index] = FORM[name]
            workbook.active.append(values)
        workbook.active.append(["   "] * len(NAMES))
        self.assertIn("OK: 16 filas.", self.upload(workbook))
        records = self.records()
        self.assertEqual(len(records), len(NAMES))
        for name, record in zip(NAMES, records):
            with self.subTest(only_field=name):
                self.assertEqual(str(record[name]), FORM[name])
                for other in NAMES:
                    if other != name:
                        self.assertIsNone(record[other])
                page = self.client.get(f"/gestion-5s/edit/entradas_salidas/{record['id']}")
                self.assertEqual(page.status_code, 200)
                values = FormValues(page.get_data(as_text=True)).values
                self.assertEqual(values[name], FORM[name])
                self.assertTrue(all(values[other] == "" for other in NAMES if other != name))
        exported = self.csv_rows()
        self.assertEqual(len(exported), 16)
        for name, header in zip(NAMES, HEADERS):
            self.assertIn({label: FORM[name] if label == header else "" for label in HEADERS}, exported)
        self.assertEqual(self.client.get("/gestion-5s/registros?vista=entradas_salidas").status_code, 200)

    def test_incomplete_records_can_be_completed_filtered_charted_and_deleted(self):
        incomplete = self.create({"nombre": "Solo nombre"})
        departure = self.create({"nombre": "Solo salida", "fecha_salida": "2026-09-09"})
        self.create({"nombre": "Fechado", "fecha_ingreso": "2026-09-08"})
        undated = self.create({"nombre": "Sin fecha permanente"})
        self.assertEqual(len(self.csv_rows()), 4)
        self.assertEqual(len(self.csv_rows(**{"from": "2026-09-08", "to": "2026-09-09"})), 1)
        labels, series = self.dashboard(**{"from": "2026-09-08", "to": "2026-09-09"})
        self.assertEqual(labels, ["2026-09-08", "2026-09-09"])
        self.assertEqual(series["entradas"], [1, 0])
        self.assertEqual(series["salidas"], [0, 1])

        route = f"/gestion-5s/edit/entradas_salidas/{incomplete['id']}"
        data = FormValues(self.client.get(route).get_data(as_text=True)).values
        data.update(fecha_ingreso="2026-09-08", fecha_salida="2026-09-09", n_tarjeta="00045")
        self.assertEqual(self.client.post(route, data=data).status_code, 302)
        after = next(record for record in self.records() if record["id"] == incomplete["id"])
        self.assertEqual(after["creado"], incomplete["creado"])
        self.assertIsNone(after["hora_entrada"])
        self.assertIsNone(after["hora_salida"])
        self.assertEqual(after["n_tarjeta"], "00045")

        response = self.client.post(f"/gestion-5s/delete/entradas_salidas/{departure['id']}", follow_redirects=True)
        self.assertEqual(response.status_code, 200)
        for mode, ids in (("selected", [str(undated["id"])]), ("all", [])):
            page = self.client.get("/gestion-5s/registros?vista=entradas_salidas").get_data(as_text=True)
            data = FormValues(page).values
            data.update(mode=mode, ids=ids)
            preview = self.client.post("/gestion-5s/delete/entradas_salidas/bulk/confirm", data=data)
            self.assertEqual(preview.status_code, 200)
            confirmation = FormValues(preview.get_data(as_text=True)).values
            response = self.client.post("/gestion-5s/delete/entradas_salidas/bulk", data=confirmation, follow_redirects=True)
            self.assertEqual(response.status_code, 200)
            self.assertNotIn(undated["id"], [record["id"] for record in self.records()])
        self.assertEqual(self.records(), [])

    def make_legacy_schema(self):
        old = web.EntradaSalidaEntry.__table__.to_metadata(MetaData())
        old.c.fecha_ingreso.nullable = False
        old.c.hora_entrada.nullable = False
        old.drop(self.engine)
        old.create(self.engine)
        before = self.create(FORM)
        with self.engine.begin() as conn:
            conn.execute(text("CREATE INDEX ix_entradas_empresa ON entradas_salidas (empresa)"))
            conn.execute(text("CREATE TABLE auditoria_entradas (registro_id INTEGER)"))
            conn.execute(text("CREATE TRIGGER auditar_entradas AFTER UPDATE ON entradas_salidas "
                              "BEGIN INSERT INTO auditoria_entradas (registro_id) VALUES (NEW.id); END"))
        return before

    def test_legacy_database_migration_preserves_rows_indexes_triggers_and_allows_nulls(self):
        before = self.make_legacy_schema()
        web.ensure_optional_entry_exit_fields(self.engine)
        web.ensure_optional_entry_exit_fields(self.engine)
        self.assertEqual(self.records(), [before])
        columns = {column["name"]: column for column in inspect(self.engine).get_columns("entradas_salidas")}
        for name in ("fecha_ingreso", "hora_entrada"):
            self.assertTrue(columns[name]["nullable"])
        self.assertEqual({index["name"] for index in inspect(self.engine).get_indexes("entradas_salidas")},
                         {"ix_entradas_empresa", "ix_entradas_salidas_fecha_ingreso", "ix_entradas_salidas_fecha_salida"})
        workbook = Workbook()
        workbook.active.append(HEADERS)
        workbook.active.append(["Importado incompleto" if name == "nombre" else None for name in NAMES])
        self.assertIn("OK: 1 filas.", self.upload(workbook))
        self.assertIsNone(self.records()[1]["fecha_ingreso"])
        self.assertIsNone(self.records()[1]["hora_entrada"])

        route = f"/gestion-5s/edit/entradas_salidas/{before['id']}"
        data = FormValues(self.client.get(route).get_data(as_text=True)).values
        data.update(fecha_ingreso="", hora_entrada="", fecha_salida="", hora_salida="")
        self.assertEqual(self.client.post(route, data=data).status_code, 302)
        web.ensure_optional_entry_exit_fields(self.engine)
        self.assertIsNone(self.records()[0]["fecha_ingreso"])
        self.assertEqual(self.records()[0]["creado"], before["creado"])
        with self.engine.connect() as conn:
            self.assertEqual(conn.scalar(text("SELECT registro_id FROM auditoria_entradas")), before["id"])

    def test_failed_sqlite_migration_restores_original_table_and_records(self):
        before = self.make_legacy_schema()

        def fail_rename(conn, cursor, statement, parameters, context, executemany):
            if statement.startswith("ALTER TABLE entradas_salidas_nullable RENAME"):
                raise RuntimeError("Fallo simulado antes de terminar la migración")

        event.listen(self.engine, "before_cursor_execute", fail_rename)
        try:
            with self.assertRaisesRegex(RuntimeError, "Fallo simulado"):
                web.ensure_optional_entry_exit_fields(self.engine)
        finally:
            event.remove(self.engine, "before_cursor_execute", fail_rename)
        self.assertEqual(self.records(), [before])
        self.assertNotIn("entradas_salidas_nullable", inspect(self.engine).get_table_names())
        columns = {column["name"]: column for column in inspect(self.engine).get_columns("entradas_salidas")}
        self.assertFalse(columns["fecha_ingreso"]["nullable"])
        self.assertFalse(columns["hora_entrada"]["nullable"])
        self.assertIn("ix_entradas_empresa", {index["name"] for index in inspect(self.engine).get_indexes("entradas_salidas")})
        with self.engine.connect() as conn:
            self.assertEqual(conn.scalar(text("SELECT COUNT(*) FROM sqlite_master WHERE name='auditar_entradas'")), 1)

    def test_dashboard_counts_each_event_on_its_date_and_respects_filters(self):
        empty = self.client.get("/gestion-5s/dashboard")
        self.assertEqual(empty.status_code, 200)
        for incoming, outgoing in (("2026-09-01", "2026-09-09"), ("2026-09-08", "2026-09-10"),
                                   ("2026-09-08", ""), ("2026-09-07", "2026-09-07")):
            self.create(dict(FORM, fecha_ingreso=incoming, fecha_salida=outgoing,
                             hora_entrada="08:00", hora_salida="09:00" if outgoing else ""))
        labels, series = self.dashboard(**{"from": "2026-09-08", "to": "2026-09-09"})
        self.assertEqual(labels, ["2026-09-08", "2026-09-09"])
        self.assertEqual(series["entradas"], [2, 0])
        self.assertEqual(series["salidas"], [0, 1])
        labels, series = self.dashboard(semana=89)
        self.assertEqual(labels, ["2026-09-07", "2026-09-08", "2026-09-09", "2026-09-10"])
        self.assertEqual(sum(series["entradas"]), 3)
        self.assertEqual(sum(series["salidas"]), 3)
        self.assertEqual(len(self.csv_rows(**{"from": "2026-09-08", "to": "2026-09-09"})), 2)
        self.assertEqual(len(self.csv_rows(semana=89)), 3)


if __name__ == "__main__":
    unittest.main()
