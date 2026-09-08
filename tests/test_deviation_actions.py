import csv
import html
import unittest
from io import BytesIO, StringIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, inspect, text

from tests.test_integration import app
from tests.test_edit_records import FormValues
from gestion5s.web import DesviacionEntry, SessionLocal, ensure_deviation_actions_column


class DeviationActionsTest(unittest.TestCase):
    def setUp(self):
        self.client = app.test_client()

    def tearDown(self):
        with SessionLocal() as db:
            db.query(DesviacionEntry).filter(DesviacionEntry.n_solicitud.startswith("ACCIONES-TEST-")).delete(synchronize_session=False)
            db.commit()

    def record(self, number):
        with SessionLocal() as db:
            record = db.query(DesviacionEntry).filter_by(n_solicitud=number).one()
            return {column.name: getattr(record, column.name) for column in record.__table__.columns}

    def upload(self, workbook):
        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        response = self.client.post(
            "/gestion-5s/import/desviaciones",
            data={"file": (payload, "desviaciones.xlsx")},
            content_type="multipart/form-data", follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertNotIn("Encabezados inválidos", response.get_data(as_text=True))
        self.assertNotIn("Error importando", response.get_data(as_text=True))

    def test_actions_save_display_export_and_edit_without_losing_other_fields(self):
        number = "ACCIONES-TEST-MANUAL"
        actions = "Señalizar el sector & avisar a mantención.\nVerificar <cierre> de la reparación."
        page = self.client.get("/gestion-5s/panel?tab=desviaciones").get_data(as_text=True)
        self.assertIn('name="acciones"', page)
        response = self.client.post("/gestion-5s/panel?tab=desviaciones", data={
            "n_solicitud": number, "fecha": "2026-09-08", "id": "ID-123",
            "descripcion_problema": "Daño en una puerta", "acciones": actions,
        })
        self.assertEqual(response.status_code, 302)
        before = self.record(number)
        self.assertEqual(before["acciones"], actions)

        listing = self.client.get("/gestion-5s/registros?vista=desviaciones").get_data(as_text=True)
        self.assertEqual(listing.count("<th>Acciones</th>"), 1)
        self.assertIn("<th>Opciones</th>", listing)
        self.assertIn(html.escape(actions), listing)
        download = self.client.get("/gestion-5s/download/desviaciones.csv")
        rows = csv.DictReader(StringIO(download.data.decode("utf-8-sig")))
        exported = next(row for row in rows if row["n_solicitud"] == number)
        self.assertEqual(exported["acciones"], actions)

        route = f"/gestion-5s/edit/desviaciones/{before['id']}"
        data = FormValues(self.client.get(route).get_data(as_text=True)).values
        self.assertEqual(data["acciones"], actions)
        data["acciones"] = "Reparación terminada."
        self.assertEqual(self.client.post(route, data=data).status_code, 302)
        after = self.record(number)
        self.assertEqual(after["acciones"], data["acciones"])
        for key in ("id", "creado", "fecha", "id_interno", "descripcion_problema"):
            self.assertEqual(after[key], before[key])
        self.assertIn(data["acciones"], self.client.get("/gestion-5s/registros?vista=desviaciones").get_data(as_text=True))

        data = FormValues(self.client.get(route).get_data(as_text=True)).values
        data["acciones"] = ""
        self.assertEqual(self.client.post(route, data=data).status_code, 302)
        self.assertIsNone(self.record(number)["acciones"])

    def test_downloaded_template_imports_actions_and_preserves_old_column_positions(self):
        response = self.client.get("/gestion-5s/template/desviaciones.xlsx")
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.data))
        headers = [cell.value for cell in workbook.active[1]]
        self.assertEqual(headers[-2:], ["CORREO_DESTINO", "ACCIONES"])
        self.assertEqual(len(headers), 14)
        values = {
            "N_SOLICITUD": "ACCIONES-TEST-EXCEL", "FECHA": "2026-09-08", "ID": "EXT-001",
            "EMPRESA_CONTRATISTA": "Contratista", "DESCRIPCION_PROBLEMA": "Desviación detectada",
            "TIPO_RIESGO": "Moderado", "TIPO_SOLICITUD": "Reparación", "PABELLON": "A",
            "HABITACION": "101", "VIA_SOLICITUD": "Correo", "QUIEN_INFORMA": "Recepción",
            "RIESGO_MATERIAL": "Puerta", "CORREO_DESTINO": "recepcion@ejemplo.cl",
            "ACCIONES": "Programar reparación.\nComprobar su ejecución.",
        }
        workbook.active.append([values[header] for header in headers])
        self.upload(workbook)
        record = self.record(values["N_SOLICITUD"])
        self.assertEqual(record["acciones"], values["ACCIONES"])
        self.assertEqual(record["correo_destino"], values["CORREO_DESTINO"])
        self.assertEqual(record["tipo_riesgo"], values["TIPO_RIESGO"])
        self.assertEqual(record["id_interno"], values["ID"])
        self.assertIn(values["ACCIONES"], self.client.get("/gestion-5s/registros?vista=desviaciones").get_data(as_text=True))

    def test_legacy_template_and_optional_blank_actions_remain_usable(self):
        response = self.client.get("/gestion-5s/template/desviaciones.xlsx")
        headers = [cell.value for cell in load_workbook(BytesIO(response.data)).active[1]]
        for suffix, selected_headers in (("ANTIGUA", headers[:-1]), ("VACIA", headers)):
            with self.subTest(template=suffix):
                number = "ACCIONES-TEST-" + suffix
                values = {"N_SOLICITUD": number, "FECHA": "2026-09-08", "CORREO_DESTINO": "recepcion@ejemplo.cl"}
                workbook = Workbook()
                workbook.active.append(selected_headers)
                workbook.active.append([values.get(header, "") for header in selected_headers])
                self.upload(workbook)
                record = self.record(number)
                self.assertFalse(record["acciones"])
                self.assertEqual(record["correo_destino"], values["CORREO_DESTINO"])
                self.assertEqual(self.client.get(f"/gestion-5s/edit/desviaciones/{record['id']}").status_code, 200)

    def test_existing_database_migration_keeps_rows_and_is_repeatable(self):
        engine = create_engine("sqlite://")
        try:
            with engine.begin() as conn:
                conn.execute(text("CREATE TABLE desviaciones (id INTEGER PRIMARY KEY, fecha DATE NOT NULL, descripcion_problema TEXT, creado DATETIME NOT NULL)"))
                conn.execute(text("INSERT INTO desviaciones VALUES (42, '2026-09-01', 'Registro existente', '2026-09-01 10:00:00')"))
                before = dict(conn.execute(text("SELECT * FROM desviaciones WHERE id=42")).mappings().one())
            ensure_deviation_actions_column(engine)
            ensure_deviation_actions_column(engine)
            self.assertIn("acciones", {column["name"] for column in inspect(engine).get_columns("desviaciones")})
            with engine.begin() as conn:
                after = dict(conn.execute(text("SELECT * FROM desviaciones WHERE id=42")).mappings().one())
                self.assertIsNone(after.pop("acciones"))
                self.assertEqual(after, before)
                conn.execute(text("UPDATE desviaciones SET acciones='Revisado' WHERE id=42"))
            ensure_deviation_actions_column(engine)
            with engine.connect() as conn:
                self.assertEqual(conn.scalar(text("SELECT acciones FROM desviaciones WHERE id=42")), "Revisado")
                self.assertEqual(conn.scalar(text("SELECT COUNT(*) FROM desviaciones")), 1)
        finally:
            engine.dispose()


if __name__ == "__main__":
    unittest.main()
