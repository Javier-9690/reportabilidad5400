import csv
import html
import json
import re
import unittest
from datetime import date, datetime
from io import BytesIO, StringIO
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import CALENDAR_MAC_1904, to_excel
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from tests.test_integration import app
from tests.test_edit_records import FormValues
from gestion5s import web


HEADERS = [
    "FECHA DE BLOQUEO", "HABITACIÓN", "OT", "EMPRESA", "ID", "MOTIVO", "COMUNICADO",
    "FECHA LIBERADA POR INGECLEAN", "FECHA LIBERADA POR FACILITY", "FECHA LIBERADA POR MANTENCION",
    "FECHA LIBERADA PROCESO INVESTIGACION", "OBSERVACIÓN",
]
NAMES = [
    "fecha_bloqueo", "habitacion", "ot", "empresa", "id_interno", "motivo", "comunicado",
    "fecha_liberada_ingeclean", "fecha_liberada_facility", "fecha_liberada_mantencion",
    "fecha_liberada_investigacion", "observacion",
]
FORM = dict(zip(NAMES, [
    "2026-09-09", "007", "000098", "Empresa & Servicios", "00042", "Revisión <eléctrica>.\nCoordinar trabajos.",
    "Recepción informada.\nConfirmado por supervisión.", "2026-09-10", "2026-09-12", "2026-09-11", "2026-09-13",
    "Comprobar cierre & limpieza.\nDejar constancia.",
]))


class BlockedRoomsTest(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite://")
        web.Base.metadata.create_all(self.engine)
        self.sessions = sessionmaker(bind=self.engine)
        self.session_patch = patch.object(web, "SessionLocal", self.sessions)
        self.session_patch.start()
        self.client = app.test_client()

    def tearDown(self):
        self.session_patch.stop()
        self.engine.dispose()

    def records(self):
        with self.sessions() as db:
            return [{column.name: getattr(record, column.name) for column in record.__table__.columns}
                    for record in db.query(web.HabitacionBloqueadaEntry).order_by(web.HabitacionBloqueadaEntry.id).all()]

    def create(self, data):
        response = self.client.post("/gestion-5s/panel?tab=habitaciones_bloqueadas", data=data)
        self.assertEqual(response.status_code, 302, response.get_data(as_text=True))
        return self.records()[-1]

    def upload(self, workbook):
        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        response = self.client.post("/gestion-5s/import/habitaciones_bloqueadas",
                                    data={"file": (payload, "habitaciones_bloqueadas.xlsx")},
                                    content_type="multipart/form-data", follow_redirects=True)
        self.assertEqual(response.status_code, 200)
        return response.get_data(as_text=True)

    def exported(self, **filters):
        response = self.client.get("/gestion-5s/download/habitaciones_bloqueadas.csv", query_string=filters)
        self.assertEqual(response.status_code, 200)
        reader = csv.DictReader(StringIO(response.data.decode("utf-8-sig")))
        self.assertEqual(reader.fieldnames, HEADERS)
        return list(reader)

    def assert_labels(self, page):
        labels = re.findall(r'<label[^>]*class="form-label[^\"]*"[^>]*>(.*?)</label>', page, re.S)
        self.assertEqual([html.unescape(label.strip()) for label in labels], HEADERS)
        self.assertEqual([name for name in FormValues(page).values if name in NAMES], NAMES)

    def test_all_twelve_fields_save_display_export_and_edit(self):
        response = self.client.get("/gestion-5s/panel?tab=habitaciones_bloqueadas")
        self.assertEqual(response.status_code, 200)
        self.assert_labels(response.get_data(as_text=True))
        before = self.create(FORM)
        for name in NAMES:
            self.assertEqual(str(before[name]), FORM[name])
        page = self.client.get("/gestion-5s/registros?vista=habitaciones_bloqueadas").get_data(as_text=True)
        self.assertEqual(re.findall(r"<th>(.*?)</th>", page), ["Acciones", *HEADERS, "Registro"])
        cells = re.findall(r"<td(?:\s[^>]*)?>(.*?)</td>", page, re.S)[1:13]
        self.assertEqual([html.unescape(re.sub(r"<[^>]+>", "", cell)).strip() for cell in cells],
                         [FORM[name] for name in NAMES])
        self.assertIn(f"/gestion-5s/edit/habitaciones_bloqueadas/{before['id']}", page)
        self.assertIn(f"/gestion-5s/delete/habitaciones_bloqueadas/{before['id']}", page)
        self.assertIn("Eliminar seleccionados", page)
        self.assertIn("Eliminar todos", page)
        self.assertEqual(self.exported(), [dict(zip(HEADERS, FORM.values()))])

        route = f"/gestion-5s/edit/habitaciones_bloqueadas/{before['id']}"
        page = self.client.get(route).get_data(as_text=True)
        self.assert_labels(page)
        data = FormValues(page).values
        self.assertEqual({name: data[name] for name in NAMES}, FORM)
        changed = dict(FORM, observacion="Seguimiento actualizado", fecha_liberada_facility="", ot="000099")
        data.update(changed)
        self.assertEqual(self.client.post(route, data=data, follow_redirects=True).status_code, 200)
        after = self.records()[0]
        self.assertEqual(after["id"], before["id"])
        self.assertEqual(after["creado"], before["creado"])
        self.assertIsNone(after["fecha_liberada_facility"])
        self.assertEqual(self.exported(), [dict(zip(HEADERS, changed.values()))])

    def test_template_import_reads_all_five_dates_and_preserves_text_identifiers(self):
        response = self.client.get("/gestion-5s/template/habitaciones_bloqueadas.xlsx")
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.data))
        self.assertEqual([cell.value for cell in workbook.active[1]], HEADERS)
        self.assertEqual(workbook.active["B2"].number_format, "@")
        self.assertEqual(workbook.active["C2"].number_format, "@")
        self.assertEqual(workbook.active["E2"].number_format, "@")
        workbook.epoch = CALENDAR_MAC_1904
        data = dict(FORM, fecha_bloqueo=datetime(2026, 9, 9), fecha_liberada_ingeclean=date(2026, 9, 10),
                    fecha_liberada_facility=datetime(2026, 9, 12), fecha_liberada_mantencion="11/09/2026",
                    fecha_liberada_investigacion=to_excel(datetime(2026, 9, 13), epoch=workbook.epoch))
        for column, name in enumerate(NAMES, 1):
            workbook.active.cell(2, column).value = data[name]
        self.assertIn("OK: 1 filas.", self.upload(workbook))
        record = self.records()[0]
        for name in NAMES:
            self.assertEqual(str(record[name]), FORM[name])

    def test_empty_fields_import_and_undated_records_remain_available(self):
        workbook = Workbook()
        workbook.active.append(HEADERS)
        workbook.active.append([None] * len(NAMES))
        for index, name in enumerate(NAMES):
            values = [None if column % 2 else "   " for column in range(len(NAMES))]
            values[index] = FORM[name]
            workbook.active.append(values)
        self.assertIn("OK: 12 filas.", self.upload(workbook))
        records = self.records()
        self.assertEqual(len(records), 12)
        for name, record in zip(NAMES, records):
            self.assertEqual(str(record[name]), FORM[name])
            self.assertTrue(all(record[other] is None for other in NAMES if other != name))
            self.assertEqual(self.client.get(f"/gestion-5s/edit/habitaciones_bloqueadas/{record['id']}").status_code, 200)
        self.assertEqual(len(self.exported()), 12)
        self.assertEqual(len(self.exported(**{"from": "2026-09-09", "to": "2026-09-09"})), 1)
        self.assertEqual(self.client.get("/gestion-5s/registros?vista=habitaciones_bloqueadas").status_code, 200)

    def test_invalid_dates_do_not_save_or_partially_import(self):
        before = self.create(FORM)
        changes = [{"fecha_bloqueo": "2026-02-30"}]
        changes += [{name: "2026-09-08"} for name in NAMES if name.startswith("fecha_liberada")]
        for change in changes:
            with self.subTest(change=change):
                invalid = dict(FORM, **change)
                response = self.client.post("/gestion-5s/panel?tab=habitaciones_bloqueadas", data=invalid)
                self.assertEqual(response.status_code, 422)
                self.assertIn("is-invalid", response.get_data(as_text=True))
                route = f"/gestion-5s/edit/habitaciones_bloqueadas/{before['id']}"
                data = FormValues(self.client.get(route).get_data(as_text=True)).values
                data.update(invalid)
                self.assertEqual(self.client.post(route, data=data).status_code, 422)
                workbook = Workbook()
                workbook.active.append(HEADERS)
                workbook.active.append([FORM[name] for name in NAMES])
                workbook.active.append([invalid[name] for name in NAMES])
                self.assertIn("Error importando habitaciones_bloqueadas: Fila 3:", self.upload(workbook))
                self.assertEqual(self.records(), [before])

    def test_dashboard_counts_entries_and_blocks_without_departure_or_release_dates(self):
        self.assertEqual(self.client.post("/gestion-5s/panel?tab=entradas_salidas", data={
            "nombre": "Solo salida", "fecha_salida": "2026-09-11",
        }).status_code, 302)
        page = self.client.get("/gestion-5s/dashboard").get_data(as_text=True)
        self.assertNotIn("const series =", page)
        self.assertEqual(self.client.post("/gestion-5s/panel?tab=entradas_salidas", data={
            "fecha_ingreso": "2026-09-09", "fecha_salida": "2026-09-12",
        }).status_code, 302)
        self.create({"fecha_bloqueo": "2026-08-31", "fecha_liberada_facility": "2026-09-09"})
        self.create({"fecha_bloqueo": "2026-09-09", "habitacion": "007", "fecha_liberada_facility": "2026-09-20"})
        self.create({"fecha_bloqueo": "2026-09-09", "habitacion": "007"})
        self.create({"fecha_bloqueo": "2026-09-10", "habitacion": "008"})
        self.create({"habitacion": "Sin fecha"})
        for filters in ({"from": "2026-09-09", "to": "2026-09-11"}, {"semana": 89}):
            with self.subTest(filters=filters):
                response = self.client.get("/gestion-5s/dashboard", query_string=filters)
                self.assertEqual(response.status_code, 200)
                page = response.get_data(as_text=True)
                self.assertIn('id="entradaHabitacionesChart"', page)
                self.assertIn('id="habitacionesBloqueadasChart"', page)
                self.assertIn("Entrada a habitaciones", page)
                self.assertIn("Habitaciones bloqueadas", page)
                self.assertNotIn("Registros de salida", page)
                labels = json.loads(re.search(r"const labels = (.*?);", page).group(1))
                series = json.loads(re.search(r"const series = (.*?);", page).group(1))
                self.assertNotIn("salidas", series)
                self.assertEqual(labels, ["2026-09-09", "2026-09-10"])
                self.assertEqual(series["entradas"], [1, 0])
                self.assertEqual(series["habitaciones_bloqueadas"], [2, 1])
                self.assertEqual(len(self.exported(**filters)), 3)


if __name__ == "__main__":
    unittest.main()
