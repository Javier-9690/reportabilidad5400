import csv
import html
import json
import re
import unittest
from datetime import datetime
from io import BytesIO, StringIO
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import CALENDAR_MAC_1904, to_excel
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from tests.test_integration import app
from tests.test_edit_records import FormValues
from gestion5s import web


SPECS = {
    "ordenamiento": {
        "title": "Ordenamiento",
        "headers": ["FECHA EJECUCIÓN", "EMPRESA", "HABITACIÓN", "NOMBRE", "RUT", "TURNO",
                    "REASIGNACIÓN", "MOTIVO CAMBIO", "PROCESADO", "PENDIENTE"],
        "form": {
            "fecha_ejecucion": "2026-09-09", "empresa": "Empresa & Servicios", "habitacion": "007",
            "nombre": "Persona de prueba", "rut": "00123456-7", "turno": "Noche", "reasignacion": "008",
            "motivo_cambio": "Cambio <solicitado>.\nCoordinar con recepción.",
            "procesado": "Confirmado por recepción", "pendiente": "Entrega de llave",
        },
        "dates": ["fecha_ejecucion"],
        "text_columns": [3, 5, 7],
    },
    "habitaciones_liberadas": {
        "title": "Habitaciones liberadas",
        "headers": ["Habitacion", "Empresa", "Entrega /devolucion", "Fecha de devolucion",
                    "Comentario", "USO A PARTIR DE", "Observación"],
        "form": {
            "habitacion": "007", "empresa": "Empresa & Servicios", "entrega_devolucion": "Devolución",
            "fecha_devolucion": "2026-09-09", "comentario": "Llaves <recibidas>.\nSin pendientes.",
            "uso_a_partir_de": "2026-09-10", "observacion": "Revisión & limpieza.\nLista para uso.",
        },
        "dates": ["fecha_devolucion", "uso_a_partir_de"],
        "text_columns": [1],
    },
}


class RoomOperationsTest(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite://")
        web.Base.metadata.create_all(self.engine)
        self.sessions = sessionmaker(bind=self.engine)
        self.session_patch = patch.object(web, "SessionLocal", self.sessions)
        self.session_patch.start()
        self.client = app.test_client()

    def tearDown(self):
        self.session_patch.stop()
        self.engine.dispose()

    def records(self, entity):
        with self.sessions() as db:
            model = web.ENTITY_MODEL[entity]
            return [{column.name: getattr(record, column.name) for column in record.__table__.columns}
                    for record in db.query(model).order_by(model.id).all()]

    def create(self, entity, data):
        response = self.client.post(f"/gestion-5s/panel?tab={entity}", data=data)
        self.assertEqual(response.status_code, 302, response.get_data(as_text=True))
        return self.records(entity)[-1]

    def upload(self, entity, workbook):
        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        response = self.client.post(f"/gestion-5s/import/{entity}",
                                    data={"file": (payload, f"{entity}.xlsx")},
                                    content_type="multipart/form-data", follow_redirects=True)
        self.assertEqual(response.status_code, 200)
        return response.get_data(as_text=True)

    def exported(self, entity, **filters):
        response = self.client.get(f"/gestion-5s/download/{entity}.csv", query_string=filters)
        self.assertEqual(response.status_code, 200)
        reader = csv.DictReader(StringIO(response.data.decode("utf-8-sig")))
        self.assertEqual(reader.fieldnames, SPECS[entity]["headers"])
        return list(reader)

    def assert_labels(self, entity, page):
        labels = re.findall(r'<label[^>]*class="form-label[^\"]*"[^>]*>(.*?)</label>', page, re.S)
        self.assertEqual([html.unescape(label.strip()) for label in labels], SPECS[entity]["headers"])
        names = list(SPECS[entity]["form"])
        self.assertEqual([name for name in FormValues(page).values if name in names], names)

    def test_all_columns_save_display_export_and_edit_in_requested_order(self):
        for entity, spec in SPECS.items():
            with self.subTest(entity=entity):
                form, headers = spec["form"], spec["headers"]
                response = self.client.get(f"/gestion-5s/panel?tab={entity}")
                self.assertEqual(response.status_code, 200)
                self.assert_labels(entity, response.get_data(as_text=True))
                before = self.create(entity, form)
                self.assertEqual({name: str(before[name]) for name in form}, form)
                page = self.client.get(f"/gestion-5s/registros?vista={entity}").get_data(as_text=True)
                self.assertEqual(re.findall(r"<th>(.*?)</th>", page), ["Acciones", *headers, "Registro"])
                cells = re.findall(r"<td(?:\s[^>]*)?>(.*?)</td>", page, re.S)[1:len(form) + 1]
                self.assertEqual([html.unescape(re.sub(r"<[^>]+>", "", cell)).strip() for cell in cells],
                                 list(form.values()))
                self.assertIn(f"/gestion-5s/edit/{entity}/{before['id']}", page)
                self.assertIn(f"/gestion-5s/delete/{entity}/{before['id']}", page)
                self.assertIn("Eliminar seleccionados", page)
                self.assertIn("Eliminar todos", page)
                self.assertEqual(self.exported(entity), [dict(zip(headers, form.values()))])

                route = f"/gestion-5s/edit/{entity}/{before['id']}"
                page = self.client.get(route).get_data(as_text=True)
                self.assert_labels(entity, page)
                data = FormValues(page).values
                self.assertEqual({name: data[name] for name in form}, form)
                changed = dict(form, habitacion="0009", **{name: "" for name in spec["dates"]})
                data.update(changed)
                self.assertEqual(self.client.post(route, data=data, follow_redirects=True).status_code, 200)
                after = self.records(entity)[0]
                self.assertEqual(after["id"], before["id"])
                self.assertEqual(after["creado"], before["creado"])
                self.assertTrue(all(after[name] is None for name in spec["dates"]))
                self.assertEqual(self.exported(entity), [dict(zip(headers, changed.values()))])

    def test_downloaded_templates_import_excel_and_text_dates_without_losing_codes(self):
        for entity, spec in SPECS.items():
            with self.subTest(entity=entity):
                response = self.client.get(f"/gestion-5s/template/{entity}.xlsx")
                self.assertEqual(response.status_code, 200)
                workbook = load_workbook(BytesIO(response.data))
                self.assertEqual([cell.value for cell in workbook.active[1]], spec["headers"])
                for column in spec["text_columns"]:
                    self.assertEqual(workbook.active.cell(2, column).number_format, "@")
                workbook.epoch = CALENDAR_MAC_1904
                for row, kind in enumerate(("excel", "serial", "text"), 2):
                    for column, (name, value) in enumerate(spec["form"].items(), 1):
                        cell = workbook.active.cell(row, column)
                        if name in spec["dates"]:
                            self.assertEqual(cell.number_format, "DD/MM/YYYY")
                            parsed = datetime.fromisoformat(value)
                            if kind == "excel":
                                value = parsed
                            elif kind == "serial":
                                value = to_excel(parsed, epoch=workbook.epoch)
                                cell.number_format = "General"
                            else:
                                value = parsed.strftime("%d/%m/%Y")
                        cell.value = value
                self.assertIn("OK: 3 filas.", self.upload(entity, workbook))
                for record in self.records(entity):
                    self.assertEqual({name: str(record[name]) for name in spec["form"]}, spec["form"])

    def test_any_single_field_can_be_imported_with_all_other_fields_empty(self):
        for entity, spec in SPECS.items():
            with self.subTest(entity=entity):
                names = list(spec["form"])
                response = self.client.get(f"/gestion-5s/template/{entity}.xlsx")
                workbook = load_workbook(BytesIO(response.data))
                for row, (name, value) in enumerate(spec["form"].items(), 3):
                    for column, other in enumerate(names, 1):
                        workbook.active.cell(row, column).value = value if other == name else ("  " if column % 2 else None)
                self.assertIn(f"OK: {len(names)} filas.", self.upload(entity, workbook))
                records = self.records(entity)
                self.assertEqual(len(records), len(names))
                for name, record in zip(names, records):
                    self.assertEqual(str(record[name]), spec["form"][name])
                    self.assertTrue(all(record[other] is None for other in names if other != name))
                    self.assertEqual(self.client.get(f"/gestion-5s/edit/{entity}/{record['id']}").status_code, 200)
                self.assertEqual(len(self.exported(entity)), len(names))
                self.assertEqual(len(self.exported(entity, **{"from": "2026-09-09", "to": "2026-09-09"})), 1)
                partial = self.create(entity, {"habitacion": "Sin fecha"})
                self.assertTrue(all(partial[name] is None for name in spec["dates"]))

    def test_invalid_populated_dates_and_empty_forms_do_not_write_partial_data(self):
        for entity, spec in SPECS.items():
            with self.subTest(entity=entity):
                before = self.create(entity, spec["form"])
                self.assertEqual(self.client.post(f"/gestion-5s/panel?tab={entity}", data={}).status_code, 422)
                route = f"/gestion-5s/edit/{entity}/{before['id']}"
                for name in spec["dates"]:
                    with self.subTest(field=name):
                        invalid = dict(spec["form"], **{name: "2026-02-30"})
                        response = self.client.post(f"/gestion-5s/panel?tab={entity}", data=invalid)
                        self.assertEqual(response.status_code, 422)
                        self.assertIn("is-invalid", response.get_data(as_text=True))
                        data = FormValues(self.client.get(route).get_data(as_text=True)).values
                        data.update(invalid)
                        self.assertEqual(self.client.post(route, data=data).status_code, 422)
                        workbook = Workbook()
                        workbook.active.append(spec["headers"])
                        workbook.active.append(list(spec["form"].values()))
                        workbook.active.append(list(invalid.values()))
                        self.assertIn(f"Error importando {entity}: Fila 3:", self.upload(entity, workbook))
                        self.assertEqual(self.records(entity), [before])

    def test_dashboard_and_filters_use_execution_and_return_dates_only(self):
        self.create("ordenamiento", {"habitacion": "Sin fecha"})
        self.create("habitaciones_liberadas", {"habitacion": "Sin fecha", "uso_a_partir_de": "2026-09-15"})
        self.assertNotIn("const series =", self.client.get("/gestion-5s/dashboard").get_data(as_text=True))
        for entity in SPECS:
            primary_date = SPECS[entity]["dates"][0]
            for day in ("2026-08-31", "2026-09-09", "2026-09-09", "2026-09-10"):
                data = {primary_date: day, "habitacion": "007"}
                if entity == "habitaciones_liberadas":
                    data["uso_a_partir_de"] = "2026-09-15"
                self.create(entity, data)
        for filters in ({"from": "2026-09-09", "to": "2026-09-11"}, {"semana": 89}):
            with self.subTest(filters=filters):
                response = self.client.get("/gestion-5s/dashboard", query_string=filters)
                self.assertEqual(response.status_code, 200)
                page = response.get_data(as_text=True)
                self.assertIn('id="ordenamientoChart"', page)
                self.assertIn('id="habitacionesLiberadasChart"', page)
                labels = json.loads(re.search(r"const labels = (.*?);", page).group(1))
                series = json.loads(re.search(r"const series = (.*?);", page).group(1))
                self.assertEqual(labels, ["2026-09-09", "2026-09-10"])
                self.assertNotIn("salidas", series)
                for entity, spec in SPECS.items():
                    self.assertEqual(series[entity], [2, 1])
                    self.assertRegex(page, rf'<div class="number[^\"]*">3</div>\s*<div class="label">{spec["title"]}</div>')
                    self.assertEqual(len(self.exported(entity, **filters)), 3)
                    self.assertEqual(len(self.exported(entity)), 5)
                    listing = self.client.get("/gestion-5s/registros", query_string={"vista": entity, **filters})
                    self.assertEqual(listing.status_code, 200)
                    self.assertNotIn("Sin fecha", listing.get_data(as_text=True))


if __name__ == "__main__":
    unittest.main()
