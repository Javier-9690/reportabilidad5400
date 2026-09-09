import os
import io
import csv
import re
import secrets
import hashlib
from urllib.parse import parse_qs, urlsplit
from sqlalchemy import inspect, text
from statistics import mean
from datetime import datetime, date, time, timedelta, timezone

from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, send_file, jsonify, abort, session
)

from gestion5s.editing import (
    EDIT_CONFIG, EXTENSION_FIELDS, ENTRY_EXIT_FIELDS, BLOCKED_ROOM_FIELDS,
    ORDERING_FIELDS, RELEASED_ROOM_FIELDS, OPTIONAL_RECORD_FIELDS,
    edit_fields, parse_edit_values, record_version,
)
from itsdangerous import BadData, URLSafeTimedSerializer

# ---------- BD ----------
from sqlalchemy import (
    create_engine, Column, Integer, String, Date, DateTime, Time, Float, Text, func, MetaData, Table
)
from sqlalchemy.schema import CreateTable
from sqlalchemy.orm import sessionmaker, declarative_base
from sqlalchemy.exc import SQLAlchemyError

# Excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils.datetime import from_excel


# -----------------------------------------------------------------------------
# App / Config
# -----------------------------------------------------------------------------
app = Flask(__name__, template_folder="templates", static_folder=None)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret")


def normalize_db_url(url: str) -> str:
    """Render / Supabase suelen dar postgres://; SQLAlchemy espera postgresql+psycopg2://
       y en entornos gestionados pedimos sslmode=require por defecto."""
    if url.startswith("postgres://"):
        url = url.replace("postgres://", "postgresql+psycopg2://", 1)
    elif url.startswith("postgresql://"):
        url = url.replace("postgresql://", "postgresql+psycopg2://", 1)
    if url.startswith("postgresql+") and "sslmode=" not in url:
        url += ("&" if "?" in url else "?") + "sslmode=require"
    return url


DATABASE_URL = os.environ.get("DATABASE_URL")
if not DATABASE_URL:
    # Desarrollo local: comparte el mismo archivo SQLite que Reportabilidad 5400.
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    instance_dir = os.path.join(project_root, "instance")
    os.makedirs(instance_dir, exist_ok=True)
    DATABASE_URL = "sqlite:///" + os.path.join(instance_dir, "dotacion_reportes.db")

ENGINE = create_engine(normalize_db_url(DATABASE_URL), pool_pre_ping=True)
SessionLocal = sessionmaker(bind=ENGINE, autocommit=False, autoflush=False)
Base = declarative_base()


# -----------------------------------------------------------------------------
# Semanas / Utilidades tiempo
# -----------------------------------------------------------------------------
def now_utc():
    return datetime.now(timezone.utc).replace(tzinfo=None)


WEEK_MAP = {
    42: ("2025-10-13", "2025-10-19"),
    43: ("2025-10-20", "2025-10-26"),
    44: ("2025-10-27", "2025-11-02"),
    45: ("2025-11-03", "2025-11-09"),
    46: ("2025-11-10", "2025-11-16"),
    47: ("2025-11-17", "2025-11-23"),
    48: ("2025-11-24", "2025-11-30"),
    49: ("2025-12-01", "2025-12-07"),
    50: ("2025-12-08", "2025-12-14"),
    51: ("2025-12-15", "2025-12-21"),
    52: ("2025-12-22", "2025-12-28"),
    53: ("2025-12-29", "2026-01-04"),
    54: ("2026-01-05", "2026-01-11"),
    55: ("2026-01-12", "2026-01-18"),
    56: ("2026-01-19", "2026-01-25"),
    57: ("2026-01-26", "2026-02-01"),
    58: ("2026-02-02", "2026-02-08"),
    59: ("2026-02-09", "2026-02-15"),
    60: ("2026-02-16", "2026-02-22"),
    61: ("2026-02-23", "2026-03-01"),
    62: ("2026-03-02", "2026-03-08"),
    63: ("2026-03-09", "2026-03-15"),
    64: ("2026-03-16", "2026-03-22"),
    65: ("2026-03-23", "2026-03-29"),
    66: ("2026-03-30", "2026-04-05"),
    67: ("2026-04-06", "2026-04-12"),
    68: ("2026-04-13", "2026-04-19"),
    69: ("2026-04-20", "2026-04-26"),
    70: ("2026-04-27", "2026-05-03"),
    71: ("2026-05-04", "2026-05-10"),
    72: ("2026-05-11", "2026-05-17"),
    73: ("2026-05-18", "2026-05-24"),
    74: ("2026-05-25", "2026-05-31"),
    75: ("2026-06-01", "2026-06-07"),
    76: ("2026-06-08", "2026-06-14"),
    77: ("2026-06-15", "2026-06-21"),
    78: ("2026-06-22", "2026-06-28"),
    79: ("2026-06-29", "2026-07-05"),
    80: ("2026-07-06", "2026-07-12"),
    81: ("2026-07-13", "2026-07-19"),
    82: ("2026-07-20", "2026-07-26"),
    83: ("2026-07-27", "2026-08-02"),
    84: ("2026-08-03", "2026-08-09"),
    85: ("2026-08-10", "2026-08-16"),
    86: ("2026-08-17", "2026-08-23"),
    87: ("2026-08-24", "2026-08-30"),
    88: ("2026-08-31", "2026-09-06"),
    89: ("2026-09-07", "2026-09-13"),
    90: ("2026-09-14", "2026-09-20"),
    91: ("2026-09-21", "2026-09-27"),
    92: ("2026-09-28", "2026-10-04"),
    93: ("2026-10-05", "2026-10-11"),
    94: ("2026-10-12", "2026-10-18"),
    95: ("2026-10-19", "2026-10-25"),
    96: ("2026-10-26", "2026-11-01"),
}

def week_range(week_number: int):
    if week_number not in WEEK_MAP:
        return (None, None)
    s, e = WEEK_MAP[week_number]
    return (date.fromisoformat(s), date.fromisoformat(e))


# Tiempo mm:ss
TIME_RE = re.compile(r"^\d{1,2}:\d{2}$")

def mmss_to_seconds(s: str) -> int:
    """Convierte formato mm:ss a segundos, maneja diferentes formatos."""
    if not s:
        return 0
    
    # Si es número (formato decimal de Excel)
    if isinstance(s, (int, float)):
        # Excel almacena el tiempo como fracción de día: 0.5 = 12 horas, 0.25 = 6 horas, etc.
        total_seconds = s * 86400  # 86400 segundos en un día
        return int(total_seconds)
    
    s = str(s).strip()
    
    # Si ya está en formato de segundos
    if s.isdigit():
        return int(s)
    
    # **CORRECCIÓN: Formato mm:ss donde mm puede ser > 59**
    if ":" in s:
        parts = s.split(":")
        if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
            return int(parts[0]) * 60 + int(parts[1])
    
    # Formato hh:mm:ss
    if s.count(":") == 2:
        parts = s.split(":")
        if all(part.isdigit() for part in parts):
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
    
    return 0

def excel_time_to_seconds(excel_time):
    """Convierte tiempo de Excel (decimal) a segundos."""
    if excel_time is None:
        return 0
    try:
        # Excel almacena tiempo como fracción de 24 horas
        return int(float(excel_time) * 86400)  # 24 * 60 * 60 = 86400
    except (ValueError, TypeError):
        return 0

def safe_convert_time(time_str):
    """Convierte strings de tiempo a segundos, manejando múltiples formatos."""
    if not time_str or str(time_str).strip() == "":
        return 0
    
    # Si es un objeto time, convertirlo a segundos
    if isinstance(time_str, time):
        return time_str.hour * 3600 + time_str.minute * 60 + time_str.second
    
    time_str = str(time_str).strip()
    
    # Si es un número (formato Excel)
    try:
        if "." in time_str or ":" not in time_str:
            return excel_time_to_seconds(time_str)
    except:
        pass
    
    # **CORRECCIÓN PRINCIPAL: Formato minutos:segundos donde minutos puede ser > 59**
    if ":" in time_str:
        parts = time_str.split(":")
        
        # Si hay exactamente 2 partes y ambas son números
        if len(parts) == 2 and all(part.strip().isdigit() for part in parts):
            minutes = int(parts[0])
            seconds = int(parts[1])
            return minutes * 60 + seconds
        
        # Formato hh:mm:ss
        elif len(parts) == 3 and all(part.strip().isdigit() for part in parts):
            hours = int(parts[0])
            minutes = int(parts[1])
            seconds = int(parts[2])
            return hours * 3600 + minutes * 60 + seconds
    
    # Intentar convertir como número de segundos
    try:
        return int(float(time_str))
    except (ValueError, TypeError):
        return 0

def seconds_to_mmss(x: int) -> str:
    m, s = divmod(max(0, int(x)), 60)
    return f"{m:02d}:{s:02d}"


# -----------------------------------------------------------------------------
# FUNCIONES ROBUSTAS PARA MANEJO DE FECHAS Y NORMALIZACIÓN
# -----------------------------------------------------------------------------
def normalize_header(header):
    """Normaliza encabezados removiendo tildes, espacios extra y caracteres especiales"""
    if header is None:
        return ""
    header = str(header).strip().upper()
    # Reemplazar caracteres con tildes
    replacements = {
        'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U',
        'À': 'A', 'È': 'E', 'Ì': 'I', 'Ò': 'O', 'Ù': 'U',
        'Ä': 'A', 'Ë': 'E', 'Ï': 'I', 'Ö': 'O', 'Ü': 'U',
        'Ñ': 'N'
    }
    for old, new in replacements.items():
        header = header.replace(old, new)
    # Remover espacios extra y caracteres no ASCII
    header = ''.join(char for char in header if ord(char) < 128)
    header = header.replace(' ', '_').replace('-', '_')
    # Remover múltiples guiones bajos consecutivos
    while '__' in header:
        header = header.replace('__', '_')
    return header

def safe_convert_date(date_str):
    """
    Convierte strings de fecha en objetos date, manejando múltiples formatos.
    """
    if not date_str or str(date_str).strip() == "":
        return None
    
    date_str = str(date_str).strip()
    
    # Si ya es un objeto date, retornarlo
    if isinstance(date_str, date):
        return date_str
    
    # Si es un objeto datetime, extraer la fecha
    if isinstance(date_str, datetime):
        return date_str.date()
    
    # Lista de formatos a probar
    formats = [
        '%Y-%m-%d',          # 2025-10-01
        '%d/%m/%Y',          # 01/10/2025
        '%d-%m-%Y',          # 01-10-2025
        '%Y/%m/%d',          # 2025/10/01
        '%Y-%m-%d %H:%M:%S', # 2025-10-01 00:00:00
        '%d/%m/%Y %H:%M:%S', # 01/10/2025 00:00:00
        '%d-%m-%Y %H:%M:%S', # 01-10-2025 00:00:00
        '%Y/%m/%d %H:%M:%S', # 2025/10/01 00:00:00
        '%Y-%m-%d %H:%M',    # 2025-10-01 00:00
        '%d/%m/%Y %H:%M',    # 01/10/2025 00:00
        '%d-%m-%Y %H:%M',    # 01-10-2025 00:00
        '%Y/%m/%d %H:%M',    # 2025/10/01 00:00
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    
    # Si ninguno funciona, intentar con fromisoformat para formatos estándar
    try:
        return date.fromisoformat(date_str.split(' ')[0])  # Tomar solo la parte de la fecha
    except ValueError:
        pass
    
    raise ValueError(f"No se pudo convertir la fecha: {date_str}")

def safe_convert_datetime(datetime_str):
    """
    Convierte strings de fecha/hora en objetos datetime, manejando múltiples formatos.
    """
    if not datetime_str or str(datetime_str).strip() == "":
        return None
    
    datetime_str = str(datetime_str).strip()
    
    # Si ya es un objeto datetime, retornarlo
    if isinstance(datetime_str, datetime):
        return datetime_str
    
    # Lista de formatos a probar
    formats = [
        '%Y-%m-%d %H:%M:%S',    # 2025-10-01 14:30:00
        '%d/%m/%Y %H:%M:%S',    # 01/10/2025 14:30:00
        '%d-%m-%Y %H:%M:%S',    # 01-10-2025 14:30:00
        '%Y/%m/%d %H:%M:%S',    # 2025/10/01 14:30:00
        '%Y-%m-%d %H:%M',       # 2025-10-01 14:30
        '%d/%m/%Y %H:%M',       # 01/10/2025 14:30
        '%d-%m-%Y %H:%M',       # 01-10-2025 14:30
        '%Y/%m/%d %H:%M',       # 2025/10/01 14:30
        '%Y-%m-%d',             # 2025-10-01 (hora 00:00:00)
        '%d/%m/%Y',             # 01/10/2025 (hora 00:00:00)
        '%d-%m-%Y',             # 01-10-2025 (hora 00:00:00)
        '%Y/%m/%d',             # 2025/10/01 (hora 00:00:00)
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(datetime_str, fmt)
        except ValueError:
            continue
    
    # Si ninguno funciona, intentar con fromisoformat
    try:
        if 'T' in datetime_str:
            return datetime.fromisoformat(datetime_str.replace('Z', '+00:00'))
        else:
            # Intentar agregar hora si solo es fecha
            return datetime.strptime(datetime_str + ' 00:00:00', '%Y-%m-%d %H:%M:%S')
    except ValueError:
        pass
    
    raise ValueError(f"No se pudo convertir la fecha/hora: {datetime_str}")

def safe_time_hhmm(val):
    """Convierte string de tiempo a objeto time, manejando múltiples formatos de hora"""
    s = ("" if val is None else str(val)).strip()
    if not s: 
        return time(0, 0)
    
    # Si ya es un objeto time, retornarlo
    if isinstance(val, time):
        return val
    
    # Intentar diferentes formatos de hora
    formats_to_try = [
        "%H:%M:%S",    # 22:40:00
        "%H:%M",       # 22:40
        "%I:%M:%S %p", # 10:40:00 PM (formato 12h)
        "%I:%M %p",    # 10:40 PM
        "%I:%M:%S%p",  # 10:40:00PM (sin espacio)
        "%I:%M%p",     # 10:40PM
    ]
    
    for fmt in formats_to_try:
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    
    # Si ninguno de los formatos estándar funciona, intentar analizar manualmente
    if ":" in s:
        parts = s.split(":")
        # Limpiar cada parte
        parts = [p.strip() for p in parts]
        
        # Tener entre 2 y 3 partes (HH:MM o HH:MM:SS)
        if 2 <= len(parts) <= 3:
            try:
                hour = int(parts[0])
                minute = int(parts[1])
                second = int(parts[2]) if len(parts) > 2 else 0
                
                # Validar rangos
                if 0 <= hour <= 23 and 0 <= minute <= 59 and 0 <= second <= 59:
                    return time(hour, minute, second)
            except (ValueError, IndexError):
                pass
    
    # Si todo falla, retornar hora por defecto
    return time(0, 0)


def record_excel_values(row, epoch, fields):
    """Normaliza celdas Excel para aplicar la misma validación que al formulario."""
    values = {}
    for field, value in zip(fields, row):
        name, label = field["name"], field["label"]
        try:
            if isinstance(value, str):
                value = value.strip()
            if value is None or value == "":
                values[name] = ""
            elif field["kind"] == "date":
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    value = from_excel(value, epoch=epoch)
                converted = safe_convert_date(value)
                if converted is None:
                    raise ValueError("Fecha inválida.")
                values[name] = converted.isoformat()
            elif field["kind"] == "time":
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    if not 0 <= value < 1:
                        raise ValueError("Usa una hora entre 00:00 y 23:59:59.")
                    value = from_excel(value, epoch=epoch)
                if isinstance(value, datetime):
                    value = value.time()
                values[name] = value.isoformat() if isinstance(value, time) else str(value).strip()
            else:
                values[name] = str(value).strip()
        except (ValueError, TypeError, OverflowError) as exc:
            raise ValueError(f"{label}: fecha u hora inválida.") from exc
    return values


# -----------------------------------------------------------------------------
# Modelos (tablas)
# -----------------------------------------------------------------------------
class CensusEntry(Base):
    __tablename__ = "census_entries"
    id = Column(Integer, primary_key=True)
    fecha = Column(Date, nullable=False)
    censo_dia = Column(Integer, nullable=False, default=0)
    censo_noche = Column(Integer, nullable=False, default=0)
    total = Column(Integer, nullable=False, default=0)
    creado = Column(DateTime, nullable=False, default=now_utc)

class EventSeguridad(Base):
    __tablename__ = "eventos_seguridad"
    id = Column(Integer, primary_key=True)
    fecha = Column(Date, nullable=False)
    horario = Column(String(50), nullable=False)
    que_ocurrio = Column(Text, nullable=False)
    nombre_afectado = Column(String(200), nullable=True)
    accion = Column(Text, nullable=True)
    creado = Column(DateTime, nullable=False, default=now_utc)

class DuplicidadEntry(Base):
    __tablename__ = "duplicidades"
    id = Column(Integer, primary_key=True)
    semana = Column(Integer, nullable=False)
    fecha = Column(Date, nullable=False)
    id_interno = Column(String(100), nullable=True)
    empresa_contratista = Column(String(200), nullable=True)
    descripcion_problema = Column(Text, nullable=True)
    tipo_riesgo = Column(String(200), nullable=True)
    pabellon = Column(String(100), nullable=True)
    habitacion = Column(String(100), nullable=True)
    ingresar_contacto = Column(String(200), nullable=True)
    nombre_usuario = Column(String(200), nullable=True)
    responsable = Column(String(200), nullable=True)
    estatus = Column(String(50), nullable=True)
    notificacion_usuario = Column(String(200), nullable=True)
    plan_accion = Column(Text, nullable=True)
    fecha_cierre = Column(Date, nullable=True)
    creado = Column(DateTime, nullable=False, default=now_utc)

class EncuestaEntry(Base):
    __tablename__ = "encuestas"
    id = Column(Integer, primary_key=True)
    fecha_hora = Column(DateTime, nullable=False)
    q1_respuesta = Column(Text, nullable=True)
    q1_puntaje = Column(Integer, nullable=True)
    q2_respuesta = Column(Text, nullable=True)
    q2_puntaje = Column(Integer, nullable=True)
    q3_respuesta = Column(Text, nullable=True)
    q3_puntaje = Column(Integer, nullable=True)
    q4_respuesta = Column(Text, nullable=True)
    q4_puntaje = Column(Integer, nullable=True)
    q5_respuesta = Column(Text, nullable=True)
    q5_puntaje = Column(Integer, nullable=True)
    total = Column(Integer, nullable=True)
    promedio = Column(Float, nullable=True)
    comentarios = Column(Text, nullable=True)
    creado = Column(DateTime, nullable=False, default=now_utc)

class AtencionEntry(Base):
    __tablename__ = "atencion_publico"
    id = Column(Integer, primary_key=True)
    fecha = Column(Date, nullable=False)
    tiempo_promedio_sec = Column(Integer, nullable=False, default=0)  # mm:ss -> seg
    cantidad = Column(Integer, nullable=False, default=0)
    creado = Column(DateTime, nullable=False, default=now_utc)

# --- Módulos agregados previamente ---
class RoboHurtoEntry(Base):
    __tablename__ = "robos_hurtos"
    id = Column(Integer, primary_key=True)
    fecha = Column(Date, nullable=False)
    hora = Column(Time, nullable=False)
    modulo = Column(String(100), nullable=True)
    habitacion = Column(String(100), nullable=True)
    empresa = Column(String(200), nullable=True)
    nombre_cliente = Column(String(200), nullable=True)
    rut = Column(String(50), nullable=True)
    medio_reclamo = Column(String(200), nullable=True)
    especies = Column(Text, nullable=True)
    observaciones = Column(Text, nullable=True)
    recepciona = Column(String(200), nullable=True)
    creado = Column(DateTime, nullable=False, default=now_utc)

class MiscelaneoEntry(Base):
    __tablename__ = "miscelaneo"
    id = Column(Integer, primary_key=True)
    ot = Column(String(100), nullable=True)
    division = Column(String(200), nullable=True)
    area = Column(String(200), nullable=True)
    lugar = Column(String(200), nullable=True)
    ubicacion = Column(String(200), nullable=True)
    disciplina = Column(String(200), nullable=True)
    especialidad = Column(String(200), nullable=True)
    falla = Column(String(200), nullable=True)
    empresa = Column(String(200), nullable=True)
    fecha_creacion = Column(Date, nullable=True)
    fecha_inicio = Column(Date, nullable=True)
    fecha_termino = Column(Date, nullable=True)
    fecha_aprobacion = Column(Date, nullable=True)
    estado = Column(String(100), nullable=True)
    comentario = Column(Text, nullable=True)
    creado = Column(DateTime, nullable=False, default=now_utc)

class DesviacionEntry(Base):
    __tablename__ = "desviaciones"
    id = Column(Integer, primary_key=True)
    n_solicitud = Column(String(100), nullable=True)
    fecha = Column(Date, nullable=False)
    id_interno = Column(String(100), nullable=True)
    empresa_contratista = Column(String(200), nullable=True)
    descripcion_problema = Column(Text, nullable=True)
    tipo_riesgo = Column(String(200), nullable=True)
    tipo_solicitud = Column(String(200), nullable=True)
    pabellon = Column(String(100), nullable=True)
    habitacion = Column(String(100), nullable=True)
    via_solicitud = Column(String(200), nullable=True)
    quien_informa = Column(String(200), nullable=True)
    riesgo_material = Column(String(200), nullable=True)
    correo_destino = Column(String(200), nullable=True)
    acciones = Column(Text, nullable=True)
    creado = Column(DateTime, nullable=False, default=now_utc)

class SolicitudOTEntry(Base):
    __tablename__ = "solicitudes_ot"
    id = Column(Integer, primary_key=True)
    n_solicitud = Column(String(100), nullable=True)
    descripcion_problema = Column(Text, nullable=True)
    tipo_solicitud = Column(String(200), nullable=True)
    modulo = Column(String(100), nullable=True)
    habitacion = Column(String(100), nullable=True)
    tipo_turno = Column(String(100), nullable=True)
    jornada = Column(String(100), nullable=True)
    via_solicitud = Column(String(200), nullable=True)
    correo_usuario = Column(String(200), nullable=True)
    tipo_tarea = Column(String(200), nullable=True)
    ot = Column(String(100), nullable=True)
    fecha_inicio = Column(Date, nullable=True)
    estado = Column(String(100), nullable=True)
    tiempo_respuesta_sec = Column(Integer, nullable=True)  # mm:ss -> seg
    satisfaccion_reclamo = Column(String(200), nullable=True)
    motivo = Column(String(200), nullable=True)
    observacion = Column(Text, nullable=True)
    creado = Column(DateTime, nullable=False, default=now_utc)

class ReclamoUsuarioEntry(Base):
    __tablename__ = "reclamos_usuarios"
    id = Column(Integer, primary_key=True)
    n_solicitud = Column(String(100), nullable=True)
    fecha = Column(Date, nullable=False)
    id_interno = Column(String(100), nullable=True)
    empresa_contratista = Column(String(200), nullable=True)
    descripcion_problema = Column(Text, nullable=True)
    tipo_solicitud = Column(String(200), nullable=True)
    pabellon = Column(String(100), nullable=True)
    habitacion = Column(String(100), nullable=True)
    via_solicitud = Column(String(200), nullable=True)
    ingresar_contacto = Column(String(200), nullable=True)
    nombre_usuario = Column(String(200), nullable=True)
    responsable = Column(String(200), nullable=True)
    estatus = Column(String(50), nullable=True)
    notificacion_usuario = Column(String(200), nullable=True)
    plan_accion = Column(Text, nullable=True)
    creado = Column(DateTime, nullable=False, default=now_utc)

# ---------------- NUEVOS 5 MÓDULOS ----------------
class ActivacionAlarmaEntry(Base):
    __tablename__ = "activacion_alarma"
    id = Column(Integer, primary_key=True)
    modulo = Column(String(100), nullable=True)
    n_habitacion = Column(String(100), nullable=True)
    nombre_recepcionista = Column(String(200), nullable=True)
    fecha = Column(Date, nullable=False)
    empresa = Column(String(200), nullable=True)
    id_interno = Column(String(100), nullable=True)
    co = Column(String(100), nullable=True)
    aviso_mantencion_h = Column(Float, nullable=True)
    llegada_mantencion_h = Column(Float, nullable=True)
    aviso_lider_h = Column(Float, nullable=True)
    llegada_lider_h = Column(Float, nullable=True)
    hora_reporte_salfa = Column(Time, nullable=True)
    tipo_evento = Column(String(200), nullable=True)
    tipo_actividad = Column(String(200), nullable=True)
    fecha_reporte = Column(Date, nullable=True)
    turno_recepcion_ingresos = Column(String(200), nullable=True)
    observaciones = Column(Text, nullable=True)
    creado = Column(DateTime, nullable=False, default=now_utc)

class ExtensionExcepcionEntry(Base):
    __tablename__ = "extension_excepcion"
    id = Column(Integer, primary_key=True)
    fecha_solicitud = Column(Date, nullable=False)
    id_interno = Column(String(100), nullable=True)
    empresa = Column(String(200), nullable=True)
    co = Column(String(100), nullable=True)
    gerencia = Column(String(200), nullable=True)
    # Se conserva el dato histórico; ya no forma parte del formulario ni la plantilla.
    proyecto = Column(String(200), nullable=True)
    centro_costos = Column(String(200), nullable=True)
    cant_clientes = Column(Integer, nullable=True)
    tipo_solicitud = Column(String(200), nullable=True)
    desde = Column(Date, nullable=True)
    hasta = Column(Date, nullable=True)
    aprobador = Column(String(200), nullable=True)
    observacion = Column(Text, nullable=True)
    creado = Column(DateTime, nullable=False, default=now_utc)

class OnboardingEntry(Base):
    __tablename__ = "onboarding"
    id = Column(Integer, primary_key=True)
    fecha_hora = Column(DateTime, nullable=False)
    nombre = Column(String(200), nullable=True)
    rut = Column(String(50), nullable=True)
    empresa = Column(String(200), nullable=True)
    id_interno = Column(String(100), nullable=True)
    archivo_pdf = Column(String(300), nullable=True)
    creado = Column(DateTime, nullable=False, default=now_utc)

class AperturaHabitacionEntry(Base):
    __tablename__ = "apertura_habitacion"
    id = Column(Integer, primary_key=True)
    fecha = Column(Date, nullable=False)
    habitacion = Column(String(100), nullable=True)
    hora = Column(Time, nullable=True)
    responsable = Column(String(200), nullable=True)
    estado_chapa = Column(Text, nullable=True)
    creado = Column(DateTime, nullable=False, default=now_utc)

class CumplimientoEECCEntry(Base):
    __tablename__ = "cumplimiento_eecc"
    id = Column(Integer, primary_key=True)
    empresa = Column(String(200), nullable=True)
    n_contrato = Column(String(100), nullable=True)
    co = Column(String(100), nullable=True)
    correo_electronico = Column(String(200), nullable=True)
    id_interno = Column(String(100), nullable=True)
    turno = Column(String(100), nullable=True)
    # fecha de negocio para filtros/dashboard
    fecha = Column(Date, nullable=False, index=True)
    creado = Column(DateTime, nullable=False, default=now_utc)


class EntradaSalidaEntry(Base):
    __tablename__ = "entradas_salidas"
    id = Column(Integer, primary_key=True)
    fecha_ingreso = Column(Date, nullable=True, index=True)
    fecha_salida = Column(Date, nullable=True, index=True)
    hora_entrada = Column(Time, nullable=True)
    hora_salida = Column(Time, nullable=True)
    empresa = Column(String(200), nullable=True)
    id_interno = Column(String(100), nullable=True)
    nombre = Column(String(200), nullable=True)
    rut = Column(String(50), nullable=True)
    turno = Column(String(100), nullable=True)
    pabellon = Column(String(100), nullable=True)
    habitacion = Column(String(100), nullable=True)
    motivo = Column(Text, nullable=True)
    autorizado = Column(String(200), nullable=True)
    pendulo = Column(String(100), nullable=True)
    n_tarjeta = Column(String(100), nullable=True)
    devolucion = Column(String(200), nullable=True)
    creado = Column(DateTime, nullable=False, default=now_utc)


class HabitacionBloqueadaEntry(Base):
    __tablename__ = "habitaciones_bloqueadas"
    id = Column(Integer, primary_key=True)
    fecha_bloqueo = Column(Date, nullable=True, index=True)
    habitacion = Column(String(100), nullable=True)
    ot = Column(String(100), nullable=True)
    empresa = Column(String(200), nullable=True)
    id_interno = Column(String(100), nullable=True)
    motivo = Column(Text, nullable=True)
    comunicado = Column(Text, nullable=True)
    fecha_liberada_ingeclean = Column(Date, nullable=True)
    fecha_liberada_facility = Column(Date, nullable=True)
    fecha_liberada_mantencion = Column(Date, nullable=True)
    fecha_liberada_investigacion = Column(Date, nullable=True)
    observacion = Column(Text, nullable=True)
    creado = Column(DateTime, nullable=False, default=now_utc)


class OrdenamientoEntry(Base):
    __tablename__ = "ordenamiento"
    id = Column(Integer, primary_key=True)
    fecha_ejecucion = Column(Date, nullable=True, index=True)
    empresa = Column(String(200), nullable=True)
    habitacion = Column(String(100), nullable=True)
    nombre = Column(String(200), nullable=True)
    rut = Column(String(50), nullable=True)
    turno = Column(String(100), nullable=True)
    reasignacion = Column(String(100), nullable=True)
    motivo_cambio = Column(Text, nullable=True)
    procesado = Column(String(200), nullable=True)
    pendiente = Column(String(200), nullable=True)
    creado = Column(DateTime, nullable=False, default=now_utc)


class HabitacionLiberadaEntry(Base):
    __tablename__ = "habitaciones_liberadas"
    id = Column(Integer, primary_key=True)
    habitacion = Column(String(100), nullable=True)
    empresa = Column(String(200), nullable=True)
    entrega_devolucion = Column(String(200), nullable=True)
    fecha_devolucion = Column(Date, nullable=True, index=True)
    comentario = Column(Text, nullable=True)
    uso_a_partir_de = Column(Date, nullable=True)
    observacion = Column(Text, nullable=True)
    creado = Column(DateTime, nullable=False, default=now_utc)


def ensure_deviation_actions_column(engine):
    """Añade el campo opcional a instalaciones existentes sin alterar sus registros."""
    with engine.begin() as conn:
        existing = {column["name"] for column in inspect(conn).get_columns("desviaciones")}
        if "acciones" not in existing:
            optional = " IF NOT EXISTS" if conn.dialect.name == "postgresql" else ""
            conn.execute(text(f"ALTER TABLE desviaciones ADD COLUMN{optional} acciones TEXT"))


def ensure_extension_columns(engine):
    """Agrega los nuevos campos sin reinterpretar ni sobrescribir datos históricos."""
    with engine.begin() as conn:
        existing = {column["name"] for column in inspect(conn).get_columns("extension_excepcion")}
        optional = " IF NOT EXISTS" if conn.dialect.name == "postgresql" else ""
        for name in ("centro_costos", "tipo_solicitud"):
            if name not in existing:
                conn.execute(text(f"ALTER TABLE extension_excepcion ADD COLUMN{optional} {name} VARCHAR(200)"))


def ensure_optional_entry_exit_fields(engine):
    """Permite fechas/horas vacías también en las bases de la versión anterior."""
    table_name = "entradas_salidas"
    with engine.begin() as conn:
        if conn.dialect.name == "sqlite":
            # SQLite requiere una transacción explícita para revertir también el DDL.
            conn.exec_driver_sql("BEGIN IMMEDIATE")
        columns = {column["name"]: column for column in inspect(conn).get_columns(table_name)}
        required = [name for name in ("fecha_ingreso", "hora_entrada") if not columns[name]["nullable"]]
        if not required:
            return
        if conn.dialect.name == "postgresql":
            for name in required:
                conn.execute(text(f"ALTER TABLE entradas_salidas ALTER COLUMN {name} DROP NOT NULL"))
        elif conn.dialect.name == "sqlite":
            # SQLite no permite quitar NOT NULL directamente. Se copia la tabla
            # dentro de la misma transacción conservando datos, índices y triggers.
            schema_objects = conn.execute(text(
                "SELECT sql FROM sqlite_master WHERE tbl_name = :name "
                "AND type IN ('index', 'trigger') AND sql IS NOT NULL"
            ), {"name": table_name}).scalars().all()
            original = Table(table_name, MetaData(), autoload_with=conn)
            replacement = original.to_metadata(MetaData(), name="entradas_salidas_nullable")
            for name in required:
                replacement.c[name].nullable = True
            conn.execute(CreateTable(replacement))
            conn.execute(replacement.insert().from_select(list(original.c.keys()), original.select()))
            conn.exec_driver_sql("DROP TABLE entradas_salidas")
            conn.exec_driver_sql("ALTER TABLE entradas_salidas_nullable RENAME TO entradas_salidas")
            for statement in schema_objects:
                conn.exec_driver_sql(statement)
        else:
            raise RuntimeError("La actualización de Entradas y salidas requiere SQLite o PostgreSQL.")


# Crear tablas si no existen y actualizar las instalaciones anteriores.
Base.metadata.create_all(ENGINE)
ensure_deviation_actions_column(ENGINE)
ensure_extension_columns(ENGINE)
ensure_optional_entry_exit_fields(ENGINE)

# Compatibilidad con instalaciones anteriores cuya tabla no tenía la fecha de negocio.
columns = {column["name"] for column in inspect(ENGINE).get_columns("cumplimiento_eecc")}
with ENGINE.begin() as conn:
    if "fecha" not in columns:
        conn.execute(text("ALTER TABLE cumplimiento_eecc ADD COLUMN fecha date"))
    conn.execute(text("""
        UPDATE cumplimiento_eecc
        SET fecha = COALESCE(DATE(creado), CURRENT_DATE)
        WHERE fecha IS NULL
    """))
    conn.execute(text("CREATE INDEX IF NOT EXISTS ix_cumplimiento_eecc_fecha ON cumplimiento_eecc (fecha)"))
    if ENGINE.dialect.name == "postgresql":
        conn.execute(text("ALTER TABLE cumplimiento_eecc ALTER COLUMN fecha SET NOT NULL"))

# -----------------------------------------------------------------------------
# Helpers filtros
# -----------------------------------------------------------------------------
def resolve_filters(args):
    semana = args.get("semana", type=int)
    d_from = args.get("from")
    d_to = args.get("to")
    if semana and semana in WEEK_MAP:
        return *week_range(semana), semana
    df = date.fromisoformat(d_from) if d_from else None
    dt = date.fromisoformat(d_to) if d_to else None
    return df, dt, None


# -----------------------------------------------------------------------------
# Rutas básicas
# -----------------------------------------------------------------------------
@app.get("/health")
def health():
    return jsonify(status="ok"), 200

@app.get("/")
def home():
    return redirect(url_for("panel", tab="censo"))

# -----------------------------------------------------------------------------
# PANEL (maneja formularios e importaciones por pestaña)
# -----------------------------------------------------------------------------
@app.route("/panel", methods=["GET", "POST"])
def panel():
    # tabs: censo | eventos | duplicidades | encuesta | atencion |
    #       robos | miscelaneo | desviaciones | solicitud_ot | reclamos |
    #       alarmas | extensiones | onboarding | apertura | cumplimiento | entradas_salidas
    tab = request.args.get("tab", "censo")
    db = SessionLocal()
    try:
        if request.method == "POST":
            # -------------------- CENSO --------------------
            if tab == "censo":
                fecha = safe_convert_date(request.form["fecha"])
                cd = int(request.form.get("censo_dia", 0) or 0)
                cn = int(request.form.get("censo_noche", 0) or 0)
                total = int(request.form.get("total", cd + cn) or (cd + cn))
                db.add(CensusEntry(fecha=fecha, censo_dia=cd, censo_noche=cn, total=total))
                db.commit(); flash("Censo guardado.")

            # -------------------- EVENTOS --------------------
            elif tab == "eventos":
                fecha = safe_convert_date(request.form["fecha"])
                horario = request.form.get("horario", "").strip()
                que = request.form.get("que_ocurrio", "").strip()
                nom = request.form.get("nombre_afectado", "").strip()
                accion = request.form.get("accion", "").strip()
                db.add(EventSeguridad(fecha=fecha, horario=horario, que_ocurrio=que,
                                      nombre_afectado=nom, accion=accion))
                db.commit(); flash("Evento de seguridad guardado.")

            # -------------------- DUPLICIDADES --------------------
            elif tab == "duplicidades":
                semana = int(request.form["semana"])
                fecha = safe_convert_date(request.form["fecha"])
                rec = DuplicidadEntry(
                    semana=semana,
                    fecha=fecha,
                    id_interno=request.form.get("id", "").strip(),
                    empresa_contratista=request.form.get("empresa_contratista", "").strip(),
                    descripcion_problema=request.form.get("descripcion_problema", "").strip(),
                    tipo_riesgo=request.form.get("tipo_riesgo", "").strip(),
                    pabellon=request.form.get("pabellon", "").strip(),
                    habitacion=request.form.get("habitacion", "").strip(),
                    ingresar_contacto=request.form.get("ingresar_contacto", "").strip(),
                    nombre_usuario=request.form.get("nombre_usuario", "").strip(),
                    responsable=request.form.get("responsable", "").strip(),
                    estatus=request.form.get("estatus", "").strip(),
                    notificacion_usuario=request.form.get("notificacion_usuario", "").strip(),
                    plan_accion=request.form.get("plan_accion", "").strip(),
                    fecha_cierre=safe_convert_date(request.form.get("fecha_cierre")),
                )
                db.add(rec); db.commit(); flash("Duplicidad guardada.")

            # -------------------- ENCUESTA --------------------
            elif tab == "encuesta":
                fh_raw = request.form.get("fecha_hora")
                fecha_hora = safe_convert_datetime(fh_raw)
                vals = {}
                total = 0; n = 0
                for i in range(1,6):
                    r = request.form.get(f"q{i}_respuesta", "")
                    p = request.form.get(f"q{i}_puntaje", "")
                    p = int(p) if (str(p).isdigit() or (isinstance(p, str) and p.strip().isdigit())) else None
                    vals[i] = (r, p)
                    if p is not None: total += p; n += 1
                promedio = (total / n) if n>0 else None
                db.add(EncuestaEntry(
                    fecha_hora=fecha_hora,
                    q1_respuesta=vals[1][0], q1_puntaje=vals[1][1],
                    q2_respuesta=vals[2][0], q2_puntaje=vals[2][1],
                    q3_respuesta=vals[3][0], q3_puntaje=vals[3][1],
                    q4_respuesta=vals[4][0], q4_puntaje=vals[4][1],
                    q5_respuesta=vals[5][0], q5_puntaje=vals[5][1],
                    total=total if n>0 else None,
                    promedio=round(promedio,2) if promedio is not None else None,
                    comentarios=request.form.get("comentarios", "").strip(),
                ))
                db.commit(); flash("Encuesta guardada.")

            # -------------------- ATENCIÓN --------------------
            elif tab == "atencion":
                fecha = safe_convert_date(request.form["fecha"])
                tiempo_input = request.form.get("tiempo_promedio", "").strip()
                
                segundos = safe_convert_time(tiempo_input)
                
                cant = int(request.form.get("cantidad", 0) or 0)
                db.add(AtencionEntry(fecha=fecha, tiempo_promedio_sec=segundos, cantidad=cant))
                db.commit(); flash("Atención guardada.")

            # -------------------- ROBOS / HURTOS --------------------
            elif tab == "robos":
                fecha = safe_convert_date(request.form["fecha"])
                hora = request.form.get("hora", "00:00")
                hora_obj = safe_time_hhmm(hora)  # Usamos la nueva función robusta
                rec = RoboHurtoEntry(
                    fecha=fecha, hora=hora_obj,
                    modulo=request.form.get("modulo","").strip(),
                    habitacion=request.form.get("habitacion","").strip(),
                    empresa=request.form.get("empresa","").strip(),
                    nombre_cliente=request.form.get("nombre_cliente","").strip(),
                    rut=request.form.get("rut","").strip(),
                    medio_reclamo=request.form.get("medio_reclamo","").strip(),
                    especies=request.form.get("especies","").strip(),
                    observaciones=request.form.get("observaciones","").strip(),
                    recepciona=request.form.get("recepciona","").strip(),
                )
                db.add(rec); db.commit(); flash("Robo/Hurto guardado.")

            # -------------------- MISCELÁNEO --------------------
            elif tab == "miscelaneo":
                rec = MiscelaneoEntry(
                    ot=request.form.get("ot","").strip(),
                    division=request.form.get("division","").strip(),
                    area=request.form.get("area","").strip(),
                    lugar=request.form.get("lugar","").strip(),
                    ubicacion=request.form.get("ubicacion","").strip(),
                    disciplina=request.form.get("disciplina","").strip(),
                    especialidad=request.form.get("especialidad","").strip(),
                    falla=request.form.get("falla","").strip(),
                    empresa=request.form.get("empresa","").strip(),
                    fecha_creacion=safe_convert_date(request.form.get("fecha_creacion")),
                    fecha_inicio=safe_convert_date(request.form.get("fecha_inicio")),
                    fecha_termino=safe_convert_date(request.form.get("fecha_termino")),
                    fecha_aprobacion=safe_convert_date(request.form.get("fecha_aprobacion")),
                    estado=request.form.get("estado","").strip(),
                    comentario=request.form.get("comentario","").strip(),
                )
                db.add(rec); db.commit(); flash("Misceláneo guardado.")

            # -------------------- DESVIACIONES --------------------
            elif tab == "desviaciones":
                rec = DesviacionEntry(
                    n_solicitud=request.form.get("n_solicitud","").strip(),
                    fecha=safe_convert_date(request.form["fecha"]),
                    id_interno=request.form.get("id","").strip(),
                    empresa_contratista=request.form.get("empresa_contratista","").strip(),
                    descripcion_problema=request.form.get("descripcion_problema","").strip(),
                    tipo_riesgo=request.form.get("tipo_riesgo","").strip(),
                    tipo_solicitud=request.form.get("tipo_solicitud","").strip(),
                    pabellon=request.form.get("pabellon","").strip(),
                    habitacion=request.form.get("habitacion","").strip(),
                    via_solicitud=request.form.get("via_solicitud","").strip(),
                    quien_informa=request.form.get("quien_informa","").strip(),
                    riesgo_material=request.form.get("riesgo_material","").strip(),
                    correo_destino=request.form.get("correo_destino","").strip(),
                    acciones=request.form.get("acciones","").strip(),
                )
                db.add(rec); db.commit(); flash("Desviación guardada.")

            # -------------------- SOLICITUD / OT USUARIO --------------------
            elif tab == "solicitud_ot":
                def to_secs(v):
                    v = (v or "").strip()
                    return safe_convert_time(v)
                rec = SolicitudOTEntry(
                    n_solicitud=request.form.get("n_solicitud","").strip(),
                    descripcion_problema=request.form.get("descripcion_problema","").strip(),
                    tipo_solicitud=request.form.get("tipo_solicitud","").strip(),
                    modulo=request.form.get("modulo","").strip(),
                    habitacion=request.form.get("habitacion","").strip(),
                    tipo_turno=request.form.get("tipo_turno","").strip(),
                    jornada=request.form.get("jornada","").strip(),
                    via_solicitud=request.form.get("via_solicitud","").strip(),
                    correo_usuario=request.form.get("correo_usuario","").strip(),
                    tipo_tarea=request.form.get("tipo_tarea","").strip(),
                    ot=request.form.get("ot","").strip(),
                    fecha_inicio=safe_convert_date(request.form.get("fecha_inicio")),
                    estado=request.form.get("estado","").strip(),
                    tiempo_respuesta_sec=to_secs(request.form.get("tiempo_respuesta")),
                    satisfaccion_reclamo=request.form.get("satisfaccion_reclamo","").strip(),
                    motivo=request.form.get("motivo","").strip(),
                    observacion=request.form.get("observacion","").strip(),
                )
                db.add(rec); db.commit(); flash("Solicitud/OT guardada.")

            # -------------------- RECLAMOS USUARIOS --------------------
            elif tab == "reclamos":
                rec = ReclamoUsuarioEntry(
                    n_solicitud=request.form.get("n_solicitud","").strip(),
                    fecha=safe_convert_date(request.form["fecha"]),
                    id_interno=request.form.get("id","").strip(),
                    empresa_contratista=request.form.get("empresa_contratista","").strip(),
                    descripcion_problema=request.form.get("descripcion_problema","").strip(),
                    tipo_solicitud=request.form.get("tipo_solicitud","").strip(),
                    pabellon=request.form.get("pabellon","").strip(),
                    habitacion=request.form.get("habitacion","").strip(),
                    via_solicitud=request.form.get("via_solicitud","").strip(),
                    ingresar_contacto=request.form.get("ingresar_contacto","").strip(),
                    nombre_usuario=request.form.get("nombre_usuario","").strip(),
                    responsable=request.form.get("responsable","").strip(),
                    estatus=request.form.get("estatus","").strip(),
                    notificacion_usuario=request.form.get("notificacion_usuario","").strip(),
                    plan_accion=request.form.get("plan_accion","").strip(),
                )
                db.add(rec); db.commit(); flash("Reclamo de usuario guardado.")

            # -------------------- ACTIVACIÓN DE ALARMA --------------------
            elif tab == "alarmas":
                fecha = safe_convert_date(request.form["fecha"])
                def f2t(v):
                    v = (v or "").strip()
                    return safe_time_hhmm(v)  # Usamos la nueva función robusta
                def f2float(v):
                    try:
                        return float(v) if (v is not None and str(v).strip()!="") else None
                    except:
                        return None
                rec = ActivacionAlarmaEntry(
                    modulo=request.form.get("modulo","").strip(),
                    n_habitacion=request.form.get("n_habitacion","").strip(),
                    nombre_recepcionista=request.form.get("nombre_recepcionista","").strip(),
                    fecha=fecha,
                    empresa=request.form.get("empresa","").strip(),
                    id_interno=request.form.get("id_interno","").strip(),
                    co=request.form.get("co","").strip(),
                    aviso_mantencion_h=f2float(request.form.get("aviso_mantencion_h")),
                    llegada_mantencion_h=f2float(request.form.get("llegada_mantencion_h")),
                    aviso_lider_h=f2float(request.form.get("aviso_lider_h")),
                    llegada_lider_h=f2float(request.form.get("llegada_lider_h")),
                    hora_reporte_salfa=f2t(request.form.get("hora_reporte_salfa")),
                    tipo_evento=request.form.get("tipo_evento","").strip(),
                    tipo_actividad=request.form.get("tipo_actividad","").strip(),
                    fecha_reporte=safe_convert_date(request.form.get("fecha_reporte")),
                    turno_recepcion_ingresos=request.form.get("turno_recepcion_ingresos","").strip(),
                    observaciones=request.form.get("observaciones","").strip(),
                )
                db.add(rec); db.commit(); flash("Activación de alarma guardada.")

            # -------------------- EXTENSIÓN / EXCEPCIÓN --------------------
            elif tab == "extensiones":
                rec = ExtensionExcepcionEntry(
                    fecha_solicitud=safe_convert_date(request.form["fecha_solicitud"]),
                    id_interno=request.form.get("id_interno","").strip(),
                    empresa=request.form.get("empresa","").strip(),
                    co=request.form.get("co","").strip(),
                    gerencia=request.form.get("gerencia","").strip(),
                    centro_costos=request.form.get("centro_costos","").strip(),
                    cant_clientes=(int(request.form.get("cant_clientes")) if request.form.get("cant_clientes") else None),
                    tipo_solicitud=request.form.get("tipo_solicitud","").strip(),
                    desde=safe_convert_date(request.form.get("desde")),
                    hasta=safe_convert_date(request.form.get("hasta")),
                    aprobador=request.form.get("aprobador","").strip(),
                    observacion=request.form.get("observacion","").strip(),
                )
                db.add(rec); db.commit(); flash("Extensión/Excepción guardada.")

            # -------------------- ONBOARDING --------------------
            elif tab == "onboarding":
                fh_raw = request.form.get("fecha_hora")
                fecha_hora = safe_convert_datetime(fh_raw)
                uploaded_pdf = request.files.get("archivo_pdf")
                rec = OnboardingEntry(
                    fecha_hora=fecha_hora,
                    nombre=request.form.get("nombre","").strip(),
                    rut=request.form.get("rut","").strip(),
                    empresa=request.form.get("empresa","").strip(),
                    id_interno=request.form.get("id_interno","").strip(),
                    archivo_pdf=(uploaded_pdf.filename or "").strip() if uploaded_pdf else "",
                )
                db.add(rec); db.commit(); flash("Onboarding guardado.")

            # -------------------- APERTURA DE HABITACIÓN --------------------
            elif tab == "apertura":
                def f2t(v):
                    v = (v or "").strip()
                    return safe_time_hhmm(v)  # Usamos la nueva función robusta
                rec = AperturaHabitacionEntry(
                    fecha=safe_convert_date(request.form["fecha"]),
                    habitacion=request.form.get("habitacion","").strip(),
                    hora=f2t(request.form.get("hora")),
                    responsable=request.form.get("responsable","").strip(),
                    estado_chapa=request.form.get("estado_chapa","").strip(),
                )
                db.add(rec); db.commit(); flash("Apertura de habitación guardada.")

            # -------------------- CUMPLIMIENTO EECC --------------------
            elif tab == "cumplimiento":
                rec = CumplimientoEECCEntry(
                    empresa=request.form.get("empresa","").strip(),
                    n_contrato=request.form.get("n_contrato","").strip(),
                    co=request.form.get("co","").strip(),
                    correo_electronico=request.form.get("correo_electronico","").strip(),
                    id_interno=request.form.get("id_interno","").strip(),
                    turno=request.form.get("turno","").strip(),
            
                    # --- NUEVO ---
                    fecha=safe_convert_date(request.form.get("fecha")),
                )
                db.add(rec); db.commit(); flash("Cumplimiento EECC guardado.")

            elif tab in OPTIONAL_RECORD_FIELDS:
                Model = ENTITY_MODEL[tab]
                fields = edit_fields(tab, Model())
                values, errors = parse_edit_values(tab, fields, request.form)
                if errors:
                    return render_template(
                        "panel.html", tab=tab, week_map=WEEK_MAP, current_tab=tab,
                        form_fields=fields, field_errors=errors, form_values=request.form,
                    ), 422
                db.add(Model(**values))
                db.commit()
                flash(f"Registro de {EDIT_CONFIG[tab][0].lower()} guardado.", "success")

            return redirect(url_for("panel", tab=tab))

        # GET
        return render_template(
            "panel.html", tab=tab, week_map=WEEK_MAP, current_tab=tab,
            form_fields=edit_fields(tab, ENTITY_MODEL[tab]()) if tab in OPTIONAL_RECORD_FIELDS else [],
            field_errors={}, form_values={},
        )
    finally:
        db.close()


# -----------------------------------------------------------------------------
# LISTADOS / REGISTROS + DESCARGAS CSV
# -----------------------------------------------------------------------------
@app.get("/registros")
def registros():
    d_from, d_to, semana_sel = resolve_filters(request.args)
    vista = request.args.get("vista", "censo")
    if vista not in ENTITY_MODEL:
        abort(404)
    db = SessionLocal()
    try:
        Model = ENTITY_MODEL[vista]
        date_column = getattr(Model, ENTITY_DATE_FIELD.get(vista, "fecha"))
        date_order = date_column.desc().nullslast() if vista == "solicitud_ot" or vista in OPTIONAL_RECORD_FIELDS else date_column.desc()
        rows = hotel_records_query(db, vista, d_from, d_to).order_by(date_order, Model.id.desc()).all()
        listing = {key: [] for key in ENTITY_LIST_KEY.values()}
        listing[ENTITY_LIST_KEY[vista]] = rows
        return render_template(
            "list.html",
            semana_sel=semana_sel, d_from=d_from, d_to=d_to, week_map=WEEK_MAP,
            vista=vista, current_tab=None, **listing,
            record_count=len(rows), entity_title=EDIT_CONFIG[vista][0],
            record_columns=OPTIONAL_RECORD_FIELDS.get(vista, ()), current_records=rows,
            record_date_label=dict(OPTIONAL_RECORD_FIELDS.get(vista, ())).get(ENTITY_DATE_FIELD.get(vista)),
            bulk_csrf=session.setdefault("hotel_bulk_csrf", secrets.token_hex(32)),
        )
    finally:
        db.close()


@app.get("/download/<string:entity>.csv")
def download_entity(entity):
    d_from, d_to, semana_sel = resolve_filters(request.args)
    if semana_sel: d_from, d_to = week_range(semana_sel)
    db = SessionLocal()
    try:
        buf = io.StringIO()
        w = None

        if entity == "censo":
            q = db.query(CensusEntry)
            if d_from: q = q.filter(CensusEntry.fecha >= d_from)
            if d_to:   q = q.filter(CensusEntry.fecha <= d_to)
            rows = q.order_by(CensusEntry.fecha).all()
            w = csv.DictWriter(buf, fieldnames=["fecha", "censo_dia", "censo_noche", "total"])
            w.writeheader()
            for r in rows:
                w.writerow({"fecha": r.fecha.isoformat(), "censo_dia": r.censo_dia, "censo_noche": r.censo_noche, "total": r.total})

        elif entity == "eventos":
            q = db.query(EventSeguridad)
            if d_from: q = q.filter(EventSeguridad.fecha >= d_from)
            if d_to:   q = q.filter(EventSeguridad.fecha <= d_to)
            rows = q.order_by(EventSeguridad.fecha).all()
            w = csv.DictWriter(buf, fieldnames=["fecha","horario","que_ocurrio","nombre_afectado","accion"])
            w.writeheader()
            for r in rows:
                w.writerow({"fecha": r.fecha.isoformat(), "horario": r.horario, "que_ocurrio": r.que_ocurrio,
                            "nombre_afectado": r.nombre_afectado or "", "accion": r.accion or ""})

        elif entity == "duplicidades":
            q = db.query(DuplicidadEntry)
            if d_from: q = q.filter(DuplicidadEntry.fecha >= d_from)
            if d_to:   q = q.filter(DuplicidadEntry.fecha <= d_to)
            rows = q.order_by(DuplicidadEntry.fecha).all()
            headers = ["semana","fecha","id","empresa_contratista","descripcion_problema","tipo_riesgo",
                       "pabellon","habitacion","ingresar_contacto","nombre_usuario","responsable","estatus",
                       "notificacion_usuario","plan_accion","fecha_cierre"]
            w = csv.DictWriter(buf, fieldnames=headers)
            w.writeheader()
            for r in rows:
                w.writerow({
                    "semana": r.semana, "fecha": r.fecha.isoformat(), "id": r.id_interno or "",
                    "empresa_contratista": r.empresa_contratista or "", "descripcion_problema": r.descripcion_problema or "",
                    "tipo_riesgo": r.tipo_riesgo or "", "pabellon": r.pabellon or "", "habitacion": r.habitacion or "",
                    "ingresar_contacto": r.ingresar_contacto or "", "nombre_usuario": r.nombre_usuario or "",
                    "responsable": r.responsable or "", "estatus": r.estatus or "",
                    "notificacion_usuario": r.notificacion_usuario or "", "plan_accion": r.plan_accion or "",
                    "fecha_cierre": r.fecha_cierre.isoformat() if r.fecha_cierre else ""
                })

        elif entity == "encuestas":
            q = db.query(EncuestaEntry)
            if d_from: q = q.filter(EncuestaEntry.fecha_hora >= datetime.combine(d_from, time.min))
            if d_to:   q = q.filter(EncuestaEntry.fecha_hora <= datetime.combine(d_to, time.max))
            rows = q.order_by(EncuestaEntry.fecha_hora).all()
            headers = ["fecha_hora","q1_respuesta","q1_puntaje","q2_respuesta","q2_puntaje",
                       "q3_respuesta","q3_puntaje","q4_respuesta","q4_puntaje","q5_respuesta","q5_puntaje",
                       "total","promedio","comentarios"]
            w = csv.DictWriter(buf, fieldnames=headers)
            w.writeheader()
            for r in rows:
                w.writerow({
                    "fecha_hora": r.fecha_hora.isoformat(timespec="minutes"),
                    "q1_respuesta": r.q1_respuesta or "", "q1_puntaje": r.q1_puntaje or "",
                    "q2_respuesta": r.q2_respuesta or "", "q2_puntaje": r.q2_puntaje or "",
                    "q3_respuesta": r.q3_respuesta or "", "q3_puntaje": r.q3_puntaje or "",
                    "q4_respuesta": r.q4_respuesta or "", "q4_puntaje": r.q4_puntaje or "",
                    "q5_respuesta": r.q5_respuesta or "", "q5_puntaje": r.q5_puntaje or "",
                    "total": r.total if r.total is not None else "",
                    "promedio": r.promedio if r.promedio is not None else "",
                    "comentarios": r.comentarios or "",
                })

        elif entity == "atencion":
            q = db.query(AtencionEntry)
            if d_from: q = q.filter(AtencionEntry.fecha >= d_from)
            if d_to:   q = q.filter(AtencionEntry.fecha <= d_to)
            rows = q.order_by(AtencionEntry.fecha).all()
            w = csv.DictWriter(buf, fieldnames=["fecha","tiempo_promedio_mmss","cantidad"])
            w.writeheader()
            for r in rows:
                w.writerow({"fecha": r.fecha.isoformat(), "tiempo_promedio_mmss": seconds_to_mmss(r.tiempo_promedio_sec),
                            "cantidad": r.cantidad})

        # ---------------- CSV de módulos previos ----------------
        elif entity == "robos":
            q = db.query(RoboHurtoEntry)
            if d_from: q = q.filter(RoboHurtoEntry.fecha >= d_from)
            if d_to:   q = q.filter(RoboHurtoEntry.fecha <= d_to)
            rows = q.order_by(RoboHurtoEntry.fecha).all()
            headers = ["fecha","hora","modulo","habitacion","empresa","nombre_cliente","rut",
                       "medio_reclamo","especies","observaciones","recepciona"]
            w = csv.DictWriter(buf, fieldnames=headers); w.writeheader()
            for r in rows:
                w.writerow({
                    "fecha": r.fecha.isoformat(),
                    "hora": r.hora.strftime("%H:%M"),
                    "modulo": r.modulo or "",
                    "habitacion": r.habitacion or "",
                    "empresa": r.empresa or "",
                    "nombre_cliente": r.nombre_cliente or "",
                    "rut": r.rut or "",
                    "medio_reclamo": r.medio_reclamo or "",
                    "especies": r.especies or "",
                    "observaciones": r.observaciones or "",
                    "recepciona": r.recepciona or "",
                })

        elif entity == "miscelaneo":
            q = db.query(MiscelaneoEntry)
            if d_from: q = q.filter(MiscelaneoEntry.fecha_creacion >= d_from)
            if d_to:   q = q.filter(MiscelaneoEntry.fecha_creacion <= d_to)
            rows = q.order_by(MiscelaneoEntry.fecha_creacion, MiscelaneoEntry.id).all()
            headers = ["ot","division","area","lugar","ubicacion","disciplina","especialidad","falla",
                       "empresa","fecha_creacion","fecha_inicio","fecha_termino","fecha_aprobacion","estado","comentario"]
            w = csv.DictWriter(buf, fieldnames=headers); w.writeheader()
            for r in rows:
                w.writerow({
                    "ot": r.ot or "", "division": r.division or "", "area": r.area or "",
                    "lugar": r.lugar or "", "ubicacion": r.ubicacion or "", "disciplina": r.disciplina or "",
                    "especialidad": r.especialidad or "", "falla": r.falla or "", "empresa": r.empresa or "",
                    "fecha_creacion": r.fecha_creacion.isoformat() if r.fecha_creacion else "",
                    "fecha_inicio": r.fecha_inicio.isoformat() if r.fecha_inicio else "",
                    "fecha_termino": r.fecha_termino.isoformat() if r.fecha_termino else "",
                    "fecha_aprobacion": r.fecha_aprobacion.isoformat() if r.fecha_aprobacion else "",
                    "estado": r.estado or "", "comentario": r.comentario or "",
                })

        elif entity == "desviaciones":
            q = db.query(DesviacionEntry)
            if d_from: q = q.filter(DesviacionEntry.fecha >= d_from)
            if d_to:   q = q.filter(DesviacionEntry.fecha <= d_to)
            rows = q.order_by(DesviacionEntry.fecha).all()
            headers = ["n_solicitud","fecha","id","empresa_contratista","descripcion_problema","tipo_riesgo",
                       "tipo_solicitud","pabellon","habitacion","via_solicitud","quien_informa","riesgo_material","correo_destino","acciones"]
            w = csv.DictWriter(buf, fieldnames=headers); w.writeheader()
            for r in rows:
                w.writerow({
                    "n_solicitud": r.n_solicitud or "", "fecha": r.fecha.isoformat(),
                    "id": r.id_interno or "", "empresa_contratista": r.empresa_contratista or "",
                    "descripcion_problema": r.descripcion_problema or "", "tipo_riesgo": r.tipo_riesgo or "",
                    "tipo_solicitud": r.tipo_solicitud or "", "pabellon": r.pabellon or "",
                    "habitacion": r.habitacion or "", "via_solicitud": r.via_solicitud or "",
                    "quien_informa": r.quien_informa or "", "riesgo_material": r.riesgo_material or "",
                    "correo_destino": r.correo_destino or "",
                    "acciones": r.acciones or "",
                })

        elif entity == "solicitud_ot":
            q = db.query(SolicitudOTEntry)
            if d_from: q = q.filter(SolicitudOTEntry.fecha_inicio >= d_from)
            if d_to:   q = q.filter(SolicitudOTEntry.fecha_inicio <= d_to)
            rows = q.order_by(SolicitudOTEntry.fecha_inicio, SolicitudOTEntry.id).all()
            headers = ["n_solicitud","descripcion_problema","tipo_solicitud","modulo","habitacion","tipo_turno",
                       "jornada","via_solicitud","correo_usuario","tipo_tarea","ot","fecha_inicio","estado",
                       "tiempo_respuesta_mmss","satisfaccion_reclamo","motivo","observacion"]
            w = csv.DictWriter(buf, fieldnames=headers); w.writeheader()
            for r in rows:
                w.writerow({
                    "n_solicitud": r.n_solicitud or "", "descripcion_problema": r.descripcion_problema or "",
                    "tipo_solicitud": r.tipo_solicitud or "", "modulo": r.modulo or "",
                    "habitacion": r.habitacion or "", "tipo_turno": r.tipo_turno or "",
                    "jornada": r.jornada or "", "via_solicitud": r.via_solicitud or "",
                    "correo_usuario": r.correo_usuario or "", "tipo_tarea": r.tipo_tarea or "",
                    "ot": r.ot or "", "fecha_inicio": r.fecha_inicio.isoformat() if r.fecha_inicio else "",
                    "estado": r.estado or "",
                    "tiempo_respuesta_mmss": seconds_to_mmss(r.tiempo_respuesta_sec or 0),
                    "satisfaccion_reclamo": r.satisfaccion_reclamo or "", "motivo": r.motivo or "",
                    "observacion": r.observacion or "",
                })

        elif entity == "reclamos":
            q = db.query(ReclamoUsuarioEntry)
            if d_from: q = q.filter(ReclamoUsuarioEntry.fecha >= d_from)
            if d_to:   q = q.filter(ReclamoUsuarioEntry.fecha <= d_to)
            rows = q.order_by(ReclamoUsuarioEntry.fecha).all()
            headers = ["n_solicitud","fecha","id","empresa_contratista","descripcion_problema","tipo_solicitud",
                       "pabellon","habitacion","via_solicitud","ingresar_contacto","nombre_usuario","responsable",
                       "estatus","notificacion_usuario","plan_accion"]
            w = csv.DictWriter(buf, fieldnames=headers); w.writeheader()
            for r in rows:
                w.writerow({
                    "n_solicitud": r.n_solicitud or "", "fecha": r.fecha.isoformat(),
                    "id": r.id_interno or "", "empresa_contratista": r.empresa_contratista or "",
                    "descripcion_problema": r.descripcion_problema or "", "tipo_solicitud": r.tipo_solicitud or "",
                    "pabellon": r.pabellon or "", "habitacion": r.habitacion or "",
                    "via_solicitud": r.via_solicitud or "", "ingresar_contacto": r.ingresar_contacto or "",
                    "nombre_usuario": r.nombre_usuario or "", "responsable": r.responsable or "",
                    "estatus": r.estatus or "", "notificacion_usuario": r.notificacion_usuario or "",
                    "plan_accion": r.plan_accion or "",
                })

        # --------- CSV NUEVOS 5 ----------
        elif entity == "alarmas":
            q = db.query(ActivacionAlarmaEntry)
            if d_from: q = q.filter(ActivacionAlarmaEntry.fecha >= d_from)
            if d_to:   q = q.filter(ActivacionAlarmaEntry.fecha <= d_to)
            rows = q.order_by(ActivacionAlarmaEntry.fecha).all()
            headers = ["MODULO","N_HABITACION","NOMBRE_RECEPCIONISTA","FECHA","EMPRESA","ID","CO",
                       "AVISO_MANTENCION_H","LLEGADA_MANTENCION_H","AVISO_LIDER_H","LLEGADA_LIDER_H",
                       "HORA_REPORTE_SALFA","TIPO_EVENTO","TIPO_ACTIVIDAD","FECHA_REPORTE",
                       "TURNO_RECEPCION_INGRESOS","OBSERVACIONES"]
            w = csv.DictWriter(buf, fieldnames=headers); w.writeheader()
            for r in rows:
                w.writerow({
                    "MODULO": r.modulo or "", "N_HABITACION": r.n_habitacion or "",
                    "NOMBRE_RECEPCIONISTA": r.nombre_recepcionista or "",
                    "FECHA": r.fecha.isoformat(), "EMPRESA": r.empresa or "",
                    "ID": r.id_interno or "", "CO": r.co or "",
                    "AVISO_MANTENCION_H": r.aviso_mantencion_h if r.aviso_mantencion_h is not None else "",
                    "LLEGADA_MANTENCION_H": r.llegada_mantencion_h if r.llegada_mantencion_h is not None else "",
                    "AVISO_LIDER_H": r.aviso_lider_h if r.aviso_lider_h is not None else "",
                    "LLEGADA_LIDER_H": r.llegada_lider_h if r.llegada_lider_h is not None else "",
                    "HORA_REPORTE_SALFA": r.hora_reporte_salfa.strftime("%H:%M") if r.hora_reporte_salfa else "",
                    "TIPO_EVENTO": r.tipo_evento or "", "TIPO_ACTIVIDAD": r.tipo_actividad or "",
                    "FECHA_REPORTE": r.fecha_reporte.isoformat() if r.fecha_reporte else "",
                    "TURNO_RECEPCION_INGRESOS": r.turno_recepcion_ingresos or "",
                    "OBSERVACIONES": r.observaciones or "",
                })

        elif entity == "extensiones":
            q = db.query(ExtensionExcepcionEntry)
            if d_from: q = q.filter(ExtensionExcepcionEntry.fecha_solicitud >= d_from)
            if d_to:   q = q.filter(ExtensionExcepcionEntry.fecha_solicitud <= d_to)
            rows = q.order_by(ExtensionExcepcionEntry.fecha_solicitud).all()
            headers = [label for _, label in EXTENSION_FIELDS]
            w = csv.DictWriter(buf, fieldnames=headers); w.writeheader()
            for r in rows:
                w.writerow({
                    label: getattr(r, name) if getattr(r, name) is not None else ""
                    for name, label in EXTENSION_FIELDS
                })

        elif entity == "onboarding":
            q = db.query(OnboardingEntry)
            if d_from: q = q.filter(OnboardingEntry.fecha_hora >= datetime.combine(d_from, time.min))
            if d_to:   q = q.filter(OnboardingEntry.fecha_hora <= datetime.combine(d_to, time.max))
            rows = q.order_by(OnboardingEntry.fecha_hora).all()
            headers = ["FECHA_HORA","NOMBRE","RUT","EMPRESA","ID","ARCHIVO_PDF"]
            w = csv.DictWriter(buf, fieldnames=headers); w.writeheader()
            for r in rows:
                w.writerow({
                    "FECHA_HORA": r.fecha_hora.isoformat(timespec="minutes"),
                    "NOMBRE": r.nombre or "", "RUT": r.rut or "", "EMPRESA": r.empresa or "",
                    "ID": r.id_interno or "", "ARCHIVO_PDF": r.archivo_pdf or "",
                })

        elif entity == "apertura":
            q = db.query(AperturaHabitacionEntry)
            if d_from: q = q.filter(AperturaHabitacionEntry.fecha >= d_from)
            if d_to:   q = q.filter(AperturaHabitacionEntry.fecha <= d_to)
            rows = q.order_by(AperturaHabitacionEntry.fecha).all()
            headers = ["FECHA","HABITACION","HORA","RESPONSABLE","ESTADO_CHAPA"]
            w = csv.DictWriter(buf, fieldnames=headers); w.writeheader()
            for r in rows:
                w.writerow({
                    "FECHA": r.fecha.isoformat(),
                    "HABITACION": r.habitacion or "",
                    "HORA": r.hora.strftime("%H:%M") if r.hora else "",
                    "RESPONSABLE": r.responsable or "",
                    "ESTADO_CHAPA": r.estado_chapa or "",
                })

        elif entity == "cumplimiento":
            q = db.query(CumplimientoEECCEntry)
            if d_from: q = q.filter(CumplimientoEECCEntry.fecha >= d_from)
            if d_to:   q = q.filter(CumplimientoEECCEntry.fecha <= d_to)
            rows = q.order_by(CumplimientoEECCEntry.fecha, CumplimientoEECCEntry.id).all()
            headers = ["FECHA","EMPRESA","N_CONTRATO","CO","CORREO_ELECTRONICO","ID","TURNO"]
            w = csv.DictWriter(buf, fieldnames=headers); w.writeheader()
            for r in rows:
                w.writerow({
                    "FECHA": r.fecha.isoformat() if r.fecha else "",
                    "EMPRESA": r.empresa or "",
                    "N_CONTRATO": r.n_contrato or "",
                    "CO": r.co or "",
                    "CORREO_ELECTRONICO": r.correo_electronico or "",
                    "ID": r.id_interno or "",
                    "TURNO": r.turno or "",
                })

        elif entity in OPTIONAL_RECORD_FIELDS:
            Model = ENTITY_MODEL[entity]
            columns = OPTIONAL_RECORD_FIELDS[entity]
            rows = hotel_records_query(db, entity, d_from, d_to).order_by(
                getattr(Model, ENTITY_DATE_FIELD[entity]), Model.id
            ).all()
            w = csv.DictWriter(buf, fieldnames=[label for _, label in columns])
            w.writeheader()
            for r in rows:
                w.writerow({label: getattr(r, name) if getattr(r, name) is not None else ""
                            for name, label in columns})

        else:
            flash("Entidad no válida.")
            return redirect(url_for("registros"))

        return send_file(
            io.BytesIO(buf.getvalue().encode("utf-8-sig")),
            mimetype="text/csv",
            as_attachment=True,
            download_name=f"{entity}.csv"
        )
    finally:
        db.close()
        
# --- Mapa entidad → Modelo para editar y eliminar ---
ENTITY_MODEL = {
    "censo": CensusEntry,
    "eventos": EventSeguridad,
    "duplicidades": DuplicidadEntry,
    "encuestas": EncuestaEntry,
    "atencion": AtencionEntry,
    "robos": RoboHurtoEntry,
    "miscelaneo": MiscelaneoEntry,
    "desviaciones": DesviacionEntry,
    "solicitud_ot": SolicitudOTEntry,
    "reclamos": ReclamoUsuarioEntry,
    # nuevos
    "alarmas": ActivacionAlarmaEntry,
    "extensiones": ExtensionExcepcionEntry,
    "onboarding": OnboardingEntry,
    "apertura": AperturaHabitacionEntry,
    "cumplimiento": CumplimientoEECCEntry,
    "entradas_salidas": EntradaSalidaEntry,
    "habitaciones_bloqueadas": HabitacionBloqueadaEntry,
    "ordenamiento": OrdenamientoEntry,
    "habitaciones_liberadas": HabitacionLiberadaEntry,
}

ENTITY_LIST_KEY = {
    "censo": "census", "eventos": "eventos", "duplicidades": "duplics",
    "encuestas": "encuestas", "atencion": "atenciones", "robos": "robos",
    "miscelaneo": "miscelaneo", "desviaciones": "desviaciones",
    "solicitud_ot": "solicitudes_ot", "reclamos": "reclamos", "alarmas": "alarmas",
    "extensiones": "extensiones", "onboarding": "onboarding", "apertura": "apertura",
    "cumplimiento": "cumplimiento", "entradas_salidas": "entradas_salidas",
    "habitaciones_bloqueadas": "habitaciones_bloqueadas",
    "ordenamiento": "ordenamiento", "habitaciones_liberadas": "habitaciones_liberadas",
}
ENTITY_DATE_FIELD = {
    "encuestas": "fecha_hora", "onboarding": "fecha_hora",
    "miscelaneo": "fecha_creacion", "solicitud_ot": "fecha_inicio",
    "extensiones": "fecha_solicitud",
    "entradas_salidas": "fecha_ingreso",
    "habitaciones_bloqueadas": "fecha_bloqueo",
    "ordenamiento": "fecha_ejecucion",
    "habitaciones_liberadas": "fecha_devolucion",
}


def hotel_records_query(db, entity, d_from=None, d_to=None):
    """Una sola regla de fechas para mostrar y eliminar registros."""
    Model = ENTITY_MODEL[entity]
    column = getattr(Model, ENTITY_DATE_FIELD.get(entity, "fecha"))
    query = db.query(Model)
    if isinstance(column.type, DateTime):
        if d_from:
            query = query.filter(column >= datetime.combine(d_from, time.min))
        if d_to:
            query = query.filter(column <= datetime.combine(d_to, time.max))
    else:
        if d_from:
            query = query.filter(column >= d_from)
        if d_to:
            query = query.filter(column <= d_to)
    return query


def records_return_url(entity, target):
    """Vuelve al listado del módulo con sus filtros, dentro de la aplicación."""
    filters = {}
    list_path = url_for("registros")
    # Admite los formularios anteriores, que enviaban /registros sin el prefijo.
    allowed_paths = {list_path, list_path.removeprefix(request.script_root)}
    try:
        parts = urlsplit(target or "")
        if not parts.scheme and not parts.netloc and parts.path in allowed_paths:
            query = parse_qs(parts.query)
            for name in ("from", "to"):
                if query.get(name):
                    try:
                        filters[name] = date.fromisoformat(query[name][0]).isoformat()
                    except ValueError:
                        pass
            if query.get("semana") and query["semana"][0].isdigit():
                week = int(query["semana"][0])
                if week in WEEK_MAP:
                    filters["semana"] = week
    except ValueError:
        pass
    return url_for("registros", vista=entity, **filters)


def verify_bulk_csrf():
    expected = session.get("hotel_bulk_csrf", "")
    if not expected or not secrets.compare_digest(
        request.form.get("csrf_token", "").encode(), expected.encode()
    ):
        abort(400, description="La sesión del formulario expiró. Vuelve al listado.")


def bulk_delete_filters(form):
    """No transforma filtros inválidos en una eliminación sin restricciones."""
    filters = {}
    try:
        for name in ("from", "to"):
            raw = form.get(name, "").strip()
            if raw:
                filters[name] = date.fromisoformat(raw).isoformat()
        raw_week = form.get("semana", "").strip()
        if raw_week:
            week = int(raw_week)
            if week not in WEEK_MAP:
                raise ValueError("Semana inválida")
            filters["semana"] = str(week)
            start, end = week_range(week)
            filters.update({"from": start.isoformat(), "to": end.isoformat()})
        d_from = date.fromisoformat(filters["from"]) if "from" in filters else None
        d_to = date.fromisoformat(filters["to"]) if "to" in filters else None
        if d_from and d_to and d_to < d_from:
            raise ValueError("Rango inválido")
    except (ValueError, TypeError):
        abort(400, description="Los filtros no son válidos. Vuelve al listado y revisa las fechas.")
    return filters, d_from, d_to


def bulk_records_digest(records):
    digest = hashlib.sha256()
    for record in sorted(records, key=lambda row: row.id):
        digest.update(record_version(record).encode("ascii"))
    return digest.hexdigest()


def records_in_batches(query, Model, ids, lock=False):
    records = []
    for offset in range(0, len(ids), 500):
        batch = query.filter(Model.id.in_(ids[offset:offset + 500])).order_by(Model.id)
        records.extend((batch.with_for_update() if lock else batch).all())
    return records


def bulk_delete_serializer():
    return URLSafeTimedSerializer(app.secret_key, salt="hotel-bulk-delete")


@app.post("/delete/<string:entity>/bulk/confirm")
def confirm_bulk_delete(entity):
    Model = ENTITY_MODEL.get(entity)
    if Model is None:
        abort(404)
    verify_bulk_csrf()
    filters, d_from, d_to = bulk_delete_filters(request.form)
    return_url = url_for("registros", vista=entity, **filters)
    mode = request.form.get("mode")
    if mode not in ("selected", "all"):
        abort(400, description="Selecciona una opción de eliminación.")

    db = SessionLocal()
    try:
        query = hotel_records_query(db, entity, d_from, d_to)
        if mode == "selected":
            raw_ids = request.form.getlist("ids")
            if not raw_ids:
                flash("Selecciona al menos un registro para eliminar.", "warning")
                return redirect(return_url)
            if any(not re.fullmatch(r"[0-9]{1,10}", value) or not 1 <= int(value) <= 2147483647 for value in raw_ids):
                abort(400, description="La selección de registros no es válida.")
            ids = sorted({int(value) for value in raw_ids})
            records = records_in_batches(query, Model, ids)
            if len(records) != len(ids):
                flash("La selección cambió o incluye registros fuera del listado. Vuelve a seleccionarlos.", "warning")
                return redirect(return_url)
        else:
            records = query.order_by(Model.id).all()
            ids = [record.id for record in records]
        if not records:
            flash("No hay registros para eliminar con estos filtros.", "info")
            return redirect(return_url)

        token = bulk_delete_serializer().dumps({
            "entity": entity, "mode": mode, "ids": ids, "filters": filters,
            "digest": bulk_records_digest(records), "csrf": session["hotel_bulk_csrf"],
        })
        return render_template(
            "bulk_delete_confirm.html", entity=entity, entity_title=EDIT_CONFIG[entity][0],
            count=len(ids), mode=mode, d_from=d_from, d_to=d_to,
            return_url=return_url, confirmation_token=token,
            bulk_csrf=session["hotel_bulk_csrf"],
        )
    finally:
        db.close()


@app.post("/delete/<string:entity>/bulk")
def delete_records_bulk(entity):
    Model = ENTITY_MODEL.get(entity)
    if Model is None:
        abort(404)
    verify_bulk_csrf()
    try:
        payload = bulk_delete_serializer().loads(request.form.get("confirmation_token", ""), max_age=900)
    except BadData:
        flash("La confirmación venció o no es válida. Revisa el listado y vuelve a confirmar.", "warning")
        return redirect(url_for("registros", vista=entity))
    if payload.get("entity") != entity or payload.get("csrf") != session["hotel_bulk_csrf"]:
        abort(400, description="La confirmación no corresponde a esta selección.")

    ids = payload["ids"]
    filters, d_from, d_to = bulk_delete_filters(payload["filters"])
    return_url = url_for("registros", vista=entity, **filters)
    db = SessionLocal()
    try:
        records = records_in_batches(hotel_records_query(db, entity, d_from, d_to), Model, ids, lock=True)
        if len(records) != len(ids) or bulk_records_digest(records) != payload["digest"]:
            flash("Los registros cambiaron desde la confirmación. No se eliminó ninguno; revisa el listado.", "warning")
            return redirect(return_url)

        deleted = 0
        for offset in range(0, len(ids), 500):
            deleted += db.query(Model).filter(Model.id.in_(ids[offset:offset + 500])).delete(synchronize_session=False)
        if deleted != len(ids):
            db.rollback()
            flash("La selección cambió. No se eliminó ningún registro; vuelve a intentarlo.", "warning")
        else:
            db.commit()
            flash(f"Se eliminaron {deleted} registros de {EDIT_CONFIG[entity][0]}.", "success")
    except SQLAlchemyError:
        db.rollback()
        app.logger.exception("Error al eliminar varios registros de hotelería")
        flash("No se pudo completar la eliminación. No se guardó ningún borrado; vuelve a intentarlo.", "danger")
    finally:
        db.close()
    return redirect(return_url)


@app.route("/edit/<string:entity>/<int:rid>", methods=["GET", "POST"])
def edit_record(entity, rid):
    Model = ENTITY_MODEL.get(entity)
    if Model is None or entity not in EDIT_CONFIG:
        abort(404)

    csrf_token = session.setdefault("hotel_edit_csrf", secrets.token_hex(32))
    if request.method == "POST" and not secrets.compare_digest(
        request.form.get("csrf_token", "").encode(), csrf_token.encode()
    ):
        abort(400, description="La sesión del formulario expiró. Vuelve a abrir el registro.")

    db = SessionLocal()
    try:
        query = db.query(Model).filter(Model.id == rid)
        # PostgreSQL bloquea la fila durante el guardado para evitar sobrescrituras.
        record = (query.with_for_update() if request.method == "POST" else query).first()
        if record is None:
            abort(404, description="El registro no existe o fue eliminado.")

        fields = edit_fields(entity, record)
        version = record_version(record)
        total_auto = entity == "censo" and record.total == record.censo_dia + record.censo_noche
        return_url = records_return_url(entity, request.form.get("next") if request.method == "POST" else request.args.get("next"))
        errors, status, conflict = {}, 200, False

        if request.method == "POST":
            total_auto = request.form.get("censo_total_auto") == "1"
            if request.form.get("record_version") != version:
                errors["_form"] = "Este registro cambió desde que lo abriste. Carga la versión actual antes de editarlo de nuevo."
                status, conflict = 409, True
            else:
                values, errors = parse_edit_values(entity, fields, request.form)
                if errors:
                    status = 422
                else:
                    try:
                        for name, value in values.items():
                            setattr(record, name, value)
                        db.commit()
                    except SQLAlchemyError:
                        db.rollback()
                        app.logger.exception("Error al editar un registro de hotelería")
                        errors["_form"] = "No se pudieron guardar los cambios. Inténtalo nuevamente."
                        status = 503
                    else:
                        flash("Registro actualizado correctamente.", "success")
                        return redirect(return_url)

            # Conserva lo escrito si hay errores; no sustituye el token de una versión antigua.
            version = request.form.get("record_version", "")
            for field in fields:
                field["value"] = request.form.get(field["name"], field["value"])

        return render_template(
            "edit.html", entity=entity, entity_title=EDIT_CONFIG[entity][0],
            record_id=rid, fields=fields, errors=errors, return_url=return_url,
            csrf_token=csrf_token, version=version, total_auto=total_auto, conflict=conflict,
        ), status
    finally:
        db.close()


@app.post("/delete/<string:entity>/<int:rid>")
def delete_record(entity, rid):
    mapping = ENTITY_MODEL
    Model = mapping.get(entity)
    if not Model:
        flash("Entidad inválida.")
        return redirect(url_for("registros"))

    db = SessionLocal()
    try:
        obj = db.get(Model, rid)
        if not obj:
            flash("Registro no encontrado.")
        else:
            db.delete(obj)
            db.commit()
            flash("Registro eliminado.", "success")
    except Exception as e:
        db.rollback()
        flash(f"No se pudo eliminar: {e}")
    finally:
        db.close()

    return redirect(records_return_url(entity, request.form.get("next")))

# -----------------------------------------------------------------------------
# PLANTILLAS EXCEL + IMPORTACIÓN POR MÓDULO
# -----------------------------------------------------------------------------
TEMPLATES = {
    # existentes
    "censo": ["FECHA", "CENSO_DIA", "CENSO_NOCHE", "TOTAL"],
    "eventos": ["FECHA", "HORARIO", "QUE_OCURRIO", "NOMBRE_AFECTADO", "ACCION"],
    "duplicidades": [
        "SEMANA","FECHA","ID","EMPRESA_CONTRATISTA","DESCRIPCION_PROBLEMA","TIPO_RIESGO",
        "PABELLON","HABITACION","INGRESAR_CONTACTO","NOMBRE_USUARIO","RESPONSABLE","ESTATUS",
        "NOTIFICACION_USUARIO","PLAN_ACCION","FECHA_CIERRE"
    ],
    "encuesta": [
        "FECHA_HORA",
        "Q1_RESPUESTA","Q1_PUNTAJE","Q2_RESPUESTA","Q2_PUNTAJE","Q3_RESPUESTA","Q3_PUNTAJE",
        "Q4_RESPUESTA","Q4_PUNTAJE","Q5_RESPUESTA","Q5_PUNTAJE","TOTAL","PROMEDIO","COMENTARIOS"
    ],
    "atencion": ["FECHA","TIEMPO_PROMEDIO_MMSS","CANTIDAD"],

    # agregados previos
    "robos": [
        "FECHA","HORA","MODULO","HABITACION","EMPRESA","NOMBRE_CLIENTE","RUT",
        "MEDIO_RECLAMO","ESPECIES","OBSERVACIONES","RECEPCIONA"
    ],
    "miscelaneo": [
        "OT","DIVISION","AREA","LUGAR","UBICACION","DISCIPLINA","ESPECIALIDAD","FALLA","EMPRESA",
        "FECHA_CREACION","FECHA_INICIO","FECHA_TERMINO","FECHA_APROBACION","ESTADO","COMENTARIO"
    ],
    "desviaciones": [
        "N_SOLICITUD","FECHA","ID","EMPRESA_CONTRATISTA","DESCRIPCION_PROBLEMA","TIPO_RIESGO",
        "TIPO_SOLICITUD","PABELLON","HABITACION","VIA_SOLICITUD","QUIEN_INFORMA","RIESGO_MATERIAL","CORREO_DESTINO","ACCIONES"
    ],
    "solicitud_ot": [
        "N_SOLICITUD","DESCRIPCION_PROBLEMA","TIPO_SOLICITUD","MODULO","HABITACION","TIPO_TURNO","JORNADA",
        "VIA_SOLICITUD","CORREO_USUARIO","TIPO_TAREA","OT","FECHA_INICIO","ESTADO","TIEMPO_RESPUESTA_MMSS",
        "SATISFACCION_RECLAMO","MOTIVO","OBSERVACION"
    ],
    "reclamos": [
        "N_SOLICITUD","FECHA","ID","EMPRESA_CONTRATISTA","DESCRIPCION_PROBLEMA","TIPO_SOLICITUD","PABELLON",
        "HABITACION","VIA_SOLICITUD","INGRESAR_CONTACTO","NOMBRE_USUARIO","RESPONSABLE","ESTATUS",
        "NOTIFICACION_USUARIO","PLAN_ACCION"
    ],

    # --------- NUEVOS 5 ---------
   "alarmas": [
    "MODULO", "N_HABITACION", "NOMBRE_RECEPCIONISTA", "FECHA", "EMPRESA", "ID", "CO",
    "AVISO_MANTENCION_H", "LLEGADA_MANTENCION_H", "AVISO_LIDER_H", "LLEGADA_LIDER_H",
    "HORA_REPORTE_SALFA", "TIPO_EVENTO", "TIPO_ACTIVIDAD", "FECHA_REPORTE",
    "TURNO_RECEPCION_INGRESOS", "OBSERVACIONES"  # <-- Asegúrate de que OBSERVACIONES esté aquí
    ],
    "extensiones": [label for _, label in EXTENSION_FIELDS],
    "onboarding": ["FECHA_HORA","NOMBRE","RUT","EMPRESA","ID","ARCHIVO_PDF"],
    "apertura": ["FECHA","HABITACION","HORA","RESPONSABLE","ESTADO_CHAPA"],
    "cumplimiento": ["FECHA","EMPRESA","N_CONTRATO","CO","CORREO_ELECTRONICO","ID","TURNO"],
    "entradas_salidas": [label for _, label in ENTRY_EXIT_FIELDS],
    "habitaciones_bloqueadas": [label for _, label in BLOCKED_ROOM_FIELDS],
    "ordenamiento": [label for _, label in ORDERING_FIELDS],
    "habitaciones_liberadas": [label for _, label in RELEASED_ROOM_FIELDS],
}

@app.get("/template/<string:entity>.xlsx")
def template_xlsx(entity):
    entity = entity.lower()
    if entity not in TEMPLATES:
        flash("Entidad no válida para plantilla.")
        return redirect(url_for("panel", tab="censo"))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla"
    
    # Obtener los encabezados exactos para esta entidad
    headers = TEMPLATES[entity]
    
    # **CORRECCIÓN: Asegurar que la columna OBSERVACIONES esté presente para alarmas**
    if entity == "alarmas" and "OBSERVACIONES" not in headers:
        headers.append("OBSERVACIONES")
    
    # Escribir los encabezados
    ws.append(headers)

    if entity in OPTIONAL_RECORD_FIELDS:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            ws.column_dimensions[cell.column_letter].width = max(20, len(cell.value) + 3)
        # Formatos de las primeras filas para introducir fechas, horas e identificadores.
        formats = {
            "entradas_salidas": {1: "DD/MM/YYYY", 2: "DD/MM/YYYY", 3: "HH:MM:SS", 4: "HH:MM:SS",
                                 6: "@", 8: "@", 10: "@", 11: "@", 15: "@"},
            "habitaciones_bloqueadas": {1: "DD/MM/YYYY", 2: "@", 3: "@", 5: "@",
                                       8: "DD/MM/YYYY", 9: "DD/MM/YYYY", 10: "DD/MM/YYYY", 11: "DD/MM/YYYY"},
            "ordenamiento": {1: "DD/MM/YYYY", 3: "@", 5: "@", 7: "@"},
            "habitaciones_liberadas": {1: "@", 4: "DD/MM/YYYY", 6: "DD/MM/YYYY"},
        }[entity]
        for row in range(2, 202):
            for column, number_format in formats.items():
                ws.cell(row, column).number_format = number_format
    
    # **CORRECCIÓN: Para alarmas, agregar una fila de ejemplo con el formato correcto**
    if entity == "alarmas":
        # Fila de ejemplo con datos de muestra
        example_row = [
            "5",                    # MODULO
            "Hall central",         # N_HABITACION  
            "Marta Montenegro",     # NOMBRE_RECEPCIONISTA
            "2025-10-15",           # FECHA (formato YYYY-MM-DD)
            "Empresa Ejemplo",      # EMPRESA
            "GC123456",             # ID
            "CO123",                # CO
            0.5,                    # AVISO_MANTENCION_H (formato decimal Excel)
            0.75,                   # LLEGADA_MANTENCION_H (formato decimal Excel)
            0.5,                    # AVISO_LIDER_H (formato decimal Excel)
            0.75,                   # LLEGADA_LIDER_H (formato decimal Excel)
            "15:30",                # HORA_REPORTE_SALFA (formato HH:MM)
            "Polvo",                # TIPO_EVENTO
            "Silenciar",            # TIPO_ACTIVIDAD
            "2025-10-15",           # FECHA_REPORTE (formato YYYY-MM-DD)
            "DIA",                  # TURNO_RECEPCION_INGRESOS
            "Observación de ejemplo" # OBSERVACIONES
        ]
        ws.append(example_row)
        
    if entity == "cumplimiento":
        ws.append([
            "2025-10-20",     # FECHA (reconoce varios formatos al importar)
            "Empresa X",
            "C-123",
            "CO45",
            "contacto@x.cl",
            "ID987",
            "NOCHE",
        ])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(
        out, 
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, 
        download_name=f"plantilla_{entity}.xlsx"
    )

@app.post("/import/<string:entity>")
def import_xlsx(entity):
    entity = entity.lower()
    if entity not in TEMPLATES:
        flash("Entidad no válida para importación.")
        return redirect(url_for("panel", tab="censo"))

    f = request.files.get("file")
    if not f or f.filename == "":
        flash("Sube un archivo .xlsx.")
        return redirect(url_for("panel", tab=entity if entity != "eventos" else "eventos"))

    try:
        wb = load_workbook(filename=io.BytesIO(f.read()), data_only=True)
        ws = wb.active
        
        # **CORRECCIÓN: Normalizar encabezados para manejar tildes y caracteres especiales**
        headers = [normalize_header(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        expected = [normalize_header(h) for h in TEMPLATES[entity]]
        
        # La plantilla anterior de desviaciones sigue siendo válida; Acciones queda vacío.
        legacy_deviations = entity == "desviaciones" and headers == expected[:-1]
        if headers != expected and not legacy_deviations:
            flash(f"Encabezados inválidos. Esperado: {', '.join(TEMPLATES[entity])}")
            return redirect(url_for("panel", tab=entity if entity != "eventos" else "eventos"))

        inserted = 0
        db = SessionLocal()
        try:
            def to_int(v, default=0):
                try:
                    if v in (None, ""): return default
                    return int(v)
                except:
                    return default

            def to_float(v):
                try:
                    if v in (None, ""): return None
                    return float(v)
                except:
                    return None

            for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if all(cell is None or str(cell).strip()=="" for cell in row):
                    continue

                # ---------------- existentes ----------------
                if entity == "censo":
                    fecha = safe_convert_date(row[0])
                    cd = to_int(row[1]); cn = to_int(row[2])
                    total = to_int(row[3], cd+cn)
                    db.add(CensusEntry(fecha=fecha, censo_dia=cd, censo_noche=cn, total=total))

                elif entity == "eventos":
                    db.add(EventSeguridad(
                        fecha=safe_convert_date(row[0]),
                        horario=str(row[1] or "").strip(),
                        que_ocurrio=str(row[2] or "").strip(),
                        nombre_afectado=str(row[3] or "").strip(),
                        accion=str(row[4] or "").strip()
                    ))

                elif entity == "duplicidades":
                    db.add(DuplicidadEntry(
                        semana=to_int(row[0], 0),
                        fecha=safe_convert_date(row[1]),
                        id_interno=str(row[2] or "").strip(),
                        empresa_contratista=str(row[3] or "").strip(),
                        descripcion_problema=str(row[4] or "").strip(),
                        tipo_riesgo=str(row[5] or "").strip(),
                        pabellon=str(row[6] or "").strip(),
                        habitacion=str(row[7] or "").strip(),
                        ingresar_contacto=str(row[8] or "").strip(),
                        nombre_usuario=str(row[9] or "").strip(),
                        responsable=str(row[10] or "").strip(),
                        estatus=str(row[11] or "").strip(),
                        notificacion_usuario=str(row[12] or "").strip(),
                        plan_accion=str(row[13] or "").strip(),
                        fecha_cierre=safe_convert_date(row[14]),
                    ))

                elif entity == "encuesta":
                    fh_raw = safe_convert_datetime(row[0])
                    def t_int(v):
                        try: return int(v)
                        except: return None
                    db.add(EncuestaEntry(
                        fecha_hora=fh_raw,
                        q1_respuesta=str(row[1] or ""), q1_puntaje=t_int(row[2]),
                        q2_respuesta=str(row[3] or ""), q2_puntaje=t_int(row[4]),
                        q3_respuesta=str(row[5] or ""), q3_puntaje=t_int(row[6]),
                        q4_respuesta=str(row[7] or ""), q4_puntaje=t_int(row[8]),
                        q5_respuesta=str(row[9] or ""), q5_puntaje=t_int(row[10]),
                        total=t_int(row[11]),
                        promedio=to_float(row[12]),
                        comentarios=str(row[13] or "")
                    ))

                elif entity == "atencion":
                    fecha = safe_convert_date(row[0])
                    
                    # Manejar el tiempo de diferentes formatos
                    tiempo_val = row[1]
                    segundos = safe_convert_time(tiempo_val)
                    
                    cant = to_int(row[2])
                    db.add(AtencionEntry(fecha=fecha, tiempo_promedio_sec=segundos, cantidad=cant))

                # ---------------- agregados previos ----------------
                elif entity == "robos":
                    # Usamos la función safe_time_hhmm robusta
                    hora_str = str(row[1] or "").strip() if row[1] is not None else ""
                    hora_obj = safe_time_hhmm(hora_str)
                    
                    db.add(RoboHurtoEntry(
                        fecha=safe_convert_date(row[0]),
                        hora=hora_obj,
                        modulo=str(row[2] or "").strip(),
                        habitacion=str(row[3] or "").strip(),
                        empresa=str(row[4] or "").strip(),
                        nombre_cliente=str(row[5] or "").strip(),
                        rut=str(row[6] or "").strip(),
                        medio_reclamo=str(row[7] or "").strip(),
                        especies=str(row[8] or "").strip(),
                        observaciones=str(row[9] or "").strip(),
                        recepciona=str(row[10] or "").strip(),
                    ))

                elif entity == "miscelaneo":
                    db.add(MiscelaneoEntry(
                        ot=str(row[0] or "").strip(),
                        division=str(row[1] or "").strip(),
                        area=str(row[2] or "").strip(),
                        lugar=str(row[3] or "").strip(),
                        ubicacion=str(row[4] or "").strip(),
                        disciplina=str(row[5] or "").strip(),
                        especialidad=str(row[6] or "").strip(),
                        falla=str(row[7] or "").strip(),
                        empresa=str(row[8] or "").strip(),
                        fecha_creacion=safe_convert_date(row[9]),
                        fecha_inicio=safe_convert_date(row[10]),
                        fecha_termino=safe_convert_date(row[11]),
                        fecha_aprobacion=safe_convert_date(row[12]),
                        estado=str(row[13] or "").strip(),
                        comentario=str(row[14] or "").strip(),
                    ))

                elif entity == "desviaciones":
                    db.add(DesviacionEntry(
                        n_solicitud=str(row[0] or "").strip(),
                        fecha=safe_convert_date(row[1]),
                        id_interno=str(row[2] or "").strip(),
                        empresa_contratista=str(row[3] or "").strip(),
                        descripcion_problema=str(row[4] or "").strip(),
                        tipo_riesgo=str(row[5] or "").strip(),
                        tipo_solicitud=str(row[6] or "").strip(),
                        pabellon=str(row[7] or "").strip(),
                        habitacion=str(row[8] or "").strip(),
                        via_solicitud=str(row[9] or "").strip(),
                        quien_informa=str(row[10] or "").strip(),
                        riesgo_material=str(row[11] or "").strip(),
                        correo_destino=str(row[12] or "").strip(),
                        acciones=str(row[13] if len(row) > 13 and row[13] is not None else "").strip(),
                    ))

                elif entity == "solicitud_ot":
                    tiempo_val = row[13]
                    secs = safe_convert_time(tiempo_val) if tiempo_val not in (None, "") else None
                    
                    db.add(SolicitudOTEntry(
                        n_solicitud=str(row[0] or "").strip(),
                        descripcion_problema=str(row[1] or "").strip(),
                        tipo_solicitud=str(row[2] or "").strip(),
                        modulo=str(row[3] or "").strip(),
                        habitacion=str(row[4] or "").strip(),
                        tipo_turno=str(row[5] or "").strip(),
                        jornada=str(row[6] or "").strip(),
                        via_solicitud=str(row[7] or "").strip(),
                        correo_usuario=str(row[8] or "").strip(),
                        tipo_tarea=str(row[9] or "").strip(),
                        ot=str(row[10] or "").strip(),
                        fecha_inicio=safe_convert_date(row[11]),
                        estado=str(row[12] or "").strip(),
                        tiempo_respuesta_sec=secs,
                        satisfaccion_reclamo=str(row[14] or "").strip(),
                        motivo=str(row[15] or "").strip(),
                        observacion=str(row[16] or "").strip(),
                    ))

                elif entity == "reclamos":
                    db.add(ReclamoUsuarioEntry(
                        n_solicitud=str(row[0] or "").strip(),
                        fecha=safe_convert_date(row[1]),
                        id_interno=str(row[2] or "").strip(),
                        empresa_contratista=str(row[3] or "").strip(),
                        descripcion_problema=str(row[4] or "").strip(),
                        tipo_solicitud=str(row[5] or "").strip(),
                        pabellon=str(row[6] or "").strip(),
                        habitacion=str(row[7] or "").strip(),
                        via_solicitud=str(row[8] or "").strip(),
                        ingresar_contacto=str(row[9] or "").strip(),
                        nombre_usuario=str(row[10] or "").strip(),
                        responsable=str(row[11] or "").strip(),
                        estatus=str(row[12] or "").strip(),
                        notificacion_usuario=str(row[13] or "").strip(),
                        plan_accion=str(row[14] or "").strip(),
                    ))

                # ---------------- NUEVOS 5 ----------------
                elif entity == "alarmas":
                    def ffloat(v):
                        try:
                            return float(v) if (v not in (None, "") ) else None
                        except:
                            return None
                    db.add(ActivacionAlarmaEntry(
                        modulo=str(row[0] or "").strip(),
                        n_habitacion=str(row[1] or "").strip(),
                        nombre_recepcionista=str(row[2] or "").strip(),
                        fecha=safe_convert_date(row[3]),
                        empresa=str(row[4] or "").strip(),
                        id_interno=str(row[5] or "").strip(),
                        co=str(row[6] or "").strip(),
                        aviso_mantencion_h=ffloat(row[7]),
                        llegada_mantencion_h=ffloat(row[8]),
                        aviso_lider_h=ffloat(row[9]),
                        llegada_lider_h=ffloat(row[10]),
                        hora_reporte_salfa=safe_time_hhmm(row[11]),
                        tipo_evento=str(row[12] or "").strip(),
                        tipo_actividad=str(row[13] or "").strip(),
                        fecha_reporte=safe_convert_date(row[14]),
                        turno_recepcion_ingresos=str(row[15] or "").strip(),
                        observaciones=str(row[16] or "").strip(),
                    ))

                elif entity == "extensiones":
                    # Los nombres mantienen alineados los datos si cambia el orden de la plantilla.
                    values = dict(zip((name for name, _ in EXTENSION_FIELDS), row))
                    for name in values:
                        if name in ("fecha_solicitud", "desde", "hasta"):
                            values[name] = safe_convert_date(values[name])
                        elif name == "cant_clientes":
                            values[name] = to_int(values[name], None)
                        else:
                            values[name] = str(values[name]).strip() if values[name] is not None else ""
                    db.add(ExtensionExcepcionEntry(
                        **values,
                    ))

                elif entity == "onboarding":
                    fh_raw = safe_convert_datetime(row[0])
                    db.add(OnboardingEntry(
                        fecha_hora=fh_raw,
                        nombre=str(row[1] or "").strip(),
                        rut=str(row[2] or "").strip(),
                        empresa=str(row[3] or "").strip(),
                        id_interno=str(row[4] or "").strip(),
                        archivo_pdf=str(row[5] or "").strip(),
                    ))

                elif entity == "apertura":
                    db.add(AperturaHabitacionEntry(
                        fecha=safe_convert_date(row[0]),
                        habitacion=str(row[1] or "").strip(),
                        hora=safe_time_hhmm(row[2]),
                        responsable=str(row[3] or "").strip(),
                        estado_chapa=str(row[4] or "").strip(),
                    ))

                elif entity == "cumplimiento":
                    db.add(CumplimientoEECCEntry(
                        fecha=safe_convert_date(row[0]),                    # <-- NUEVO
                        empresa=str(row[1] or "").strip(),
                        n_contrato=str(row[2] or "").strip(),
                        co=str(row[3] or "").strip(),
                        correo_electronico=str(row[4] or "").strip(),
                        id_interno=str(row[5] or "").strip(),
                        turno=str(row[6] or "").strip(),
                    ))

                elif entity in OPTIONAL_RECORD_FIELDS:
                    try:
                        Model = ENTITY_MODEL[entity]
                        fields = edit_fields(entity, Model())
                        form = record_excel_values(row, wb.epoch, fields)
                        values, errors = parse_edit_values(entity, fields, form)
                        if errors:
                            labels = dict(OPTIONAL_RECORD_FIELDS[entity])
                            raise ValueError("; ".join(f"{labels.get(name, 'Registro')}: {error}" for name, error in errors.items()))
                        db.add(Model(**values))
                    except (ValueError, TypeError, OverflowError) as exc:
                        raise ValueError(f"Fila {row_number}: {exc}") from exc

                inserted += 1

            db.commit()
            flash(f"Importación de {entity} OK: {inserted} filas.")
        except Exception as e:
            db.rollback()
            flash(f"Error importando {entity}: {e}")
        finally:
            db.close()

    except Exception as e:
        flash(f"No se pudo leer el Excel: {e}")

    tab = "eventos" if entity == "eventos" else entity
    return redirect(url_for("panel", tab=tab))


# -----------------------------------------------------------------------------
# DASHBOARD (actualizado para mostrar estadísticas de todos los módulos)
# -----------------------------------------------------------------------------
@app.get("/dashboard")
def dashboard():
    d_from, d_to, semana_sel = resolve_filters(request.args)
    if semana_sel: d_from, d_to = week_range(semana_sel)
    db = SessionLocal()
    try:
        per_day = {}
        def bucket(dkey):
            return per_day.setdefault(dkey, {
                "censo": 0,
                "eventos": 0,
                "duplicidades": 0,
                "encuestas": 0,
                "atencion_cant": 0,
                "atencion_tiempos": [],
                "robos": 0,
                "miscelaneo": 0,
                "desviaciones": 0,
                "solicitudes_ot": 0,
                "reclamos": 0,
                "alarmas": 0,
                "extensiones": 0,
                "onboarding": 0,
                "apertura": 0,
                "cumplimiento": 0,
                "entradas": 0,
                "habitaciones_bloqueadas": 0,
                "ordenamiento": 0,
                "habitaciones_liberadas": 0,
            })

        # Censo
        q = db.query(CensusEntry)
        if d_from: q = q.filter(CensusEntry.fecha >= d_from)
        if d_to:   q = q.filter(CensusEntry.fecha <= d_to)
        for r in q.all():
            b = bucket(r.fecha.isoformat())
            b["censo"] += (r.total or (r.censo_dia + r.censo_noche))

        # Eventos
        q = db.query(EventSeguridad)
        if d_from: q = q.filter(EventSeguridad.fecha >= d_from)
        if d_to:   q = q.filter(EventSeguridad.fecha <= d_to)
        for r in q.all():
            bucket(r.fecha.isoformat())["eventos"] += 1

        # Duplicidades
        q = db.query(DuplicidadEntry)
        if d_from: q = q.filter(DuplicidadEntry.fecha >= d_from)
        if d_to:   q = q.filter(DuplicidadEntry.fecha <= d_to)
        for r in q.all():
            bucket(r.fecha.isoformat())["duplicidades"] += 1

        # Encuestas
        q = db.query(EncuestaEntry)
        if d_from: q = q.filter(EncuestaEntry.fecha_hora >= datetime.combine(d_from, time.min))
        if d_to:   q = q.filter(EncuestaEntry.fecha_hora <= datetime.combine(d_to, time.max))
        for r in q.all():
            bucket(r.fecha_hora.date().isoformat())["encuestas"] += 1

        # Atención
        q = db.query(AtencionEntry)
        if d_from: q = q.filter(AtencionEntry.fecha >= d_from)
        if d_to:   q = q.filter(AtencionEntry.fecha <= d_to)
        for r in q.all():
            b = bucket(r.fecha.isoformat())
            b["atencion_cant"] += r.cantidad
            b["atencion_tiempos"].append(r.tiempo_promedio_sec)

        # Robos
        q = db.query(RoboHurtoEntry)
        if d_from: q = q.filter(RoboHurtoEntry.fecha >= d_from)
        if d_to:   q = q.filter(RoboHurtoEntry.fecha <= d_to)
        for r in q.all():
            bucket(r.fecha.isoformat())["robos"] += 1

        # Miscelaneo (filtrar por fecha de negocio)
        q = db.query(MiscelaneoEntry)
        if d_from: q = q.filter(MiscelaneoEntry.fecha_creacion >= d_from)
        if d_to:   q = q.filter(MiscelaneoEntry.fecha_creacion <= d_to)
        for r in q.all():
            # Agrupar por la fecha de negocio; si viene nula, caer al timestamp de creación
            key_date = r.fecha_creacion or r.creado.date()
            bucket(key_date.isoformat())["miscelaneo"] += 1


        # Desviaciones
        q = db.query(DesviacionEntry)
        if d_from: q = q.filter(DesviacionEntry.fecha >= d_from)
        if d_to:   q = q.filter(DesviacionEntry.fecha <= d_to)
        for r in q.all():
            bucket(r.fecha.isoformat())["desviaciones"] += 1

        # Solicitudes OT
        # Solicitudes OT (filtrar por fecha de negocio)
        q = db.query(SolicitudOTEntry)
        if d_from: q = q.filter(SolicitudOTEntry.fecha_inicio >= d_from)
        if d_to:   q = q.filter(SolicitudOTEntry.fecha_inicio <= d_to)
        for r in q.all():
            # Agrupar por la fecha de negocio; si viene nula, caer al timestamp de creación
            key_date = r.fecha_inicio or r.creado.date()
            bucket(key_date.isoformat())["solicitudes_ot"] += 1


        # Reclamos
        q = db.query(ReclamoUsuarioEntry)
        if d_from: q = q.filter(ReclamoUsuarioEntry.fecha >= d_from)
        if d_to:   q = q.filter(ReclamoUsuarioEntry.fecha <= d_to)
        for r in q.all():
            bucket(r.fecha.isoformat())["reclamos"] += 1

        # Alarmas
        q = db.query(ActivacionAlarmaEntry)
        if d_from: q = q.filter(ActivacionAlarmaEntry.fecha >= d_from)
        if d_to:   q = q.filter(ActivacionAlarmaEntry.fecha <= d_to)
        for r in q.all():
            bucket(r.fecha.isoformat())["alarmas"] += 1

        # Extensiones
        q = db.query(ExtensionExcepcionEntry)
        if d_from: q = q.filter(ExtensionExcepcionEntry.fecha_solicitud >= d_from)
        if d_to:   q = q.filter(ExtensionExcepcionEntry.fecha_solicitud <= d_to)
        for r in q.all():
            bucket(r.fecha_solicitud.isoformat())["extensiones"] += 1

        # Onboarding
        q = db.query(OnboardingEntry)
        if d_from: q = q.filter(OnboardingEntry.fecha_hora >= datetime.combine(d_from, time.min))
        if d_to:   q = q.filter(OnboardingEntry.fecha_hora <= datetime.combine(d_to, time.max))
        for r in q.all():
            bucket(r.fecha_hora.date().isoformat())["onboarding"] += 1

        # Apertura
        q = db.query(AperturaHabitacionEntry)
        if d_from: q = q.filter(AperturaHabitacionEntry.fecha >= d_from)
        if d_to:   q = q.filter(AperturaHabitacionEntry.fecha <= d_to)
        for r in q.all():
            bucket(r.fecha.isoformat())["apertura"] += 1

        # Cumplimiento
        q = db.query(CumplimientoEECCEntry)
        if d_from: q = q.filter(CumplimientoEECCEntry.fecha >= d_from)
        if d_to:   q = q.filter(CumplimientoEECCEntry.fecha <= d_to)
        for r in q.all():
            bucket(r.fecha.isoformat())["cumplimiento"] += 1

        # Cada módulo se cuenta por su fecha de negocio, sin asignar fechas a los vacíos.
        for Model, column, key in (
            (EntradaSalidaEntry, EntradaSalidaEntry.fecha_ingreso, "entradas"),
            (HabitacionBloqueadaEntry, HabitacionBloqueadaEntry.fecha_bloqueo, "habitaciones_bloqueadas"),
            (OrdenamientoEntry, OrdenamientoEntry.fecha_ejecucion, "ordenamiento"),
            (HabitacionLiberadaEntry, HabitacionLiberadaEntry.fecha_devolucion, "habitaciones_liberadas"),
        ):
            q = db.query(column, func.count(Model.id)).filter(column.isnot(None))
            if d_from: q = q.filter(column >= d_from)
            if d_to: q = q.filter(column <= d_to)
            for day, count in q.group_by(column).all():
                bucket(day.isoformat())[key] = count

        if not per_day:
            return render_template("dashboard.html", have_data=False, week_map=WEEK_MAP,
                                   d_from=d_from, d_to=d_to, semana_sel=semana_sel, current_tab=None)

        ordered_days = sorted(per_day.keys())
        series_data = {
            "censo": [],
            "eventos": [],
            "duplicidades": [],
            "encuestas": [],
            "atencion_cant": [],
            "atencion_min": [],
            "robos": [],
            "miscelaneo": [],
            "desviaciones": [],
            "solicitudes_ot": [],
            "reclamos": [],
            "alarmas": [],
            "extensiones": [],
            "onboarding": [],
            "apertura": [],
            "cumplimiento": [],
            "entradas": [],
            "habitaciones_bloqueadas": [],
            "ordenamiento": [],
            "habitaciones_liberadas": [],
        }

        for k in ordered_days:
            g = per_day[k]
            series_data["censo"].append(g["censo"])
            series_data["eventos"].append(g["eventos"])
            series_data["duplicidades"].append(g["duplicidades"])
            series_data["encuestas"].append(g["encuestas"])
            series_data["atencion_cant"].append(g["atencion_cant"])
            series_data["robos"].append(g["robos"])
            series_data["miscelaneo"].append(g["miscelaneo"])
            series_data["desviaciones"].append(g["desviaciones"])
            series_data["solicitudes_ot"].append(g["solicitudes_ot"])
            series_data["reclamos"].append(g["reclamos"])
            series_data["alarmas"].append(g["alarmas"])
            series_data["extensiones"].append(g["extensiones"])
            series_data["onboarding"].append(g["onboarding"])
            series_data["apertura"].append(g["apertura"])
            series_data["cumplimiento"].append(g["cumplimiento"])
            series_data["entradas"].append(g["entradas"])
            series_data["habitaciones_bloqueadas"].append(g["habitaciones_bloqueadas"])
            series_data["ordenamiento"].append(g["ordenamiento"])
            series_data["habitaciones_liberadas"].append(g["habitaciones_liberadas"])
            
            prom_s = int(mean(g["atencion_tiempos"])) if g["atencion_tiempos"] else 0
            series_data["atencion_min"].append(round(prom_s/60.0, 2))

        # Calcular totales para las tarjetas
        cards = {
            "censo_total": sum(series_data["censo"]),
            "eventos_total": sum(series_data["eventos"]),
            "duplicidades_total": sum(series_data["duplicidades"]),
            "encuestas_total": sum(series_data["encuestas"]),
            "atencion_cant_total": sum(series_data["atencion_cant"]),
            "robos_total": sum(series_data["robos"]),
            "miscelaneo_total": sum(series_data["miscelaneo"]),
            "desviaciones_total": sum(series_data["desviaciones"]),
            "solicitudes_ot_total": sum(series_data["solicitudes_ot"]),
            "reclamos_total": sum(series_data["reclamos"]),
            "alarmas_total": sum(series_data["alarmas"]),
            "extensiones_total": sum(series_data["extensiones"]),
            "onboarding_total": sum(series_data["onboarding"]),
            "apertura_total": sum(series_data["apertura"]),
            "cumplimiento_total": sum(series_data["cumplimiento"]),
            "entradas_total": sum(series_data["entradas"]),
            "habitaciones_bloqueadas_total": sum(series_data["habitaciones_bloqueadas"]),
            "ordenamiento_total": sum(series_data["ordenamiento"]),
            "habitaciones_liberadas_total": sum(series_data["habitaciones_liberadas"]),
            "atencion_tiempo_prom_global": (
                seconds_to_mmss(int(mean([int(x*60) for x in series_data["atencion_min"] if x>0])))
                if any(x>0 for x in series_data["atencion_min"]) else "00:00"
            ),
        }

        return render_template("dashboard.html",
                               have_data=True,
                               week_map=WEEK_MAP,
                               labels=ordered_days,
                               series=series_data,
                               cards=cards,
                               d_from=d_from, d_to=d_to, semana_sel=semana_sel,
                               current_tab=None)
    finally:
        db.close()


# -----------------------------------------------------------------------------
# MAIN
# -----------------------------------------------------------------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
